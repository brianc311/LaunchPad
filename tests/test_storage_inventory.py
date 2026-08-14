from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from launchpad.health_alert_state import (
    empty_state,
    grandfather_fingerprints,
    issue_fingerprint_for_issue,
    set_limit_new_issues,
)
from launchpad.storage_inventory import (
    BLANK_SITE_LABEL,
    StorageInventoryProgress,
    build_inventory_row,
    build_issues_notes,
    empty_storage_inventory_progress,
    export_storage_inventory_xlsx,
    format_phone_home_cell,
    format_smtp_cell,
    group_inventory_rows_by_site,
    inventory_commands_for_profile,
    inventory_totals,
    is_storage_inventory_eligible,
    parse_hpe_showrcopy_protection,
    parse_svc_lsemailserver,
    parse_svc_lsrcrelationship,
    parse_svc_lssystem_identity,
    row_has_issues,
    site_status,
    split_inventory_issue_fields,
)


def test_storage_inventory_eligible_without_monitor():
    flash = {"card_type": "ssh", "device_profile": "flashsystem_7200"}
    par = {"card_type": "ssh", "device_profile": "hpe_3par_8400"}
    assert is_storage_inventory_eligible(flash)
    assert is_storage_inventory_eligible(par)
    assert not is_storage_inventory_eligible(
        {"card_type": "ssh", "device_profile": "vultr_vps"}
    )
    assert not is_storage_inventory_eligible(
        {"card_type": "rdp", "device_profile": "flashsystem_7200"}
    )

def test_inventory_commands_svc_includes_smtp_and_rcrelationship():
    cmds = inventory_commands_for_profile("flashsystem_7200")
    assert cmds["smtp"] == ["lsemailserver -delim :"]
    assert cmds["data_protection"] == ["lsrcrelationship -delim :"]
    assert cmds["call_home"] == ["lscloudcallhome -delim :"]


def test_inventory_commands_hpe_smtp_empty_call_home_empty():
    cmds = inventory_commands_for_profile("hpe_3par_8400")
    assert cmds["smtp"] == []
    assert cmds["call_home"] == []
    assert cmds["data_protection"] == ["showrcopy"]


def test_parse_svc_identity_and_smtp_and_rcrelationship():
    model, serial = parse_svc_lssystem_identity(
        "id:78E31NF\nname:v7kand-g3v1\nproduct_name:IBM FlashSystem 7200\n"
    )
    assert model == "IBM FlashSystem 7200"
    assert serial == "78E31NF"
    cfg, status, details = parse_svc_lsemailserver(
        "id:name:IP_address:port\n0:smtp1:172.29.62.98:25\n"
    )
    assert cfg == "yes"
    assert "172.29.62.98" in details
    cfg2, _, details2 = parse_svc_lsemailserver("id:name:IP_address:port\n")
    assert cfg2 == "no"
    assert "Not configured" in details2
    yes_cfg, _, _ = parse_svc_lsrcrelationship(
        "id:name:master_cluster_id:master_cluster_name\n0:rel1:1:clusterA\n"
    )
    assert yes_cfg == "yes"
    no_cfg, _, _ = parse_svc_lsrcrelationship(
        "id:name:master_cluster_id:master_cluster_name\n"
    )
    assert no_cfg == "no"


def test_parse_hpe_showrcopy_protection_empty_is_unknown():
    cfg, _status, details = parse_hpe_showrcopy_protection("")
    assert cfg == "unknown"
    assert "empty" in details


def test_parse_hpe_showrcopy_protection_not_configured():
    cfg, _status, details = parse_hpe_showrcopy_protection(
        "Remote Copy is not configured on this system.\n"
    )
    assert cfg == "no"
    assert "not configured" in details


def test_parse_hpe_showrcopy_protection_configured_with_targets():
    cfg, _status, details = parse_hpe_showrcopy_protection(
        "Group: RCG1\nTarget: sync1\nStatus: Started\n"
    )
    assert cfg == "yes"
    assert "RCG1" in details
    assert "sync1" in details


def test_parse_hpe_showrcopy_protection_unrecognized_is_unknown():
    cfg, _status, details = parse_hpe_showrcopy_protection(
        "Line one of noise\nLine two of noise\nLine three of noise\n"
    )
    assert cfg == "unknown"
    assert "unrecognized" in details


def test_issues_notes_and_totals():
    notes = build_issues_notes(
        phone_configured="no",
        data_protection_configured="no",
        smtp_configured="no",
        dns_configured="yes",
        ntp_configured="no",
        health_issues=[{"message": "Running at 91.0% capacity"}],
        extra_errors=[],
    )
    assert "Phone Home not configured" in notes
    assert "Data Protection not configured" in notes
    assert "SMTP not configured" in notes
    assert "NTP not configured" in notes
    assert "Running at 91.0% capacity" in notes
    assert format_phone_home_cell(configured="no", details="", vendor="IBM") == (
        "No — Not configured"
    )
    assert format_smtp_cell(configured="yes", details="172.29.62.98") == "172.29.62.98"
    rows = [
        {"issues": ""},
        {"issues": notes},
    ]
    assert row_has_issues(rows[1]) is True
    assert inventory_totals(rows) == {"total_devices": 2, "devices_with_issues": 1}


def test_export_xlsx_sheets_meta_and_red_issue_row():
    rows = [
        {
            "site": "SiteA",
            "host": "array1",
            "ip": "10.0.0.1",
            "model": "IBM FlashSystem 7200",
            "serial": "ABC",
            "location": "SiteA",
            "phone_home": "Yes — IBM",
            "data_protection": "Yes",
            "smtp": "10.1.1.1",
            "issues": "",
        },
        {
            "site": "SiteB",
            "host": "array2",
            "ip": "10.0.0.2",
            "model": "IBM FlashSystem 7200",
            "serial": "DEF",
            "location": "SiteB",
            "phone_home": "No — Not configured",
            "data_protection": "No — Not configured",
            "smtp": "No IP — Not configured",
            "issues": "Phone Home not configured; SMTP not configured",
        },
    ]
    wb = load_workbook(BytesIO(export_storage_inventory_xlsx(rows, generated_at="2026-08-10T12:00:00")))
    assert wb.sheetnames == ["Inventory", "Issues Summary"]
    inv = wb["Inventory"]
    assert "Total Devices: 2" in str(inv["A1"].value)
    assert "Devices with Issues: 1" in str(inv["A1"].value)
    # Find issue data row by host array2 and assert red-ish fill
    found = False
    for row in inv.iter_rows(min_row=2, max_row=inv.max_row):
        vals = [c.value for c in row]
        if "array2" in vals:
            found = True
            assert row[0].fill.fgColor.rgb in ("00FFCDD2", "FFCDD2")
    assert found
    summary = wb["Issues Summary"]
    assert summary.max_row == 2  # header + one issue


def test_group_inventory_rows_by_site_sorts_and_blank_label():
    rows = [
        {"site": "zeta", "host": "z"},
        {"site": "", "host": "orphan"},
        {"site": "Alpha", "host": "a"},
        {"site": "  ", "host": "also-orphan"},
    ]
    grouped = group_inventory_rows_by_site(rows)
    assert [name for name, _ in grouped] == [BLANK_SITE_LABEL, "Alpha", "zeta"]
    assert BLANK_SITE_LABEL == "(no site)"
    blank_hosts = [row["host"] for row in dict(grouped)[BLANK_SITE_LABEL]]
    assert blank_hosts == ["orphan", "also-orphan"]
    assert group_inventory_rows_by_site([]) == []
    assert group_inventory_rows_by_site(None) == []


def test_site_status_red_orange_green_and_na_ignored():
    assert site_status([{"issues": "SMTP not configured", "phone_home": "unknown"}]) == "red"
    assert site_status([
        {"issues": "", "phone_home": "unknown", "data_protection": "Yes", "smtp": "10.0.0.1"},
    ]) == "orange"
    assert site_status([
        {"issues": "", "phone_home": "n/a", "data_protection": "Yes", "smtp": "10.0.0.1"},
    ]) == "green"
    assert site_status([
        {"issues": "", "phone_home": "Yes — IBM", "data_protection": "Unknown", "smtp": "n/a"},
    ]) == "orange"
    assert site_status([]) == "green"


def _health(message: str, category: str = "drive") -> dict:
    return {"severity": "critical", "category": category, "message": message}


def test_split_config_only_is_recent():
    fields = split_inventory_issue_fields(
        phone_configured="no",
        data_protection_configured="yes",
        smtp_configured="yes",
        dns_configured="yes",
        ntp_configured="yes",
        health_issues=[],
        extra_errors=["scan failed"],
        card_id=7,
        alert_state=empty_state(),
    )
    assert "Phone Home not configured" in fields["issues"]
    assert "scan failed" in fields["issues"]
    assert fields["issues_recent"] == fields["issues"]
    assert fields["issues_older"] == ""


def test_split_leftover_health_is_older():
    issue = _health("Drive 0 is offline")
    state = empty_state()
    fp = issue_fingerprint_for_issue(7, issue)
    state = grandfather_fingerprints(state, [fp])
    fields = split_inventory_issue_fields(
        phone_configured="yes",
        data_protection_configured="yes",
        smtp_configured="yes",
        dns_configured="yes",
        ntp_configured="yes",
        health_issues=[issue],
        extra_errors=[],
        card_id=7,
        alert_state=state,
        now=datetime(2026, 8, 14, 12, 0, 0).timestamp(),
    )
    assert "Drive 0 is offline" in fields["issues"]
    assert fields["issues_recent"] == ""
    assert "Drive 0 is offline" in fields["issues_older"]


def test_split_mixed_and_limit_off():
    leftover = _health("old leftover")
    fresh = _health("new alert", category="capacity")
    state = empty_state()
    state = grandfather_fingerprints(
        state, [issue_fingerprint_for_issue(3, leftover)]
    )
    mixed = split_inventory_issue_fields(
        phone_configured="no",
        data_protection_configured="yes",
        smtp_configured="yes",
        dns_configured="yes",
        ntp_configured="yes",
        health_issues=[leftover, fresh],
        extra_errors=[],
        card_id=3,
        alert_state=state,
    )
    assert "Phone Home not configured" in mixed["issues_recent"]
    assert "new alert" in mixed["issues_recent"]
    assert "old leftover" not in mixed["issues_recent"]
    assert "old leftover" in mixed["issues_older"]
    off = set_limit_new_issues(state, False)
    unlimited = split_inventory_issue_fields(
        phone_configured="yes",
        data_protection_configured="yes",
        smtp_configured="yes",
        dns_configured="yes",
        ntp_configured="yes",
        health_issues=[leftover],
        extra_errors=[],
        card_id=3,
        alert_state=off,
    )
    assert "old leftover" in unlimited["issues_recent"]
    assert unlimited["issues_older"] == ""


def test_build_inventory_row_includes_split_fields():
    unknown = ("yes", "", "")
    row = build_inventory_row(
        site="A",
        host="h",
        ip="10.0.0.1",
        model="m",
        serial="s",
        location="A",
        vendor="IBM",
        profile="flashsystem_7200",
        card_id=1,
        phone=("no", "", ""),
        data_protection=unknown,
        smtp=unknown,
        dns=unknown,
        ntp=unknown,
        health_issues=[],
        extra_errors=[],
    )
    assert "issues_recent" in row
    assert "issues_older" in row
    assert row["issues_recent"] == row["issues"]


def test_storage_inventory_progress_counts_cards():
    assert empty_storage_inventory_progress() == {
        "running": False,
        "done": 0,
        "total": 0,
        "current": "",
    }
    prog = StorageInventoryProgress()
    prog.begin(2)
    snap = prog.snapshot()
    assert snap["running"] is True
    assert snap["total"] == 2
    assert snap["done"] == 0
    prog.start_card("Anderson")
    assert prog.snapshot()["current"] == "Anderson"
    prog.finish_card()
    prog.start_card("Windsor")
    prog.finish_card()
    assert prog.snapshot()["done"] == 2
    prog.end()
    done = prog.snapshot()
    assert done["running"] is False
    assert done["done"] == 2
