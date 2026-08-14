from io import BytesIO

from openpyxl import load_workbook

from launchpad.health_server import HealthCard, HealthServer
from launchpad.storage_inventory_page import STORAGE_INVENTORY_PATH


def _unlock(server: HealthServer) -> None:
    server.set_settings_backend(lambda _key, default: default, lambda _key, _value: None)


def test_storage_inventory_page_route_constant():
    assert STORAGE_INVENTORY_PATH == "/storage-inventory"


def test_scan_storage_inventory_requires_unlock(monkeypatch):
    server = HealthServer()
    monkeypatch.setattr(server, "is_unlocked", lambda: False)
    try:
        server.scan_storage_inventory_live()
        assert False, "expected RuntimeError"
    except RuntimeError as exc:
        assert "unlock" in str(exc).lower()


def test_scan_storage_inventory_svc_happy_path(monkeypatch):
    server = HealthServer()
    _unlock(server)
    card = HealthCard(
        card_id=1,
        name="Hartford",
        host="10.0.0.1",
        port=22,
        username="u",
        key_path="/tmp/key",
        device_profile="flashsystem_7200",
    )
    server._cards[1] = card
    server.set_monitor_enabled(card_id=1, enabled=True)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)

    def _runner(_card):
        def run(command):
            if "lscloudcallhome" in command:
                return "id:status\n0:enabled\n"
            if "lsdnsserver" in command:
                return "id:name:IP_address\n0:dns1:10.1.1.1\n"
            if "lsemailserver" in command:
                return "id:name:IP_address:port\n0:smtp1:172.29.62.98:25\n"
            if "lsrcrelationship" in command:
                return "id:name:master_cluster_id\n0:rel1:1\n"
            if "lssystem" in command:
                return (
                    "id:78E37V9\nname:v7kcon-g3v1\n"
                    "product_name:IBM FlashSystem 7200\n"
                    "cluster_ntp_IP_address:10.3.3.3\n"
                )
            return ""

        return run

    monkeypatch.setattr(server, "_lun_run_command", _runner)
    result = server.scan_storage_inventory_live()
    assert result["errors"] == []
    assert len(result["rows"]) == 1
    row = result["rows"][0]
    assert row["ip"] == "10.0.0.1"
    assert "7200" in row["model"]
    assert row["serial"] == "78E37V9"
    assert "172.29.62.98" in row["smtp"]
    assert row["data_protection"].lower().startswith("yes")
    cached = server.get_storage_inventory_cache()
    assert cached is not None
    assert len(cached["rows"]) == 1


def test_scan_storage_inventory_success_includes_health_issue(monkeypatch):
    server = HealthServer()
    _unlock(server)
    card = HealthCard(
        card_id=1,
        name="Hartford",
        host="10.0.0.1",
        port=22,
        username="u",
        key_path="/tmp/key",
        device_profile="flashsystem_7200",
    )
    server._cards[1] = card
    server.set_monitor_enabled(card_id=1, enabled=True)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)
    monkeypatch.setattr(
        server,
        "_storage_inventory_health_issues",
        lambda _card: [
            {"severity": "warn", "category": "capacity", "message": "Running at 92.0% capacity"}
        ],
    )

    def _runner(_card):
        def run(command):
            if "lscloudcallhome" in command:
                return "id:status\n0:enabled\n"
            if "lsdnsserver" in command:
                return "id:name:IP_address\n0:dns1:10.1.1.1\n"
            if "lsemailserver" in command:
                return "id:name:IP_address:port\n0:smtp1:172.29.62.98:25\n"
            if "lsrcrelationship" in command:
                return "id:name:master_cluster_id\n0:rel1:1\n"
            if "lssystem" in command:
                return (
                    "id:78E37V9\nname:v7kcon-g3v1\n"
                    "product_name:IBM FlashSystem 7200\n"
                    "cluster_ntp_IP_address:10.3.3.3\n"
                )
            return ""

        return run

    monkeypatch.setattr(server, "_lun_run_command", _runner)
    result = server.scan_storage_inventory_live()
    assert result["errors"] == []
    row = result["rows"][0]
    assert "Running at 92.0% capacity" in row["issues"]


def test_scan_storage_inventory_failure_retains_health_issue(monkeypatch):
    server = HealthServer()
    _unlock(server)
    card = HealthCard(
        card_id=2,
        name="Anderson",
        host="10.0.0.2",
        port=22,
        username="u",
        key_path="/tmp/key",
        device_profile="flashsystem_7200",
    )
    server._cards[2] = card
    server.set_monitor_enabled(card_id=2, enabled=True)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)
    monkeypatch.setattr(
        server,
        "_storage_inventory_health_issues",
        lambda _card: [
            {"severity": "critical", "category": "capacity", "message": "Running at 95.0% capacity"}
        ],
    )

    def _boom(_card):
        raise RuntimeError("ssh connection refused")

    monkeypatch.setattr(server, "_scan_storage_inventory_card", _boom)
    result = server.scan_storage_inventory_live()
    assert len(result["errors"]) == 1
    assert result["errors"][0]["error"] == "ssh connection refused"
    row = result["rows"][0]
    assert "Running at 95.0% capacity" in row["issues"]
    assert "ssh connection refused" in row["issues"]


def test_scan_storage_inventory_svc_per_topic_failure_keeps_siblings(monkeypatch):
    server = HealthServer()
    _unlock(server)
    card = HealthCard(
        card_id=1,
        name="Hartford",
        host="10.0.0.1",
        port=22,
        username="u",
        key_path="/tmp/key",
        device_profile="flashsystem_7200",
    )
    server._cards[1] = card
    server.set_monitor_enabled(card_id=1, enabled=True)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)

    def _runner(_card):
        def run(command):
            if "lsemailserver" in command:
                raise ValueError("CMMVC5701E No object ID was specified.")
            if "lscloudcallhome" in command:
                return "id:status\n0:enabled\n"
            if "lsdnsserver" in command:
                return "id:name:IP_address\n0:dns1:10.1.1.1\n"
            if "lsrcrelationship" in command:
                return "id:name:master_cluster_id\n0:rel1:1\n"
            if "lssystem" in command:
                return (
                    "id:78E37V9\nname:v7kcon-g3v1\n"
                    "product_name:IBM FlashSystem 7200\n"
                    "cluster_ntp_IP_address:10.3.3.3\n"
                )
            return ""

        return run

    monkeypatch.setattr(server, "_lun_run_command", _runner)
    result = server.scan_storage_inventory_live()
    assert result["errors"] == []
    row = result["rows"][0]
    assert row["phone_home"].lower().startswith("yes")
    assert row["data_protection"].lower().startswith("yes")
    assert row["smtp"] == "unknown"
    assert "No object ID was specified" in row["issues"]


def test_scan_storage_inventory_prefers_card_serial_over_cluster_id(monkeypatch):
    server = HealthServer()
    _unlock(server)
    card = HealthCard(
        card_id=1,
        name="Hartford",
        host="10.0.0.1",
        port=22,
        username="u",
        key_path="/tmp/key",
        device_profile="flashsystem_7200",
        serial_number="78E31NF",
    )
    server._cards[1] = card
    server.set_monitor_enabled(card_id=1, enabled=True)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)

    def _runner(_card):
        def run(command):
            if "lscloudcallhome" in command:
                return "id:status\n0:enabled\n"
            if "lsdnsserver" in command:
                return "id:name:IP_address\n0:dns1:10.1.1.1\n"
            if "lsemailserver" in command:
                return "id:name:IP_address:port\n0:smtp1:172.29.62.98:25\n"
            if "lsrcrelationship" in command:
                return "id:name:master_cluster_id\n0:rel1:1\n"
            if "lssystem" in command:
                return (
                    "id:0000020420A18C4E\nname:v7kcon-g3v1\n"
                    "product_name:IBM FlashSystem 7200\n"
                    "cluster_ntp_IP_address:10.3.3.3\n"
                )
            return ""

        return run

    monkeypatch.setattr(server, "_lun_run_command", _runner)
    result = server.scan_storage_inventory_live()
    assert result["errors"] == []
    row = result["rows"][0]
    assert row["serial"] == "78E31NF"


def test_scan_storage_inventory_includes_unmonitored_flashsystem_and_3par(monkeypatch):
    server = HealthServer()
    _unlock(server)
    server._cards[1] = HealthCard(
        card_id=1,
        name="Hartford",
        host="10.0.0.1",
        port=22,
        username="u",
        key_path="/tmp/key",
        device_profile="flashsystem_7200",
    )
    server._cards[2] = HealthCard(
        card_id=2,
        name="Tempe",
        host="10.0.0.2",
        port=22,
        username="u",
        key_path="/tmp/key",
        device_profile="hpe_3par_8400",
    )
    server._cards[3] = HealthCard(
        card_id=3,
        name="SVR-WEB",
        host="45.76.232.99",
        port=22,
        username="root",
        key_path="/tmp/key",
        device_profile="vultr_vps",
    )
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)

    def _fake_scan(card):
        return {
            "site": card.name,
            "host": card.name,
            "ip": str(card.host or ""),
            "profile": card.device_profile,
            "issues": "",
        }

    monkeypatch.setattr(server, "_scan_storage_inventory_card", _fake_scan)
    result = server.scan_storage_inventory_live()
    sites = {row["site"] for row in result["rows"]}
    assert sites == {"Hartford", "Tempe"}
    assert result["total_devices"] == 2


def test_export_storage_inventory_uses_cache_without_unlock(monkeypatch):
    server = HealthServer()
    monkeypatch.setattr(server, "is_unlocked", lambda: False)
    server.set_storage_inventory_cache(
        {
            "generated_at": "2026-08-10T12:00:00",
            "rows": [
                {
                    "site": "Hartford",
                    "host": "Hartford",
                    "ip": "10.0.0.1",
                    "model": "IBM FlashSystem 7200",
                    "serial": "ABC",
                    "location": "Hartford",
                    "phone_home": "Yes — IBM",
                    "data_protection": "Yes",
                    "smtp": "172.29.62.98",
                    "issues": "",
                },
                {
                    "site": "Bad",
                    "host": "Bad",
                    "ip": "10.0.0.2",
                    "model": "IBM FlashSystem 7200",
                    "serial": "DEF",
                    "location": "Bad",
                    "phone_home": "No — Not configured",
                    "data_protection": "No — Not configured",
                    "smtp": "No IP — Not configured",
                    "issues": "Phone Home not configured",
                },
            ],
            "errors": [],
            "total_devices": 2,
            "devices_with_issues": 1,
        }
    )
    body, filename, content_type = server.export_storage_inventory_bytes()
    assert filename.endswith(".xlsx")
    assert "sheet" in content_type or "spreadsheet" in content_type or "octet" in content_type
    wb = load_workbook(BytesIO(body))
    assert wb.sheetnames == ["Inventory", "Issues Summary"]
