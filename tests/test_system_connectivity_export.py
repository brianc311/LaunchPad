from io import BytesIO
import zipfile
from openpyxl import load_workbook
from launchpad.system_connectivity_export import (
    TOPIC_CSV_NAMES,
    TOPIC_SHEETS,
    export_system_connectivity_csv_zip,
    export_system_connectivity_xlsx,
    filter_payload_by_card_id,
)


def _sample():
    row_a = {
        "card_id": 1, "site": "Hartford", "card_name": "Hartford", "host": "10.0.0.1",
        "vendor": "ibm", "profile": "flashsystem_7200", "configured": "yes",
        "status": "enabled", "details": "10.1.1.1", "error": "",
    }
    row_b = {**row_a, "card_id": 2, "site": "Primera", "card_name": "Primera", "host": "10.0.0.2", "vendor": "hpe"}
    return {
        "call_home": [row_a, row_b],
        "dns": [row_a],
        "snmp": [],
        "ntp": [row_a],
        "errors": [],
    }


def test_filter_by_card_id():
    scoped = filter_payload_by_card_id(_sample(), card_id=1)
    assert len(scoped["call_home"]) == 1
    assert scoped["call_home"][0]["card_name"] == "Hartford"


def test_xlsx_six_sheets():
    body = export_system_connectivity_xlsx(_sample())
    wb = load_workbook(BytesIO(body))
    assert wb.sheetnames == ["Call Home", "DNS", "SNMP", "NTP", "Firmware", "License Key"]


def test_csv_zip_members():
    body = export_system_connectivity_csv_zip(_sample())
    with zipfile.ZipFile(BytesIO(body)) as zf:
        names = set(zf.namelist())
    assert names == {
        "call_home.csv",
        "dns.csv",
        "snmp.csv",
        "ntp.csv",
        "firmware.csv",
        "license_key.csv",
    }


def test_export_includes_firmware_sheet_and_columns():
    payload = {
        "call_home": [],
        "dns": [],
        "snmp": [],
        "ntp": [],
        "firmware": [
            {
                "site": "A",
                "card_name": "A",
                "host": "1.1.1.1",
                "vendor": "ibm",
                "profile": "flashsystem_7300",
                "configured": "yes",
                "status": "behind",
                "details": "8.6.0 → 8.6.1",
                "error": "",
                "current": "8.6.0",
                "latest": "8.6.1",
                "versions_behind": "1",
            }
        ],
        "errors": [],
    }
    assert TOPIC_SHEETS["firmware"] == "Firmware"
    wb = load_workbook(BytesIO(export_system_connectivity_xlsx(payload)))
    assert "Firmware" in wb.sheetnames
    sheet = wb["Firmware"]
    headers = [cell.value for cell in sheet[1]]
    assert "Current" in headers
    assert "Versions behind" in headers
    z = zipfile.ZipFile(BytesIO(export_system_connectivity_csv_zip(payload)))
    assert "firmware.csv" in z.namelist()


def test_export_includes_license_key_sheet_and_columns():
    payload = {
        "call_home": [],
        "dns": [],
        "snmp": [],
        "ntp": [],
        "firmware": [],
        "license_key": [
            {
                "site": "A",
                "card_name": "A",
                "host": "1.1.1.1",
                "vendor": "hpe",
                "profile": "hpe_3par_8450",
                "configured": "yes",
                "status": "ok",
                "details": "3 features",
                "error": "",
                "key_generation_date": "2017-09-19",
                "date": "",
                "time": "",
                "encryption_licensed": "",
                "feature": "Remote Copy",
                "expiration": "—",
            }
        ],
        "errors": [],
    }
    assert TOPIC_SHEETS["license_key"] == "License Key"
    assert TOPIC_CSV_NAMES["license_key"] == "license_key.csv"
    wb = load_workbook(BytesIO(export_system_connectivity_xlsx(payload)))
    assert "License Key" in wb.sheetnames
    sheet = wb["License Key"]
    headers = [cell.value for cell in sheet[1]]
    assert "Key generation date" in headers
    assert "Date" in headers
    assert "Time" in headers
    assert "Encryption licensed" in headers
    assert "Feature" in headers
    assert "Expiration" in headers
    assert headers.index("Profile") < headers.index("Key generation date")
    assert headers.index("Expiration") < headers.index("Configured")
    z = zipfile.ZipFile(BytesIO(export_system_connectivity_csv_zip(payload)))
    assert "license_key.csv" in z.namelist()
