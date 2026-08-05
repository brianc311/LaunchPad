from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

from launchpad.dell_report_export import (
    HOME_SHEET_NAME,
    ORDERED_SHEET_NAMES,
    STUB_SHEET_NAMES,
    _FIRST_DATA_COL,
    _HEADER_FILL,
    build_dell_report_workbook,
    bytes_to_gib,
    workbook_to_bytes,
)
from launchpad.dell_report_leds import UTIL_YELLOW_THRESHOLD


def _minimal_row(**overrides) -> dict:
    row = {
        "facility": "Data center -WAG1",
        "array_name": "V7K001",
        "model": "FlashSystem 9200",
        "prior_usable_gib": 100.0,
        "prior_used_gib": 50.0,
        "prior_util": 0.5,
        "curr_usable_gib": 100.0,
        "curr_used_gib": 60.0,
        "curr_util": 0.6,
        "weekly_growth": 0.2,
    }
    row.update(overrides)
    return row


def test_bytes_to_gib():
    assert bytes_to_gib(0) == 0.0
    assert bytes_to_gib(1024**3) == 1.0
    assert bytes_to_gib(2.5 * 1024**3) == 2.5


def test_stub_sheet_names_include_required_tabs():
    names = " ".join(STUB_SHEET_NAMES)
    for token in ("PowerMax", "PowerStore", "NetApp", "Data Domain", "ECS", "Forecast - Wkly"):
        assert token in names


def test_ordered_sheets_place_ibm_hp_after_netapp():
    assert ORDERED_SHEET_NAMES.index("NetApp Forecast - Wkly") < ORDERED_SHEET_NAMES.index(
        "IBM Report"
    )
    assert ORDERED_SHEET_NAMES.index("IBM Report") < ORDERED_SHEET_NAMES.index(
        "IBM Report - Wkly"
    )
    assert ORDERED_SHEET_NAMES.index("IBM Report - Wkly") < ORDERED_SHEET_NAMES.index(
        "IBM Forecast"
    )
    assert ORDERED_SHEET_NAMES.index("IBM Forecast") < ORDERED_SHEET_NAMES.index(
        "IBM Forecast - Wkly"
    )
    assert ORDERED_SHEET_NAMES.index("IBM Report") < ORDERED_SHEET_NAMES.index("HP Report")
    assert ORDERED_SHEET_NAMES.index("HP Report") < ORDERED_SHEET_NAMES.index(
        "HP Report - Wkly"
    )


def test_workbook_has_report_wkly_sheets_with_week_columns():
    from launchpad.dell_report_snapshots import upsert_week_snapshot

    store = {}
    store = upsert_week_snapshot(
        store,
        card_id=1,
        week="2026-W31",
        usable_bytes=100 * 1024**3,
        used_bytes=40 * 1024**3,
        model="M1",
        facility="Remote",
        family="ibm",
        array_name="A1",
        captured_at="2026-07-28T00:00:00+00:00",
    )
    store = upsert_week_snapshot(
        store,
        card_id=1,
        week="2026-W32",
        usable_bytes=100 * 1024**3,
        used_bytes=60 * 1024**3,
        model="M1",
        facility="Remote",
        family="ibm",
        array_name="A1",
        captured_at="2026-08-04T00:00:00+00:00",
    )
    rows = [
        _minimal_row(
            card_id=1, facility="Remote", array_name="A1", model="M1", curr_util=0.6
        )
    ]
    wb = build_dell_report_workbook(
        ibm_rows=rows,
        hp_rows=[],
        snapshot_store=store,
        report_date=datetime(2026, 8, 5, tzinfo=timezone.utc),
    )
    assert "IBM Report - Wkly" in wb.sheetnames
    assert "HP Report - Wkly" in wb.sheetnames
    ws = wb["IBM Report - Wkly"]
    headers = [ws.cell(row=9, column=c).value for c in range(3, 20)]
    assert "Facility" in headers
    assert sum(1 for h in headers if h and "Utilization" in str(h)) >= 2
    assert ws.cell(row=10, column=3).value == "Remote"
    assert ws.cell(row=10, column=5).value == "M1"


def test_hp_forecast_wkly_has_data_rows():
    wb = build_dell_report_workbook(
        ibm_rows=[],
        hp_rows=[_minimal_row(curr_util=0.25, weekly_growth=None)],
    )
    ws = wb["HP Forecast - Wkly"]
    assert ws.cell(row=10, column=4).value  # array
    assert ws.cell(row=10, column=6).value == 0.25
    assert ws.cell(row=10, column=7).value == 0.25
    assert ws.cell(row=10, column=10).value == 0.25


def test_workbook_has_ibm_hp_and_stub():
    wb = build_dell_report_workbook(
        ibm_rows=[_minimal_row()],
        hp_rows=[_minimal_row(array_name="3PAR001", model="HPE 3PAR 8450")],
    )
    assert "IBM Report" in wb.sheetnames
    assert "HP Report" in wb.sheetnames
    assert any("PowerMax" in name for name in wb.sheetnames)
    stub = wb[[name for name in wb.sheetnames if name == "PowerMax Report"][0]]
    assert stub.cell(9, _FIRST_DATA_COL).value == "Facility"
    assert stub.max_row == 9


def test_home_sheet_title():
    wb = build_dell_report_workbook(ibm_rows=[], hp_rows=[])
    home = wb.worksheets[0]
    title_cells = [
        str(cell.value)
        for row in home.iter_rows(max_row=20)
        for cell in row
        if cell.value
    ]
    joined = " ".join(title_cells)
    assert "Dell Technologies Managed Services" in joined
    assert "Capacity Management Report" in joined


def test_rows_sorted_by_facility_then_array_name():
    wb = build_dell_report_workbook(
        ibm_rows=[
            _minimal_row(facility="Z-facility", array_name="Z-array"),
            _minimal_row(facility="A-facility", array_name="B-array"),
            _minimal_row(facility="A-facility", array_name="A-array"),
        ],
        hp_rows=[],
    )
    ws = wb["IBM Report"]
    data_start = _data_start_row(ws)
    facilities = [
        ws.cell(row, _FIRST_DATA_COL).value for row in range(data_start, ws.max_row + 1)
    ]
    arrays = [
        ws.cell(row, _FIRST_DATA_COL + 1).value
        for row in range(data_start, ws.max_row + 1)
    ]
    assert facilities == ["A-facility", None, "Z-facility"]
    assert arrays == ["A-array", "B-array", "Z-array"]


def test_utilization_icon_set_leds():
    wb = build_dell_report_workbook(
        ibm_rows=[
            _minimal_row(array_name="Cold", curr_util=0.5, prior_util=0.75),
            _minimal_row(array_name="Hot", curr_util=0.95, prior_util=0.95),
        ],
        hp_rows=[],
    )
    ws = wb["IBM Report"]
    start = _data_start_row(ws)
    end = ws.max_row
    rules = [rule for group in ws.conditional_formatting for rule in group.rules]
    icon_rules = [rule for rule in rules if getattr(rule, "type", None) == "iconSet"]
    assert icon_rules
    icon_set = icon_rules[0].iconSet
    assert icon_set is not None
    assert icon_set.iconSet == "3TrafficLights1"
    assert icon_set.reverse is True
    vals = [float(cfvo.val) for cfvo in icon_set.cfvo]
    assert UTIL_YELLOW_THRESHOLD in vals
    # Values still written (icon + value); fills are not the primary LED.
    assert ws.cell(start, 11).value == 0.5
    assert ws.cell(start + 1, 11).value == 0.95
    assert end >= start


def test_banner_has_sheet_title_and_logos():
    wb = build_dell_report_workbook(ibm_rows=[_minimal_row()], hp_rows=[])
    ws = wb["IBM Report"]
    assert any(
        cell.value == "IBM Report"
        for row in ws.iter_rows(min_row=1, max_row=6, max_col=12)
        for cell in row
    )
    assert len(ws._images) >= 2


def test_facility_shown_only_on_first_row_of_group():
    wb = build_dell_report_workbook(
        ibm_rows=[
            _minimal_row(facility="A-facility", array_name="B-array"),
            _minimal_row(facility="A-facility", array_name="A-array"),
            _minimal_row(facility="Z-facility", array_name="Z-array"),
        ],
        hp_rows=[],
    )
    ws = wb["IBM Report"]
    start = _data_start_row(ws)
    facilities = [
        ws.cell(row, _FIRST_DATA_COL).value for row in range(start, ws.max_row + 1)
    ]
    arrays = [
        ws.cell(row, _FIRST_DATA_COL + 1).value for row in range(start, ws.max_row + 1)
    ]
    assert arrays == ["A-array", "B-array", "Z-array"]
    assert facilities == ["A-facility", None, "Z-facility"]


def test_gib_and_utilization_values_written():
    wb = build_dell_report_workbook(
        ibm_rows=[_minimal_row()],
        hp_rows=[],
        report_date=datetime(2026, 6, 15, tzinfo=timezone.utc),
    )
    ws = wb["IBM Report"]
    row = _data_start_row(ws)
    c = _FIRST_DATA_COL
    assert ws.cell(row, c).value == "Data center -WAG1"
    assert ws.cell(row, c + 1).value == "V7K001"
    assert ws.cell(row, c + 2).value == "FlashSystem 9200"
    assert ws.cell(row, c + 3).value == 100.0
    assert ws.cell(row, c + 4).value == 50.0
    assert ws.cell(row, c + 5).value == 0.5
    assert ws.cell(row, c + 6).value == 100.0
    assert ws.cell(row, c + 7).value == 60.0
    assert ws.cell(row, c + 8).value == 0.6
    assert ws.cell(row, c + 9).value == 0.2
    assert ws.cell(row, c + 5).number_format == "0.0%"
    assert ws.cell(row, c + 8).number_format == "0.0%"
    assert ws.cell(row, c + 9).number_format == "0.0%"
    assert "Useable Capacity" in str(ws.cell(9, c + 3).value)
    assert ws.cell(9, c + 3).fill.fgColor.rgb[-6:].upper() == _HEADER_FILL.fgColor.rgb[-6:].upper()


def test_report_embeds_logos_when_assets_present():
    wb = build_dell_report_workbook(ibm_rows=[_minimal_row()], hp_rows=[])
    assert len(wb["IBM Report"]._images) >= 1
    assert len(wb["HP Forecast"]._images) >= 1


def test_utilization_conditional_formatting_uses_icon_set():
    wb = build_dell_report_workbook(
        ibm_rows=[_minimal_row()],
        hp_rows=[_minimal_row()],
    )
    for sheet_name in ("IBM Report", "HP Report"):
        ws = wb[sheet_name]
        rules = [rule for group in ws.conditional_formatting for rule in group.rules]
        icon_rules = [rule for rule in rules if getattr(rule, "type", None) == "iconSet"]
        assert icon_rules
        assert icon_rules[0].iconSet.iconSet == "3TrafficLights1"


def test_workbook_to_bytes_roundtrip():
    wb = build_dell_report_workbook(
        ibm_rows=[_minimal_row()],
        hp_rows=[_minimal_row(array_name="3PAR001")],
    )
    payload = workbook_to_bytes(wb)
    assert isinstance(payload, bytes)
    assert payload[:2] == b"PK"
    loaded = load_workbook(BytesIO(payload))
    assert "IBM Report" in loaded.sheetnames
    assert "HP Report" in loaded.sheetnames


def _data_start_row(ws) -> int:
    for row in range(1, ws.max_row + 1):
        for col in range(1, 15):
            if ws.cell(row, col).value == "Facility":
                return row + 1
    raise AssertionError("Facility header row not found")


def _forecast_data_start_row(ws) -> int:
    for row in range(1, ws.max_row + 1):
        for col in range(1, 15):
            if ws.cell(row, col).value == "3 Month":
                return row + 1
    raise AssertionError("3 Month header row not found")


def test_workbook_includes_ibm_and_hp_forecast_sheets():
    wb = build_dell_report_workbook(
        ibm_rows=[_minimal_row(curr_util=0.61)],
        hp_rows=[_minimal_row(array_name="3PAR001", curr_util=0.82)],
    )
    assert "IBM Forecast" in wb.sheetnames
    assert "HP Forecast" in wb.sheetnames
    ibm_f = wb["IBM Forecast"]
    start = _forecast_data_start_row(ibm_f)
    for col in (6, 7, 8, 9, 10):
        assert ibm_f.cell(start, col).value == 0.61
    rules = [rule for group in ibm_f.conditional_formatting for rule in group.rules]
    assert any(getattr(rule, "type", None) == "iconSet" for rule in rules)


def test_home_lists_forecast_sheets():
    wb = build_dell_report_workbook(ibm_rows=[_minimal_row()], hp_rows=[])
    home_text = " ".join(
        str(c.value)
        for row in wb[HOME_SHEET_NAME].iter_rows(max_row=40)
        for c in row
        if c.value
    )
    assert "IBM Forecast" in home_text
    assert "HP Forecast" in home_text
    assert "PowerMax Forecast - Wkly" in home_text
