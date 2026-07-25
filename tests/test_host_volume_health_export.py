import csv
import inspect
import io
import zipfile
from io import BytesIO, TextIOWrapper

from openpyxl import load_workbook

from launchpad.health_server import HealthServer, _HealthHandler
from launchpad.host_volume_health_export import (
    HOST_HEADERS,
    VOLUME_HEADERS,
    export_host_volume_health_csv_zip,
    export_host_volume_health_xlsx,
    filter_payload_by_card_id,
)


def _sample_payload() -> dict:
    return {
        "hosts": [
            {
                "card_name": "Hartford",
                "host": "10.0.0.1",
                "vendor": "ibm",
                "host_name": "bad_host",
                "status": "offline",
            }
        ],
        "volumes": [
            {
                "card_name": "Hartford",
                "host": "10.0.0.1",
                "vendor": "ibm",
                "volume_name": "bad_vol",
                "pool_or_cpg": "Pool0",
                "status": "degraded",
            }
        ],
    }


def test_xlsx_has_hosts_and_volumes_sheets():
    workbook = load_workbook(
        BytesIO(export_host_volume_health_xlsx(_sample_payload()))
    )
    assert workbook.sheetnames == ["Hosts", "Volumes"]
    assert workbook["Hosts"].freeze_panes == "A2"
    assert workbook["Hosts"]["A2"].value == "Hartford"
    assert workbook["Hosts"]["D2"].value == "bad_host"
    assert workbook["Volumes"]["D2"].value == "bad_vol"
    assert workbook["Volumes"]["E2"].value == "Pool0"


def test_csv_zip_contains_hosts_and_volumes():
    with zipfile.ZipFile(
        BytesIO(export_host_volume_health_csv_zip(_sample_payload()))
    ) as archive:
        assert set(archive.namelist()) == {"hosts.csv", "volumes.csv"}
        with archive.open("hosts.csv") as raw:
            hosts = list(csv.DictReader(TextIOWrapper(raw, encoding="utf-8-sig")))
        with archive.open("volumes.csv") as raw:
            volumes = list(csv.DictReader(TextIOWrapper(raw, encoding="utf-8-sig")))

    assert hosts[0][HOST_HEADERS[3]] == "bad_host"
    assert volumes[0][VOLUME_HEADERS[3]] == "bad_vol"
    assert volumes[0][VOLUME_HEADERS[4]] == "Pool0"


def test_filter_payload_by_card_id():
    payload = {
        "hosts": [
            {"card_name": "SiteA", "host_name": "h1"},
            {"card_name": "SiteB", "host_name": "h2"},
        ],
        "volumes": [
            {"card_name": "SiteA", "volume_name": "v1"},
            {"card_name": "SiteB", "volume_name": "v2"},
        ],
    }
    filtered = filter_payload_by_card_id(payload, card_name="SiteB")
    assert len(filtered["hosts"]) == 1
    assert filtered["hosts"][0]["host_name"] == "h2"
    assert len(filtered["volumes"]) == 1
    assert filtered["volumes"][0]["volume_name"] == "v2"


def _call_hv_export_api(monkeypatch, server: HealthServer, query: str = ""):
    handler = object.__new__(_HealthHandler)
    handler.path = f"/api/host-volume-health/export{query}"
    sent: dict = {}

    def _send_bytes(body, *, content_type, filename, status=200):
        sent["body"] = body
        sent["content_type"] = content_type
        sent["filename"] = filename
        sent["status"] = status

    def _send_json(data, status=200):
        sent["json"] = data
        sent["status"] = status

    handler._send_bytes = _send_bytes
    handler._send_json = _send_json
    monkeypatch.setattr("launchpad.health_server.get_health_server", lambda: server)

    handler.do_GET()
    return sent


def test_export_api_requires_cached_scan(monkeypatch):
    server = HealthServer()
    sent = _call_hv_export_api(monkeypatch, server, "?format=xlsx&open=0")
    assert sent["status"] == 404
    assert "refresh" in sent["json"]["error"].lower()


def test_export_api_sends_xlsx_from_cache(monkeypatch):
    server = HealthServer()
    server.set_host_volume_health_cache(_sample_payload())
    sent = _call_hv_export_api(monkeypatch, server, "?format=xlsx&open=0")
    assert sent["status"] == 200
    assert sent["body"][:2] == b"PK"
    assert sent["content_type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert sent["filename"].startswith("Host_Volume_Health_")


def test_export_api_sends_csv_zip_from_cache(monkeypatch):
    server = HealthServer()
    server.set_host_volume_health_cache(_sample_payload())
    sent = _call_hv_export_api(monkeypatch, server, "?format=csv&open=0")
    assert sent["status"] == 200
    assert sent["body"][:2] == b"PK"
    assert sent["content_type"] == "application/zip"
    assert sent["filename"].endswith(".zip")


def test_live_scan_stores_cache(monkeypatch):
    from launchpad.health_server import HealthCard

    server = HealthServer()
    server.set_settings_backend(lambda _key, default: default, lambda _key, _value: None)
    card = HealthCard(
        card_id=1,
        name="Hartford",
        host="10.0.0.1",
        port=22,
        username="user",
        key_path="/tmp/key",
        device_profile="flashsystem_7200",
    )
    server._cards[1] = card
    server.set_monitor_enabled(card_id=1, enabled=True)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)
    monkeypatch.setattr(
        server,
        "_lun_run_command",
        lambda _card: lambda command: (
            "id:name:status\n0:bad_host:offline\n"
            if "lshost" in command
            else "id:name:mdisk_grp_name:status\n0:bad_vol:Pool0:degraded\n"
        ),
    )
    server.scan_host_volume_health_live()
    cached = server.get_host_volume_health_cache()
    assert cached is not None
    assert len(cached["hosts"]) == 1
    assert cached["hosts"][0]["host_name"] == "bad_host"


def test_page_export_buttons_call_export_api():
    from launchpad.host_volume_health_page import HOST_VOLUME_HEALTH_HTML

    assert "/api/host-volume-health/export" in HOST_VOLUME_HEALTH_HTML
    assert 'id="hv-export-xlsx-btn"' in HOST_VOLUME_HEALTH_HTML
    assert 'id="hv-export-csv-btn"' in HOST_VOLUME_HEALTH_HTML


def test_health_handler_declares_hv_export_route():
    source = inspect.getsource(_HealthHandler.do_GET)
    assert "/api/host-volume-health/export" in source
