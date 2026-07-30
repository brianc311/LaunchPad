import inspect
from io import BytesIO

from openpyxl import load_workbook

from launchpad.health_server import HealthCard, HealthServer, _HealthHandler


def _unlock(server: HealthServer) -> None:
    server.set_settings_backend(lambda _key, default: default, lambda _key, _value: None)


def _svc_card(
    card_id: int,
    name: str,
    host: str,
    *,
    category: str = "",
    device_profile: str = "flashsystem_7200",
) -> HealthCard:
    return HealthCard(
        card_id=card_id,
        name=name,
        host=host,
        port=22,
        username="user",
        key_path="/tmp/key",
        device_profile=device_profile,
        category=category,
    )


def _summary(name: str, **extra) -> dict:
    base = {
        "name": name,
        "status": "idle_or_copied",
        "flash_time": "2026-07-30 10:00:00",
        "progress_pct": None,
        "fc_map_count": 2,
        "host_map_count": 1,
        "total_size": "10.0 GB",
        "policy": "",
        "snaps_per_week": 3,
    }
    base.update(extra)
    return base


def test_live_scan_requires_unlock(monkeypatch):
    server = HealthServer()
    monkeypatch.setattr(server, "is_unlocked", lambda: False)
    try:
        server.scan_fc_cg_summary_live()
        assert False, "expected RuntimeError"
    except RuntimeError as exc:
        assert "unlock" in str(exc).lower()


def test_live_scan_happy_path_mocked_inventory(monkeypatch):
    server = HealthServer()
    _unlock(server)
    server._cards[1] = _svc_card(1, "Hartford", "10.0.0.1", category="General")
    server._cards[2] = _svc_card(2, "Anderson", "10.0.0.2", category="General")
    server.set_monitor_enabled(card_id=1, enabled=True)
    server.set_monitor_enabled(card_id=2, enabled=True)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)

    def fake_inventory(card_id: int, *, include_summaries: bool = True):
        assert include_summaries is True
        if card_id == 1:
            return {
                "ok": True,
                "warnings": [],
                "summaries": [
                    _summary("AWD1_FC"),
                    _summary("STOPPED_CG", status="stopped", progress_pct=50),
                ],
            }
        if card_id == 2:
            return {
                "ok": True,
                "warnings": [],
                "summaries": [_summary("AAN1_FC", fc_map_count=84, host_map_count=48)],
            }
        return {"ok": False, "warnings": ["missing"], "summaries": []}

    monkeypatch.setattr(server, "fc_consistgrp_inventory", fake_inventory)

    result = server.scan_fc_cg_summary_live()
    assert result["errors"] == []
    assert len(result["rows"]) == 3
    by_key = {row["row_key"]: row for row in result["rows"]}
    hartford = by_key["1:AWD1_FC"]
    assert hartford["site"] == "Hartford"
    assert hartford["site"] != "General"
    assert hartford["card_name"] == "Hartford"
    assert hartford["host"] == "10.0.0.1"
    assert hartford["card_id"] == 1
    assert hartford["name"] == "AWD1_FC"
    assert hartford["status"] == "idle_or_copied"
    assert hartford["flash_time"] == "2026-07-30 10:00:00"
    assert hartford["fc_map_count"] == 2
    assert hartford["host_map_count"] == 1
    assert hartford["total_size"] == "10.0 GB"
    assert hartford["snaps_per_week"] == 3
    assert by_key["1:STOPPED_CG"]["progress_pct"] == 50
    anderson = by_key["2:AAN1_FC"]
    assert anderson["site"] == "Anderson"
    assert anderson["site"] != "General"
    cached = server.get_fc_cg_summary_live_cache()
    assert cached is not None
    assert len(cached["rows"]) == 3


def test_export_selected_only_from_cache():
    server = HealthServer()
    _unlock(server)
    server.set_fc_cg_summary_live_cache(
        {
            "rows": [
                {
                    "site": "Hartford",
                    "card_name": "Hartford",
                    "host": "10.0.0.1",
                    "card_id": 1,
                    "name": "AWD1_FC",
                    "status": "idle_or_copied",
                    "flash_time": "",
                    "progress_pct": None,
                    "fc_map_count": 1,
                    "host_map_count": 1,
                    "total_size": "1 GB",
                    "policy": "",
                    "snaps_per_week": 1,
                    "row_key": "1:AWD1_FC",
                },
                {
                    "site": "Anderson",
                    "card_name": "Anderson",
                    "host": "10.0.0.2",
                    "card_id": 2,
                    "name": "AAN1_FC",
                    "status": "idle_or_copied",
                    "flash_time": "",
                    "progress_pct": None,
                    "fc_map_count": 2,
                    "host_map_count": 2,
                    "total_size": "2 GB",
                    "policy": "",
                    "snaps_per_week": 2,
                    "row_key": "2:AAN1_FC",
                },
            ],
            "errors": [],
        }
    )

    body, filename, content_type = server.export_fc_cg_summary_selected_bytes(
        selected=["1:AWD1_FC"],
        open_after=False,
    )
    assert filename.startswith("FC_CG_Summary_MultiSite_")
    assert filename.endswith(".xlsx")
    assert "spreadsheetml" in content_type
    workbook = load_workbook(BytesIO(body))
    assert set(workbook.sheetnames) == {"Hartford"}
    sheet = workbook["Hartford"]
    assert sheet["C2"].value == "AWD1_FC"


def test_export_selected_requires_cache():
    server = HealthServer()
    _unlock(server)
    try:
        server.export_fc_cg_summary_selected_bytes(selected=["1:CG"], open_after=False)
        assert False, "expected LookupError"
    except LookupError as exc:
        assert "refresh" in str(exc).lower()


def test_export_selected_empty_selection_raises():
    server = HealthServer()
    _unlock(server)
    server.set_fc_cg_summary_live_cache({"rows": [], "errors": []})
    try:
        server.export_fc_cg_summary_selected_bytes(selected=[], open_after=False)
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "select" in str(exc).lower()


def test_export_selected_unknown_keys_raise():
    server = HealthServer()
    _unlock(server)
    server.set_fc_cg_summary_live_cache(
        {
            "rows": [
                {
                    "site": "Hartford",
                    "card_name": "Hartford",
                    "host": "10.0.0.1",
                    "card_id": 1,
                    "name": "AWD1_FC",
                    "status": "idle_or_copied",
                    "flash_time": "",
                    "progress_pct": None,
                    "fc_map_count": 1,
                    "host_map_count": 1,
                    "total_size": "1 GB",
                    "policy": "",
                    "snaps_per_week": 1,
                    "row_key": "1:AWD1_FC",
                }
            ],
            "errors": [],
        }
    )
    try:
        server.export_fc_cg_summary_selected_bytes(
            selected=["99:MISSING"],
            open_after=False,
        )
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "matching" in str(exc).lower()


def test_api_fc_cg_summary_multisite_routes_declared():
    get_src = inspect.getsource(_HealthHandler.do_GET)
    post_src = inspect.getsource(_HealthHandler.do_POST)
    assert "/api/contingency-groups/fc-cg-summary/live" in get_src
    assert "/api/contingency-groups/fc-cg-summary/export-selected" in post_src
