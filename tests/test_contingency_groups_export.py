from io import BytesIO

from openpyxl import load_workbook

from launchpad.contingency_groups_data import (
    SNAP_SUFFIX,
    seed_contingency_groups,
)
from launchpad.contingency_groups_export import (
    build_contingency_groups_workbook,
    workbook_to_bytes,
)

WINDSOR_WWPN = "51402EC012CFD072"
WINDSOR_UID = "60050768128000A75800000000000000"


def _windsor_group() -> dict:
    for group in seed_contingency_groups():
        if group["id"] == "windsor":
            return group
    raise AssertionError("windsor seed missing")


def _column_values(ws, header: str) -> list[str]:
    col_index = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header:
            col_index = col
            break
    assert col_index is not None, f"missing column {header!r}"
    return [
        str(ws.cell(row=r, column=col_index).value or "")
        for r in range(2, ws.max_row + 1)
    ]


def _all_cell_texts(wb) -> list[str]:
    texts: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    texts.append(str(cell))
    return texts


def test_workbook_has_four_sheets_with_windsor_wwpn_and_uid():
    wb = build_contingency_groups_workbook([_windsor_group()])
    assert [ws.title for ws in wb.worksheets] == [
        "Summary",
        "Hosts",
        "Volumes",
        "Maps",
    ]

    texts = _all_cell_texts(wb)
    assert any(WINDSOR_WWPN in t for t in texts)
    assert any(WINDSOR_UID in t for t in texts)

    hosts_ws = wb["Hosts"]
    wwpn_col = None
    for col in range(1, hosts_ws.max_column + 1):
        if hosts_ws.cell(row=1, column=col).value == "WWPNs":
            wwpn_col = col
            break
    assert wwpn_col is not None
    host_wwpn_cells = [
        str(hosts_ws.cell(row=r, column=wwpn_col).value or "")
        for r in range(2, hosts_ws.max_row + 1)
    ]
    assert any(WINDSOR_WWPN in cell for cell in host_wwpn_cells)
    assert any(";" in cell and WINDSOR_WWPN in cell for cell in host_wwpn_cells)

    volumes_ws = wb["Volumes"]
    uid_col = None
    for col in range(1, volumes_ws.max_column + 1):
        if volumes_ws.cell(row=1, column=col).value == "UID":
            uid_col = col
            break
    assert uid_col is not None
    uid_cells = [
        str(volumes_ws.cell(row=r, column=uid_col).value or "")
        for r in range(2, volumes_ws.max_row + 1)
    ]
    assert WINDSOR_UID in uid_cells


def test_seeded_groups_export_snap_role_on_volumes_and_maps():
    wb = build_contingency_groups_workbook(seed_contingency_groups())

    volumes_ws = wb["Volumes"]
    roles = _column_values(volumes_ws, "Role")
    source_volumes = _column_values(volumes_ws, "Source Volume")
    volume_names = _column_values(volumes_ws, "Volume Name")

    assert "snap" in roles
    assert "source" in roles
    snap_rows = [
        (name, role, source)
        for name, role, source in zip(volume_names, roles, source_volumes)
        if role == "snap"
    ]
    assert snap_rows
    for name, role, source in snap_rows:
        assert name.endswith(SNAP_SUFFIX)
        assert role == "snap"
        assert source and not source.endswith(SNAP_SUFFIX)

    maps_ws = wb["Maps"]
    map_roles = _column_values(maps_ws, "Role")
    assert "snap" in map_roles
    assert "source" in map_roles


def test_workbook_to_bytes_round_trip():
    wb = build_contingency_groups_workbook(seed_contingency_groups())
    data = workbook_to_bytes(wb)
    loaded = load_workbook(BytesIO(data), read_only=True)
    assert len(loaded.sheetnames) == 4
    assert loaded.sheetnames[0] == "Summary"
