import zipfile
from io import BytesIO

from openpyxl import load_workbook

from launchpad.site_lookup_export import (
    export_site_lookup_csv_zip,
    export_site_lookup_xlsx,
    offline_inventory_rows,
    pools_sheet_title,
    snapshot_policies_sheet_wanted,
)


def _hpe_payload():
    return {
        "card": {"name": "HPE-site", "device_profile": "hpe_3par_8400"},
        "hosts": [
            {"host_name": "esx_ok", "status": "online", "type": "VMware", "port_count": "2"},
            {"host_name": "esx_bad", "status": "offline", "type": "VMware", "port_count": "0"},
        ],
        "volumes": [
            {"name": "vv_ok", "status": "normal", "pool": "cpg_a", "capacity": "10"},
            {"name": "vv_bad", "status": "degraded", "pool": "cpg_b", "capacity": "20"},
        ],
        "pools": [{"name": "cpg_a", "used_pct": 10}],
        "consistency_groups": [],
        "consistency_groups_available": False,
    }


def test_pools_sheet_title_hpe_vs_ibm():
    assert pools_sheet_title({"device_profile": "hpe_3par_8400"}) == "CPGs"
    assert pools_sheet_title({"device_profile": "flashsystem_7200"}) == "Pools"


def test_offline_inventory_rows_combined():
    rows = offline_inventory_rows(_hpe_payload())
    assert {(r["row_type"], r["name"]) for r in rows} == {
        ("host", "esx_bad"),
        ("volume", "vv_bad"),
    }


def test_export_xlsx_sheets_and_optional_offline():
    payload = _hpe_payload()
    wb = load_workbook(BytesIO(export_site_lookup_xlsx(payload, include_offline=False)))
    assert wb.sheetnames == ["Hosts", "Volumes", "CPGs"]
    wb2 = load_workbook(BytesIO(export_site_lookup_xlsx(payload, include_offline=True)))
    assert "Offline" in wb2.sheetnames
    assert wb2["Offline"].max_row >= 2


def test_export_csv_zip_no_offline_member():
    raw = export_site_lookup_csv_zip(_hpe_payload())
    with zipfile.ZipFile(BytesIO(raw)) as zf:
        names = set(zf.namelist())
    assert names == {"Hosts.csv", "Volumes.csv", "CPGs.csv"}
    assert "Offline.csv" not in names


def _ibm_payload():
    return {
        "card": {"name": "Hartford", "device_profile": "flashsystem_7200"},
        "hosts": [],
        "volumes": [],
        "pools": [],
        "consistency_groups": [],
        "consistency_groups_available": True,
        "snapshot_policies_available": True,
        "policies": [
            {"name": "esx_snap", "schedule": "every 1 day", "retention": "keep 7 days"}
        ],
    }


def test_snapshot_policies_sheet_wanted():
    assert snapshot_policies_sheet_wanted(_ibm_payload()) is True
    assert snapshot_policies_sheet_wanted(_hpe_payload()) is False
    assert snapshot_policies_sheet_wanted(
        {"snapshot_policies_available": False, "policies": [{"name": "x"}]}
    ) is True


def test_export_xlsx_and_csv_include_policies_for_ibm():
    wb = load_workbook(BytesIO(export_site_lookup_xlsx(_ibm_payload())))
    assert "Policies" in wb.sheetnames
    assert [cell.value for cell in wb["Policies"][1]] == [
        "Name",
        "Schedule",
        "Retention",
    ]
    assert [cell.value for cell in wb["Policies"][2]] == [
        "esx_snap",
        "every 1 day",
        "keep 7 days",
    ]
    raw = export_site_lookup_csv_zip(_ibm_payload())
    with zipfile.ZipFile(BytesIO(raw)) as zf:
        assert "Policies.csv" in zf.namelist()
        text = zf.read("Policies.csv").decode("utf-8")
    assert "esx_snap" in text
    assert "keep 7 days" in text
