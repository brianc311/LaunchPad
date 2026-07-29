from io import BytesIO

from openpyxl import load_workbook

from launchpad.fc_consistgrp_status_export import (
    STATUS_FIELDS,
    STATUS_HEADERS,
    export_fc_consistgrp_status_xlsx,
    filter_status_rows,
)


def _sample_rows() -> list[dict]:
    return [
        {
            "site": "Hartford",
            "card_name": "Hartford",
            "host": "10.0.0.1",
            "name": "AWD1_AS400_CG",
            "status": "idle_or_copied",
            "map_count": 6,
            "flash_time": "2026-01-15 08:30",
            "error": "",
            "bucket": "idle_or_copied",
        },
        {
            "site": "Primera",
            "card_name": "Primera",
            "host": "10.0.0.2",
            "name": "STOPPED_CG",
            "status": "stopped",
            "map_count": 0,
            "flash_time": "",
            "error": "",
            "bucket": "stopped",
        },
        {
            "site": "Primera",
            "card_name": "Primera",
            "host": "10.0.0.2",
            "name": "COPYING_CG",
            "status": "copying",
            "map_count": 2,
            "flash_time": "",
            "error": "",
            "bucket": "copying",
        },
        {
            "site": "Hartford",
            "card_name": "Hartford",
            "host": "10.0.0.1",
            "name": "EMPTY_CG",
            "status": "empty",
            "map_count": 0,
            "flash_time": "",
            "error": "",
            "bucket": "",
        },
    ]


def test_status_headers_and_fields():
    assert STATUS_HEADERS == (
        "Site",
        "Card",
        "Host",
        "CG name",
        "Status",
        "Maps",
        "Flash time",
        "Error",
    )
    assert STATUS_FIELDS == (
        "site",
        "card_name",
        "host",
        "name",
        "status",
        "map_count",
        "flash_time",
        "error",
    )


def test_filter_status_rows_all_returns_every_row():
    rows = _sample_rows()
    assert filter_status_rows(rows, bucket="") == rows
    assert filter_status_rows(rows, bucket="all") == rows


def test_filter_status_rows_by_bucket():
    rows = _sample_rows()
    idle = filter_status_rows(rows, bucket="idle_or_copied")
    assert [row["name"] for row in idle] == ["AWD1_AS400_CG"]

    stopped = filter_status_rows(rows, bucket="stopped")
    assert [row["name"] for row in stopped] == ["STOPPED_CG"]

    copying = filter_status_rows(rows, bucket="copying")
    assert [row["name"] for row in copying] == ["COPYING_CG"]


def test_xlsx_sheet_name_headers_and_row_values():
    body = export_fc_consistgrp_status_xlsx(_sample_rows())
    workbook = load_workbook(BytesIO(body))
    assert workbook.sheetnames == ["FC CG Status"]

    sheet = workbook["FC CG Status"]
    assert sheet.freeze_panes == "A2"
    headers = [cell.value for cell in sheet[1]]
    assert headers == list(STATUS_HEADERS)

    assert sheet["A2"].value == "Hartford"
    assert sheet["B2"].value == "Hartford"
    assert sheet["C2"].value == "10.0.0.1"
    assert sheet["D2"].value == "AWD1_AS400_CG"
    assert sheet["E2"].value == "idle_or_copied"
    assert sheet["F2"].value == 6
    assert sheet["G2"].value == "2026-01-15 08:30"
    assert sheet["H2"].value in (None, "")
