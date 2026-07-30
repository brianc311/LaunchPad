from io import BytesIO

from openpyxl import load_workbook

from launchpad.fc_cg_summary_export import (
    SUMMARY_FIELDS,
    SUMMARY_HEADERS,
    export_fc_cg_summary_multisite_xlsx,
    export_fc_cg_summary_xlsx,
    sanitize_excel_sheet_name,
)


def test_summary_headers_include_site_host_and_flash_time_and_progress():
    assert SUMMARY_HEADERS[0] == "Site"
    assert SUMMARY_HEADERS[1] == "Host"
    assert SUMMARY_HEADERS == (
        "Site",
        "Host",
        "Name",
        "Status",
        "Flash time",
        "Progress",
        "Maps",
        "Host maps",
        "Size",
        "Policy",
        "Snaps/week",
    )
    assert SUMMARY_FIELDS == (
        "site",
        "host",
        "name",
        "status",
        "flash_time",
        "progress_pct",
        "fc_map_count",
        "host_map_count",
        "total_size",
        "policy",
        "snaps_per_week",
    )


def test_export_fc_cg_summary_xlsx_sheet_headers_and_rows():
    rows = [
        {
            "site": "Anderson",
            "host": "10.0.0.1",
            "name": "AAN1_FC",
            "status": "idle_or_copied",
            "flash_time": "2026-07-30 10:00:00",
            "progress_pct": None,
            "fc_map_count": 84,
            "host_map_count": 48,
            "total_size": "5.8 TB",
            "policy": "",
            "snaps_per_week": 0.44,
        },
        {
            "site": "Jupiter",
            "host": "10.0.0.2",
            "name": "COPYING_CG",
            "status": "copying",
            "flash_time": "",
            "progress_pct": 40,
            "fc_map_count": 2,
            "host_map_count": 0,
            "total_size": "10 GB",
            "policy": "weekly",
            "snaps_per_week": 1,
        },
    ]
    body = export_fc_cg_summary_xlsx(rows)
    wb = load_workbook(BytesIO(body))
    assert wb.sheetnames == ["FC CG Summary"]
    ws = wb["FC CG Summary"]
    assert [cell.value for cell in ws[1]] == list(SUMMARY_HEADERS)
    assert ws["A2"].value == "Anderson"
    assert ws["B2"].value == "10.0.0.1"
    assert ws["C2"].value == "AAN1_FC"
    assert ws["E2"].value == "2026-07-30 10:00:00"
    assert ws["F2"].value in ("", None)
    assert ws["G2"].value == 84
    assert ws["I2"].value == "5.8 TB"
    assert ws["K2"].value == 0.44
    assert ws["E3"].value in ("", None)
    assert ws["F3"].value == "40%"


def test_export_multisite_one_sheet_per_site():
    rows = [
        {
            "site": "Anderson",
            "host": "10.0.0.1",
            "name": "CG1",
            "status": "idle_or_copied",
            "progress_pct": 100,
        },
        {
            "site": "Jupiter",
            "host": "10.0.0.2",
            "name": "CG2",
            "status": "copying",
            "progress_pct": 75,
        },
        {
            "site": "Anderson",
            "host": "10.0.0.1",
            "name": "CG3",
            "status": "stopped",
            "progress_pct": 50,
        },
    ]
    body = export_fc_cg_summary_multisite_xlsx(rows)
    wb = load_workbook(BytesIO(body))
    assert wb.sheetnames == ["Anderson", "Jupiter"]
    assert wb["Anderson"]["A2"].value == "Anderson"
    assert wb["Anderson"]["B2"].value == "10.0.0.1"
    assert wb["Anderson"]["C2"].value == "CG1"
    assert wb["Anderson"]["C3"].value == "CG3"
    assert wb["Jupiter"]["A2"].value == "Jupiter"
    assert wb["Jupiter"]["B2"].value == "10.0.0.2"
    assert wb["Jupiter"]["C2"].value == "CG2"


def test_export_multisite_empty_rows_creates_header_only_sheet():
    body = export_fc_cg_summary_multisite_xlsx([])
    wb = load_workbook(BytesIO(body))
    assert wb.sheetnames == ["Summary"]
    ws = wb["Summary"]
    assert [cell.value for cell in ws[1]] == list(SUMMARY_HEADERS)
    assert ws.max_row == 1


def test_sanitize_excel_sheet_name_strips_invalid_chars_and_truncates():
    used: set[str] = set()
    assert sanitize_excel_sheet_name("Site/A[B]", used=used) == "Site_A_B_"
    assert len(sanitize_excel_sheet_name("A" * 40, used=set())) == 31


def test_sanitize_excel_sheet_name_makes_unique_titles():
    used: set[str] = set()
    first = sanitize_excel_sheet_name("Anderson", used=used)
    second = sanitize_excel_sheet_name("Anderson", used=used)
    assert first == "Anderson"
    assert second == "Anderson_2"
    assert len(second) <= 31
