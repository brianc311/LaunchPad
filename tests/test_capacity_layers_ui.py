"""Task 4: Capacity Report toggles + Excel include_pools / show_raw flags."""

from pathlib import Path

from openpyxl import load_workbook

from launchpad.capacity_export import ExportSite, export_storage_capacity_excel_from_sites
from launchpad.capacity_report import CAPACITY_REPORT_HTML
from launchpad.health_server import HealthCard, HealthServer, _HealthHandler


def test_capacity_report_has_pool_and_raw_toggles():
    html = CAPACITY_REPORT_HTML
    assert 'id="show-pools-toggle"' not in html
    assert "Include CPG / pools" not in html
    assert 'id="show-pools-ibm-toggle"' in html
    assert 'id="show-pools-hpe-toggle"' in html
    assert 'id="show-pools-dell-toggle"' in html
    assert "Show IBM pools" in html
    assert "Show HPE CPGs / pools" in html
    assert "Show Dell pools" in html
    assert "launchpad.capacityReport.showPoolsIbm" in html
    assert "launchpad.capacityReport.showPoolsHpe" in html
    assert "launchpad.capacityReport.showPoolsDell" in html
    assert "show-pools-ibm" in html
    assert "data-pool-family" in html
    assert 'id="show-raw-toggle"' in html
    assert "Show raw capacity" in html
    assert "launchpad.capacityReport.showRaw" in html
    assert "hide-raw-capacity" in html
    assert "include_pools=" in html
    assert "show_raw=" in html


def test_capacity_report_refresh_and_export_pass_include_pools():
    html = CAPACITY_REPORT_HTML
    assert "showPoolsToggle" not in html
    assert "`&include_pools=1` +" in html
    assert html.count("`&include_pools=1` +") == 2
    assert "`/api/refresh/${cardId}?focus=capacity&include_pools=1`" in html
    assert "/api/capacity-export" in html
    assert "show_raw=" in html


def _site_with_pools_and_raw() -> ExportSite:
    return ExportSite(
        card_id=1,
        name="ArrayA",
        host="10.0.0.1",
        serial_number="S1",
        category="Remote",
        device_profile="flashsystem",
        capacity_summary={
            "name": "System",
            "used_pct": 27.0,
            "used_bytes": 270,
            "total_bytes": 1000,
        },
        pools=[
            {
                "name": "CPG1",
                "used_pct": 50.0,
                "used_bytes": 500,
                "total_bytes": 1000,
                "free_bytes": 500,
            }
        ],
        error=None,
        raw_capacity_summary={
            "name": "Raw",
            "used_pct": 40.0,
            "used_bytes": 400,
            "total_bytes": 1000,
            "free_bytes": 600,
        },
    )


def _workbook_text(path: Path) -> str:
    wb = load_workbook(path)
    parts: list[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    parts.append(str(cell))
    return "\n".join(parts)


def test_export_from_sites_omits_pools_when_include_pools_false(tmp_path: Path):
    out = tmp_path / "no-pools.xlsx"
    result = export_storage_capacity_excel_from_sites(
        [_site_with_pools_and_raw()],
        out,
        include_monitor_off=True,
        monitor_enabled={1: True},
        include_pools=False,
    )
    assert out.exists()
    assert result.pool_rows_written == 0
    assert result.pool_filled_count == 0
    blob = _workbook_text(out)
    assert "System" in blob or "27.0%" in blob
    assert "CPG1" not in blob


def test_export_from_sites_includes_raw_only_when_show_raw_true(tmp_path: Path):
    site = _site_with_pools_and_raw()
    hidden = tmp_path / "no-raw.xlsx"
    export_storage_capacity_excel_from_sites(
        [site],
        hidden,
        include_monitor_off=True,
        monitor_enabled={1: True},
        show_raw=False,
    )
    assert "40.0%" not in _workbook_text(hidden)

    shown = tmp_path / "with-raw.xlsx"
    export_storage_capacity_excel_from_sites(
        [site],
        shown,
        include_monitor_off=True,
        monitor_enabled={1: True},
        show_raw=True,
    )
    blob = _workbook_text(shown)
    assert "40.0%" in blob
    assert "Raw" in blob


def test_export_capacity_excel_bytes_passes_include_pools(monkeypatch):
    server = HealthServer()
    server.register_card(1, "On", "10.0.0.1", 22, "user", "")
    server.set_monitor_enabled(card_id=1, enabled=True)
    calls: list[dict] = []

    def _spy(card_id: int, *, focus: str = "", include_pools: bool = True) -> HealthCard:
        calls.append({"card_id": card_id, "focus": focus, "include_pools": include_pools})
        return server._cards[card_id]

    monkeypatch.setattr(server, "refresh_card", _spy)

    server.export_capacity_excel_bytes(include_monitor_off=False, include_pools=False)

    assert calls == [{"card_id": 1, "focus": "capacity", "include_pools": False}]


def test_get_capacity_export_route_honors_include_pools_and_show_raw(monkeypatch):
    server = HealthServer()
    server.register_card(1, "On", "10.0.0.1", 22, "user", "")
    server.set_monitor_enabled(card_id=1, enabled=True)
    captured: dict = {}

    def _spy(
        *,
        include_monitor_off: bool = False,
        card_id: int | None = None,
        include_pools: bool = True,
        show_raw: bool = False,
    ):
        captured["include_pools"] = include_pools
        captured["show_raw"] = show_raw
        captured["include_monitor_off"] = include_monitor_off
        captured["card_id"] = card_id
        return b"PK", "Storage_Capacity_Report_test.xlsx"

    monkeypatch.setattr(server, "export_capacity_excel_bytes", _spy)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)

    handler = object.__new__(_HealthHandler)
    handler.path = "/api/capacity-export?include_off=0&include_pools=0&show_raw=1&open=0"
    sent: dict = {}

    def _send_bytes(body, *, content_type, filename, status=200):
        sent["body"] = body
        sent["status"] = status

    def _send_json(data, status=200):
        sent["json"] = data
        sent["status"] = status

    handler._send_bytes = _send_bytes
    handler._send_json = _send_json
    monkeypatch.setattr("launchpad.health_server.get_health_server", lambda: server)

    handler.do_GET()

    assert captured["include_pools"] is False
    assert captured["show_raw"] is True
    assert sent.get("status") == 200
