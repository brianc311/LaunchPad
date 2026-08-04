import io

import pytest
from openpyxl import load_workbook

from launchpad.health_excel_export import (
    HEALTH_SUMMARY_HEADERS,
    HealthExcelSections,
    build_health_summary_workbook,
    build_health_workbook,
    command_summary_text,
    excel_safe_sheet_title,
    filter_health_summary_cards,
    health_summary_row,
    parse_health_excel_sections,
    truncate_excel_cell,
)
from launchpad.health_server import DASHBOARD_HTML, HealthServer, _HealthHandler


def _sample_card(
    card_id: int,
    *,
    name: str | None = None,
    host: str = "10.0.0.1",
    issues: list | None = None,
    command_results: list | None = None,
) -> dict:
    card = {
        "id": card_id,
        "name": name or f"Site-{card_id}",
        "host": host,
        "device_profile": "ibm_flashsystem_5200",
        "model": "IBM FlashSystem 5200",
        "health_issues": issues or [],
    }
    if command_results is not None:
        card["command_results"] = command_results
    return card


def _all_sheet_values(wb) -> list[str]:
    values: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    values.append(str(cell))
    return values


def test_health_summary_row_healthy_when_monitor_on_no_issues():
    card = _sample_card(1)
    row = health_summary_row(card, monitor_enabled={1: True})
    assert row[0] == "Site-1"
    assert row[1] == "10.0.0.1"
    assert row[2] == "IBM FlashSystem 5200"
    assert row[3] == "on"
    assert row[4] == "healthy"
    assert row[5] == 0


def test_health_summary_row_monitoring_off_status():
    card = _sample_card(2, issues=[{"message": "pool full"}])
    row = health_summary_row(card, monitor_enabled={2: False})
    assert row[3] == "off"
    assert row[4] == "monitoring off"
    assert row[5] == 1


def test_health_summary_row_has_issues_when_monitor_on():
    card = _sample_card(3, issues=[{"message": "a"}, {"message": "b"}])
    row = health_summary_row(card, monitor_enabled={3: True})
    assert row[3] == "on"
    assert row[4] == "has issues"
    assert row[5] == 2


def test_filter_health_summary_cards_by_id():
    cards = [_sample_card(1), _sample_card(2), _sample_card(3)]
    filtered = filter_health_summary_cards(cards, card_id=2)
    assert [card["id"] for card in filtered] == [2]


def test_filter_health_summary_cards_none_returns_all():
    cards = [_sample_card(1), _sample_card(2)]
    assert filter_health_summary_cards(cards, card_id=None) == cards


def test_build_health_summary_workbook_is_xlsx_with_summary_sheet():
    cards = [_sample_card(1), _sample_card(2, issues=[{"message": "warn"}])]
    body = build_health_summary_workbook(
        cards,
        monitor_enabled={1: True, 2: True},
    )
    assert body[:2] == b"PK"
    wb = load_workbook(io.BytesIO(body))
    assert "Summary" in wb.sheetnames
    ws = wb["Summary"]
    assert [cell.value for cell in ws[1]] == list(HEALTH_SUMMARY_HEADERS)
    assert ws.max_row == 3


def test_build_health_summary_workbook_single_card_filter():
    cards = filter_health_summary_cards(
        [_sample_card(1), _sample_card(2)],
        card_id=1,
    )
    body = build_health_summary_workbook(cards, monitor_enabled={1: True, 2: False})
    wb = load_workbook(io.BytesIO(body))
    ws = wb["Summary"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "Site-1"


def _register(server: HealthServer, card_id: int, name: str, *, monitor_on: bool) -> None:
    server.register_card(card_id, name, f"10.0.0.{card_id}", 22, "user", "")
    server.set_monitor_enabled(card_id=card_id, enabled=monitor_on)


def _call_health_export_api(monkeypatch, server: HealthServer, query: str = ""):
    handler = object.__new__(_HealthHandler)
    handler.path = f"/api/health-export{query}"
    sent: dict = {}

    def _send_bytes(body, *, content_type, filename, status=200):
        sent["body"] = body
        sent["content_type"] = content_type
        sent["filename"] = filename
        sent["status"] = status

    def _send_json(data, status=200):
        sent["json"] = data
        sent["status"] = status

    handler._send_bytes = _send_bytes
    handler._send_json = _send_json
    monkeypatch.setattr("launchpad.health_server.get_health_server", lambda: server)

    handler.do_GET()
    return sent


def test_get_health_export_route_sends_xlsx_bytes(monkeypatch):
    server = HealthServer()
    _register(server, 1, "Alpha", monitor_on=True)
    _register(server, 2, "Beta", monitor_on=False)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)

    sent = _call_health_export_api(monkeypatch, server, "?open=0")

    assert sent["status"] == 200
    assert sent["body"][:2] == b"PK"
    assert sent["content_type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert sent["filename"].startswith("Health_Summary_")


def test_get_health_export_route_filters_by_card_id(monkeypatch):
    server = HealthServer()
    _register(server, 1, "Alpha", monitor_on=True)
    _register(server, 2, "Beta", monitor_on=True)
    monkeypatch.setattr(server, "sync_from_app", lambda: 0)

    sent = _call_health_export_api(monkeypatch, server, "?card_id=2&open=0")
    wb = load_workbook(io.BytesIO(sent["body"]))
    ws = wb["Summary"]
    assert ws.max_row == 2
    assert ws.cell(row=2, column=1).value == "Beta"


def test_dashboard_has_health_export_excel_button():
    assert 'id="health-excel-btn"' in DASHBOARD_HTML
    assert "Export Excel" in DASHBOARD_HTML
    filter_bar_start = DASHBOARD_HTML.index('<div class="filter-bar no-print">')
    filter_bar_end = DASHBOARD_HTML.index("</div>", filter_bar_start)
    filter_bar = DASHBOARD_HTML[filter_bar_start:filter_bar_end]
    assert 'id="health-excel-btn"' in filter_bar
    assert "health-site-select" in filter_bar or 'id="health-site-select"' in filter_bar


def test_health_handler_declares_health_export_route():
    import inspect

    source = inspect.getsource(_HealthHandler.do_GET)
    assert "/api/health-export" in source


def test_parse_health_excel_sections_defaults():
    sections = parse_health_excel_sections()
    assert sections == HealthExcelSections(
        summary=True,
        issues=True,
        command_summaries=True,
        raw=False,
    )


def test_parse_health_excel_sections_raw_zero():
    sections = parse_health_excel_sections(raw="0")
    assert sections.raw is False
    assert sections.summary is True
    assert sections.issues is True
    assert sections.command_summaries is True


def test_parse_health_excel_sections_coerces_query_values():
    sections = parse_health_excel_sections(
        summary="0",
        issues="1",
        command_summaries="false",
        raw="true",
    )
    assert sections.summary is False
    assert sections.issues is True
    assert sections.command_summaries is False
    assert sections.raw is True


def test_excel_safe_sheet_title_strips_invalid_chars():
    used: set[str] = set()
    title = excel_safe_sheet_title("Site/A[B]:x*y?z\\", used=used)
    assert title == "Site_A_B__x_y_z_"
    assert title in used


def test_excel_safe_sheet_title_truncates_to_max_len():
    used: set[str] = set()
    title = excel_safe_sheet_title("A" * 40, used=used, max_len=31)
    assert len(title) == 31
    assert title in used


def test_excel_safe_sheet_title_disambiguates_collisions():
    used: set[str] = set()
    first = excel_safe_sheet_title("Anderson", used=used)
    second = excel_safe_sheet_title("Anderson", used=used)
    third = excel_safe_sheet_title("Anderson", used=used)
    assert first == "Anderson"
    assert second == "Anderson (2)"
    assert third == "Anderson (3)"
    assert len(second) <= 31
    assert len(third) <= 31


def test_truncate_excel_cell_under_limit_unchanged():
    text = "hello world"
    assert truncate_excel_cell(text) == text


def test_truncate_excel_cell_appends_truncated_marker():
    text = "A" * 50
    result = truncate_excel_cell(text, limit=30)
    assert len(result) == 30
    assert result.endswith("… (truncated)")
    assert result.startswith("A")


def test_command_summary_text_prefers_item_summary():
    item = {
        "label": "Health - Overall",
        "command": "checkhealth",
        "output": "long raw output that should not be summarized",
        "summary": "All systems healthy",
    }
    assert command_summary_text(item) == "All systems healthy"


def test_command_summary_text_falls_back_to_summarize():
    item = {
        "label": "Health - Overall",
        "command": "checkhealth",
        "output": "",
    }
    assert command_summary_text(item) == "no output"


def test_workbook_summary_plus_two_site_sheets():
    cards = [
        _sample_card(1, name="Alpha"),
        _sample_card(2, name="Beta"),
        _sample_card(3, name="Gamma"),
    ]
    body = build_health_workbook(
        cards,
        monitor_enabled={1: True, 2: True, 3: False},
        sections=HealthExcelSections(),
        detail_card_ids=[1, 2],
    )
    wb = load_workbook(io.BytesIO(body))
    assert "Summary" in wb.sheetnames
    assert len([name for name in wb.sheetnames if name != "Summary"]) == 2
    assert wb["Summary"].max_row == 3


def test_workbook_command_summary_without_raw():
    long_output = "RAW_OUTPUT_MARKER_" + ("x" * 200)
    cards = [
        _sample_card(
            1,
            name="Alpha",
            command_results=[
                {
                    "label": "Health - Overall",
                    "command": "checkhealth",
                    "output": long_output,
                    "summary": "All systems healthy",
                }
            ],
        ),
    ]
    body = build_health_workbook(
        cards,
        monitor_enabled={1: True},
        sections=HealthExcelSections(
            summary=False,
            issues=False,
            command_summaries=True,
            raw=False,
        ),
        detail_card_ids=[1],
    )
    wb = load_workbook(io.BytesIO(body))
    values = _all_sheet_values(wb)
    assert any("Summary: All systems healthy" in value for value in values)
    assert not any("RAW_OUTPUT_MARKER_" in value for value in values)
    assert not any(value.startswith("Raw:") for value in values)


def test_workbook_includes_raw_when_flag_on():
    long_output = "RAW_OUTPUT_MARKER_" + ("y" * 200)
    cards = [
        _sample_card(
            1,
            name="Alpha",
            command_results=[
                {
                    "label": "Health - Overall",
                    "command": "checkhealth",
                    "output": long_output,
                }
            ],
        ),
    ]
    body = build_health_workbook(
        cards,
        monitor_enabled={1: True},
        sections=HealthExcelSections(
            summary=False,
            issues=False,
            command_summaries=False,
            raw=True,
        ),
        detail_card_ids=[1],
    )
    wb = load_workbook(io.BytesIO(body))
    values = _all_sheet_values(wb)
    assert any("RAW_OUTPUT_MARKER_" in value for value in values)
    assert any(value.startswith("Raw:") for value in values)


def test_nothing_to_export_raises():
    cards = [_sample_card(1)]
    with pytest.raises(ValueError, match="Nothing to export"):
        build_health_workbook(
            cards,
            monitor_enabled={1: True},
            sections=HealthExcelSections(
                summary=False,
                issues=False,
                command_summaries=False,
                raw=False,
            ),
            detail_card_ids=None,
        )


def test_nothing_to_export_empty_detail_card_ids():
    cards = [_sample_card(1)]
    sections = HealthExcelSections(
        summary=False,
        issues=True,
        command_summaries=False,
        raw=False,
    )
    with pytest.raises(ValueError, match="Nothing to export"):
        build_health_workbook(
            cards,
            monitor_enabled={1: True},
            sections=sections,
            detail_card_ids=[],
        )


def test_nothing_to_export_non_matching_detail_card_ids():
    cards = [_sample_card(1)]
    sections = HealthExcelSections(
        summary=False,
        issues=True,
        command_summaries=False,
        raw=False,
    )
    with pytest.raises(ValueError, match="Nothing to export"):
        build_health_workbook(
            cards,
            monitor_enabled={1: True},
            sections=sections,
            detail_card_ids=[999],
        )


def test_raw_cell_truncates_full_value_including_prefix():
    long_output = "x" * 32767
    cards = [
        _sample_card(
            1,
            name="Alpha",
            command_results=[
                {
                    "label": "Health - Overall",
                    "command": "checkhealth",
                    "output": long_output,
                }
            ],
        ),
    ]
    body = build_health_workbook(
        cards,
        monitor_enabled={1: True},
        sections=HealthExcelSections(
            summary=False,
            issues=False,
            command_summaries=False,
            raw=True,
        ),
        detail_card_ids=[1],
    )
    wb = load_workbook(io.BytesIO(body))
    raw_values = [
        value
        for value in _all_sheet_values(wb)
        if value.startswith("Raw:")
    ]
    assert len(raw_values) == 1
    assert len(raw_values[0]) <= 32767
    assert raw_values[0].endswith("… (truncated)")
