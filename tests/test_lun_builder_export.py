import csv
from io import BytesIO, TextIOWrapper
import zipfile

from openpyxl import load_workbook

from launchpad.lun_builder_data import normalize_build
from launchpad.lun_builder_export import (
    export_lun_build_csv_zip,
    export_lun_build_xlsx,
)


def _sample_build() -> dict:
    build = normalize_build(
        {
            "id": "hartford-ct",
            "name": "Hartford, CT",
            "hosts": [
                {
                    "lpar_name": "pconsps3",
                    "slot": "5",
                    "wwpn1": "c050760c9594000e",
                    "wwpn2": "c050760c9594000f",
                    "managed_system_serial": "78A9F81",
                }
            ],
            "luns": [
                {
                    "purpose": "ora1vg",
                    "count": 2,
                    "size": "100GB",
                    "pool_or_cpg": "P0",
                    "storage_profile": "flashsystem_5200",
                    "host_names": ["pconsps3"],
                    "shared": True,
                    "cluster": "SPS",
                }
            ],
        }
    )
    assert build is not None
    return build


def test_xlsx_has_styled_hosts_and_expanded_lun_sheets():
    workbook = load_workbook(BytesIO(export_lun_build_xlsx(_sample_build())))

    assert workbook.sheetnames == ["Hosts", "LUN Plan", "By System"]
    assert workbook["Hosts"].freeze_panes == "A2"
    assert workbook["Hosts"]["A2"].value == "pconsps3"
    assert workbook["LUN Plan"].max_row == 3
    assert [workbook["LUN Plan"].cell(row, 1).value for row in (2, 3)] == [
        "pcon_sps_ora1vg_1",
        "pcon_sps_ora1vg_2",
    ]
    assert workbook["By System"].max_row == 3


def test_csv_zip_contains_hosts_and_expanded_luns():
    with zipfile.ZipFile(BytesIO(export_lun_build_csv_zip(_sample_build()))) as archive:
        assert set(archive.namelist()) == {"hosts.csv", "luns.csv"}
        with archive.open("hosts.csv") as raw:
            hosts = list(csv.DictReader(TextIOWrapper(raw, encoding="utf-8-sig")))
        with archive.open("luns.csv") as raw:
            luns = list(csv.DictReader(TextIOWrapper(raw, encoding="utf-8-sig")))

    assert hosts[0]["LPAR Name"] == "pconsps3"
    assert [row["Volume Name"] for row in luns] == ["pcon_sps_ora1vg_1", "pcon_sps_ora1vg_2"]
    assert all(row["Storage Profile"] == "flashsystem_5200" for row in luns)
