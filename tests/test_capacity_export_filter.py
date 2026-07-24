from pathlib import Path
from unittest.mock import MagicMock

from openpyxl import load_workbook

from launchpad.capacity_export import (
    ExportSite,
    card_ids_included_for_export,
    export_storage_capacity_excel,
    export_storage_capacity_excel_from_sites,
    filter_capacity_entries_by_card_id,
    keep_inventory_row,
)
from launchpad.database import Card


def _ssh_card(card_id: int, name: str) -> Card:
    return Card(
        id=card_id,
        name=name,
        card_type="ssh",
        host=f"10.0.0.{card_id}",
        port=22,
        serial_number="",
        username="",
        encrypted_password="",
        encrypted_key_passphrase="",
        encrypted_key="",
        url="",
        icon="",
        category="Cat",
        sort_order=0,
        glow_color="",
        key_file_path="",
        device_profile="FlashSystem",
        custom_commands="",
    )


def test_include_off_false_keeps_only_monitor_on():
    ids = card_ids_included_for_export(
        [1, 2, 3],
        include_monitor_off=False,
        monitor_enabled={1: True, 2: False},
    )
    assert ids == frozenset({1})


def test_include_off_true_keeps_all_ids():
    ids = card_ids_included_for_export(
        [1, 2],
        include_monitor_off=True,
        monitor_enabled={1: False, 2: False},
    )
    assert ids == frozenset({1, 2})


def test_missing_monitor_key_treated_as_off():
    ids = card_ids_included_for_export(
        [9],
        include_monitor_off=False,
        monitor_enabled={},
    )
    assert ids == frozenset()


def test_filter_capacity_entries_by_card_id():
    included = frozenset({1, 2, 3})
    assert filter_capacity_entries_by_card_id(included, card_id=None) == included
    assert filter_capacity_entries_by_card_id(included, card_id=2) == frozenset({2})
    assert filter_capacity_entries_by_card_id(included, card_id=9) == frozenset()


def test_keep_inventory_row_rules():
    included = frozenset({5})
    assert keep_inventory_row(
        matched_card_id=5,
        included_card_ids=included,
        include_monitor_off=False,
    )
    assert not keep_inventory_row(
        matched_card_id=6,
        included_card_ids=included,
        include_monitor_off=False,
    )
    assert not keep_inventory_row(
        matched_card_id=None,
        included_card_ids=included,
        include_monitor_off=False,
    )
    assert keep_inventory_row(
        matched_card_id=None,
        included_card_ids=included,
        include_monitor_off=True,
    )


def test_export_skips_monitor_off_cards(monkeypatch, tmp_path: Path):
    entry_on = MagicMock(card_id=1, name="On")
    entry_off = MagicMock(card_id=2, name="Off")
    monkeypatch.setattr(
        "launchpad.capacity_export.build_health_dashboard_entries",
        lambda db, key: [entry_on, entry_off],
    )
    monkeypatch.setattr(
        "launchpad.capacity_export._refresh_entry_capacity",
        lambda entry: ({"name": entry.name, "used_pct": 1, "used_bytes": 1, "total_bytes": 100}, [], None),
    )
    db = MagicMock()
    db.list_cards.return_value = [_ssh_card(1, "On"), _ssh_card(2, "Off")]
    monkeypatch.setattr("launchpad.capacity_export.INVENTORY_ROWS", [])

    out = tmp_path / "cap.xlsx"
    result = export_storage_capacity_excel(
        db,
        b"0" * 32,
        out,
        include_monitor_off=False,
        monitor_enabled={1: True, 2: False},
    )
    assert out.exists()
    assert result.extra_rows == 1


def _export_with_two_cards(monkeypatch, tmp_path: Path, *, include_monitor_off: bool):
    """Shared stubs: two entries/cards, empty refresh payload, caller-supplied filter."""
    entry_on = MagicMock(card_id=1, name="On")
    entry_off = MagicMock(card_id=2, name="Off")
    monkeypatch.setattr(
        "launchpad.capacity_export.build_health_dashboard_entries",
        lambda db, key: [entry_on, entry_off],
    )
    monkeypatch.setattr(
        "launchpad.capacity_export._refresh_entry_capacity",
        lambda entry: ({"name": entry.name, "used_pct": 1, "used_bytes": 1, "total_bytes": 100}, [], None),
    )
    db = MagicMock()
    db.list_cards.return_value = [_ssh_card(1, "On"), _ssh_card(2, "Off")]
    out = tmp_path / "cap.xlsx"
    result = export_storage_capacity_excel(
        db,
        b"0" * 32,
        out,
        include_monitor_off=include_monitor_off,
        monitor_enabled={1: True, 2: False},
    )
    return out, result


def _sheet_cell_values(ws) -> list[str]:
    values: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        for cell in row:
            if cell is not None:
                values.append(str(cell))
    return values


def test_export_omits_monitor_off_inventory_rows_from_workbook(monkeypatch, tmp_path: Path):
    monkeypatch.setattr(
        "launchpad.capacity_export.INVENTORY_ROWS",
        [
            ("Site-A", "on-sn", "10.0.0.1", "On device", "SN1", "IBM"),
            ("Site-B", "off-sn", "10.0.0.2", "Off device", "SN2", "IBM"),
        ],
    )
    out, result = _export_with_two_cards(monkeypatch, tmp_path, include_monitor_off=False)

    assert out.exists()
    assert result.extra_rows == 0

    ws = load_workbook(out, read_only=True).active
    sheet_text = " ".join(_sheet_cell_values(ws))

    assert "10.0.0.1" in sheet_text
    assert "On" in sheet_text
    assert "on-sn" in sheet_text
    assert "10.0.0.2" not in sheet_text
    assert "Off" not in sheet_text
    assert "off-sn" not in sheet_text
    assert ws.max_row == 2  # header + one inventory row


def test_export_refresh_called_only_for_included_entries(monkeypatch, tmp_path: Path):
    refresh_calls: list[int] = []

    def _spy_refresh(entry):
        refresh_calls.append(entry.card_id)
        return (
            {"name": entry.name, "used_pct": 1, "used_bytes": 1, "total_bytes": 100},
            [],
            None,
        )

    entry_on = MagicMock(card_id=1, name="On")
    entry_off = MagicMock(card_id=2, name="Off")
    monkeypatch.setattr(
        "launchpad.capacity_export.build_health_dashboard_entries",
        lambda db, key: [entry_on, entry_off],
    )
    monkeypatch.setattr(
        "launchpad.capacity_export._refresh_entry_capacity",
        _spy_refresh,
    )
    monkeypatch.setattr("launchpad.capacity_export.INVENTORY_ROWS", [])
    db = MagicMock()
    db.list_cards.return_value = [_ssh_card(1, "On"), _ssh_card(2, "Off")]

    export_storage_capacity_excel(
        db,
        b"0" * 32,
        tmp_path / "cap.xlsx",
        include_monitor_off=False,
        monitor_enabled={1: True, 2: False},
    )

    assert refresh_calls == [1]


def test_export_default_includes_both_cards_as_extra_rows(monkeypatch, tmp_path: Path):
    monkeypatch.setattr("launchpad.capacity_export.INVENTORY_ROWS", [])
    out, result = _export_with_two_cards(monkeypatch, tmp_path, include_monitor_off=True)

    assert out.exists()
    assert result.extra_rows == 2


def test_from_sites_respects_monitor_filter(tmp_path: Path):
    sites = [
        ExportSite(
            card_id=1,
            name="A",
            host="10.0.0.1",
            serial_number="S1",
            category="Remote",
            device_profile="flashsystem",
            capacity_summary={"name": "A", "used_pct": 2, "used_bytes": 2, "total_bytes": 100},
            pools=[],
            error=None,
        ),
        ExportSite(
            card_id=2,
            name="B",
            host="10.0.0.2",
            serial_number="S2",
            category="Remote",
            device_profile="flashsystem",
            capacity_summary=None,
            pools=[],
            error="Authentication failed.",
        ),
    ]
    out = tmp_path / "sites.xlsx"
    result = export_storage_capacity_excel_from_sites(
        sites,
        out,
        include_monitor_off=False,
        monitor_enabled={1: True, 2: False},
    )
    assert out.exists()
    assert result.extra_rows + result.filled_count >= 1
    wb = load_workbook(out)
    ws = wb["Storage Capacity"]
    blob = "\n".join(
        str(cell or "") for row in ws.iter_rows(values_only=True) for cell in row
    )
    assert "10.0.0.1" in blob or "A" in blob
    assert "10.0.0.2" not in blob


def test_from_sites_filters_by_card_id(tmp_path: Path):
    sites = [
        ExportSite(
            card_id=1,
            name="A",
            host="10.0.0.1",
            serial_number="S1",
            category="Remote",
            device_profile="flashsystem",
            capacity_summary={"name": "A", "used_pct": 2, "used_bytes": 2, "total_bytes": 100},
            pools=[],
            error=None,
        ),
        ExportSite(
            card_id=2,
            name="B",
            host="10.0.0.2",
            serial_number="S2",
            category="Remote",
            device_profile="flashsystem",
            capacity_summary={"name": "B", "used_pct": 1, "used_bytes": 1, "total_bytes": 50},
            pools=[],
            error=None,
        ),
    ]
    out = tmp_path / "one-site.xlsx"
    export_storage_capacity_excel_from_sites(
        sites,
        out,
        include_monitor_off=True,
        monitor_enabled={1: True, 2: True},
        card_id=2,
    )
    wb = load_workbook(out)
    ws = wb["Storage Capacity"]
    blob = "\n".join(
        str(cell or "") for row in ws.iter_rows(values_only=True) for cell in row
    )
    assert "10.0.0.2" in blob or "B" in blob
    assert "10.0.0.1" not in blob
