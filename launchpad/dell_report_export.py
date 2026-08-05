"""Build Dell Managed Services capacity report workbooks (.xlsx)."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import IconSetRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from launchpad.capacity_export import ExportSite
from launchpad.dell_report_capacity import select_dell_capacity_summary
from launchpad.dell_report_family import dell_report_family
from launchpad.dell_report_identity import resolve_dell_identity
from launchpad.dell_report_leds import UTIL_YELLOW_THRESHOLD
from launchpad.dell_report_snapshots import (
    has_week_snapshot,
    iso_week_key,
    ordered_weeks_for_cards,
    prior_and_current_for_card,
    upsert_week_snapshot,
    weekly_growth_fraction,
)

REPORT_TITLE = "Dell Technologies Managed Services - Capacity Management Report"
HOME_SHEET_NAME = "Home"
IBM_SHEET_NAME = "IBM Report"
HP_SHEET_NAME = "HP Report"
IBM_REPORT_WKLY_SHEET_NAME = "IBM Report - Wkly"
HP_REPORT_WKLY_SHEET_NAME = "HP Report - Wkly"
IBM_FORECAST_SHEET_NAME = "IBM Forecast"
HP_FORECAST_SHEET_NAME = "HP Forecast"
IBM_FORECAST_WKLY_SHEET_NAME = "IBM Forecast - Wkly"
HP_FORECAST_WKLY_SHEET_NAME = "HP Forecast - Wkly"

# Walgreens-style order: vendor Report / Forecast / Forecast - Wkly families.
# Live IBM/HP Report+Forecast(+Wkly) are filled; all others are empty header shells.
ORDERED_SHEET_NAMES: list[str] = [
    "PowerMax Report",
    "PowerMax Forecast",
    "PowerMax Forecast - Wkly",
    "PowerStore Report",
    "PowerStore Forecast",
    "PowerStore Forecast - Wkly",
    "PowerScale Report",
    "PowerScale Forecast",
    "PowerScale Forecast - Wkly",
    "NetApp Report",
    "NetApp Forecast",
    "NetApp Forecast - Wkly",
    IBM_SHEET_NAME,
    IBM_REPORT_WKLY_SHEET_NAME,
    IBM_FORECAST_SHEET_NAME,
    IBM_FORECAST_WKLY_SHEET_NAME,
    HP_SHEET_NAME,
    HP_REPORT_WKLY_SHEET_NAME,
    HP_FORECAST_SHEET_NAME,
    HP_FORECAST_WKLY_SHEET_NAME,
    "Data Domain Report",
    "Data Domain Forecast",
    "Data Domain Forecast - Wkly",
    "Cluster Report",
    "Cluster Forecast",
    "Cluster Forecast - Wkly",
    "Host Report",
    "Datastore Report",
    "ECS Report",
    "ECS Forecast",
    "ECS Forecast - Wkly",
]

_LIVE_SHEET_NAMES = frozenset(
    {
        IBM_SHEET_NAME,
        HP_SHEET_NAME,
        IBM_REPORT_WKLY_SHEET_NAME,
        HP_REPORT_WKLY_SHEET_NAME,
        IBM_FORECAST_SHEET_NAME,
        HP_FORECAST_SHEET_NAME,
        IBM_FORECAST_WKLY_SHEET_NAME,
        HP_FORECAST_WKLY_SHEET_NAME,
    }
)

# Back-compat for tests that import STUB_SHEET_NAMES.
STUB_SHEET_NAMES: list[str] = [
    name for name in ORDERED_SHEET_NAMES if name not in _LIVE_SHEET_NAMES
]

_ASSETS_DIR = Path(__file__).resolve().parent / "assets" / "dell_report"
_DELL_LOGO_FILE = "logo_1.png"
_WALGREENS_LOGO_FILE = "logo_3.png"

# Excel columns: A blank, B Home link, data from C (Facility) through L.
_FIRST_DATA_COL = 3
_HEADER_ROW = 9
_FORECAST_HEADER_ROW = 9
_META_ROW = 7
_DATE_ROW = 8
_BANNER_TITLE_ROW = 3

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
_SHEET_TITLE_FONT = Font(bold=True, size=18, color="5B9BD5")

_DATA_COLUMNS = (
    ("facility", None),
    ("array_name", None),
    ("model", None),
    ("prior_usable_gib", "0.00"),
    ("prior_used_gib", "0.00"),
    ("prior_util", "0.0%"),
    ("curr_usable_gib", "0.00"),
    ("curr_used_gib", "0.00"),
    ("curr_util", "0.0%"),
    ("weekly_growth", "0.0%"),
)

_HEADER_LABELS = (
    "Facility",
    "Storage Array",
    "Model Number",
    "Useable Capacity (GiB)",
    "Used Capacity (GiB)",
    "Utilization % ",
    "Useable Capacity (GiB)",
    "Used Capacity (GiB)",
    "Utilization % ",
    "Weekly Growth %",
)

_FORECAST_HEADER_LABELS = (
    "Facility",
    "Storage Array",
    "Model Number",
    "Date",
    "3 Month",
    "6 Month",
    "9 Month",
    "12 Month",
)

# Utilization % columns (1-based Excel): prior H=8, current K=11.
_UTIL_COLUMNS = (8, 11)
# Forecast util columns: Date F=6 through 12 Month J=10.
_FORECAST_UTIL_COLUMNS = (6, 7, 8, 9, 10)
_GIB = 1024**3


class DellReportEmptyError(ValueError):
    """Raised when no IBM/HP capacity rows after collection."""


def ensure_dell_report_has_rows(ibm_rows: list, hp_rows: list) -> None:
    if not ibm_rows and not hp_rows:
        raise DellReportEmptyError(
            "No Dell Report capacity data for monitored IBM/HPE sites after refresh."
        )


def bytes_to_gib(num_bytes: float) -> float:
    return num_bytes / _GIB


def collect_dell_report_rows(
    sites: list[ExportSite | dict[str, Any]],
    *,
    snapshot_store: dict,
    now: datetime | None = None,
    include_pools: bool = True,
    card_overrides: dict | None = None,
    include_card_ids: set[str] | list | None = None,
) -> tuple[list[dict], list[dict], dict]:
    """Split ibm/hp rows; upsert current week; forced blank when include set."""
    when = _coerce_utc(now)
    week = iso_week_key(when)
    captured_at = when.isoformat()
    store = dict(snapshot_store or {})
    overrides = card_overrides or {}
    include_ids = {str(x) for x in (include_card_ids or [])}
    ibm_rows: list[dict] = []
    hp_rows: list[dict] = []

    for site in sites:
        card_id = _site_value(site, "card_id")
        name = str(_site_value(site, "name") or "")
        device_profile = str(_site_value(site, "device_profile") or "")
        family = dell_report_family(device_profile)
        if family is None or card_id is None:
            continue

        summary = select_dell_capacity_summary(
            capacity_summary=_site_value(site, "capacity_summary"),
            raw_capacity_summary=_site_value(site, "raw_capacity_summary"),
            pools=_site_value(site, "pools") or [],
            include_pools=include_pools,
        )
        total_bytes = float((summary or {}).get("total_bytes") or 0)
        used_bytes = float((summary or {}).get("used_bytes") or 0)
        if not summary or total_bytes <= 0:
            if str(card_id) not in include_ids:
                continue
            ident = resolve_dell_identity(
                card_id=card_id,
                site_name=name,
                device_profile=device_profile,
                summary_name="",
                overrides=overrides,
            )
            blank = {
                "card_id": card_id,
                "facility": ident["facility"],
                "array_name": ident["array_name"],
                "model": ident["model"],
                "prior_usable_gib": None,
                "prior_used_gib": None,
                "prior_util": None,
                "curr_usable_gib": None,
                "curr_used_gib": None,
                "curr_util": None,
                "weekly_growth": None,
            }
            if family == "ibm":
                ibm_rows.append(blank)
            else:
                hp_rows.append(blank)
            continue

        ident = resolve_dell_identity(
            card_id=card_id,
            site_name=name,
            device_profile=device_profile,
            summary_name=str(summary.get("name") or ""),
            overrides=overrides,
        )
        facility = ident["facility"]
        model = ident["model"]
        array_name = ident["array_name"]

        # Always refresh the current ISO week from live capacity so CPG-off
        # raw (and identity) replace a stale same-week CPG / All-CPGs snapshot.
        store = upsert_week_snapshot(
            store,
            card_id=card_id,
            week=week,
            usable_bytes=total_bytes,
            used_bytes=used_bytes,
            model=model,
            facility=facility,
            family=family,
            array_name=array_name,
            captured_at=captured_at,
        )

        prior, current = prior_and_current_for_card(
            store, card_id, current_week=week
        )
        if current is None:
            continue

        row = _row_from_snapshots(prior, current)
        row["card_id"] = card_id
        if family == "ibm":
            ibm_rows.append(row)
        else:
            hp_rows.append(row)

    return ibm_rows, hp_rows, store


def maybe_upsert_dell_snapshot_for_card(
    card: Any,
    *,
    snapshot_store: dict,
    now: datetime | None = None,
    include_pools: bool = True,
    card_overrides: dict | None = None,
) -> dict:
    """Best-effort upsert for the current ISO week when missing."""
    when = _coerce_utc(now)
    week = iso_week_key(when)
    card_id = getattr(card, "card_id", None)
    name = str(getattr(card, "name", "") or "")
    device_profile = str(getattr(card, "device_profile", "") or "")
    family = dell_report_family(device_profile)
    if family is None or card_id is None:
        return snapshot_store

    from launchpad.flashsystem_health import analyze_health

    analysis = analyze_health(
        name,
        getattr(card, "command_results", None),
        getattr(card, "metrics", None),
    )
    summary = select_dell_capacity_summary(
        capacity_summary=analysis.get("capacity_summary"),
        raw_capacity_summary=analysis.get("raw_capacity_summary"),
        pools=analysis.get("pools") or [],
        include_pools=include_pools,
    )
    if not summary:
        return snapshot_store

    total_bytes = float(summary.get("total_bytes") or 0)
    used_bytes = float(summary.get("used_bytes") or 0)
    if total_bytes <= 0:
        return snapshot_store

    if has_week_snapshot(snapshot_store, card_id, week):
        return snapshot_store

    ident = resolve_dell_identity(
        card_id=card_id,
        site_name=name,
        device_profile=device_profile,
        summary_name=str(summary.get("name") or ""),
        overrides=card_overrides or {},
    )
    return upsert_week_snapshot(
        snapshot_store,
        card_id=card_id,
        week=week,
        usable_bytes=total_bytes,
        used_bytes=used_bytes,
        model=ident["model"],
        facility=ident["facility"],
        family=family,
        array_name=ident["array_name"],
        captured_at=when.isoformat(),
    )


def build_dell_report_workbook(
    *,
    ibm_rows: list[dict],
    hp_rows: list[dict],
    report_date: datetime | None = None,
    snapshot_store: dict | None = None,
) -> Workbook:
    when = _coerce_utc(report_date)
    store = snapshot_store or {}
    wb = Workbook()
    _build_home_sheet(wb.active, report_date=when)
    for name in ORDERED_SHEET_NAMES:
        ws = wb.create_sheet(name)
        if name == IBM_SHEET_NAME:
            _build_data_sheet(ws, ibm_rows, report_date=when)
        elif name == HP_SHEET_NAME:
            _build_data_sheet(ws, hp_rows, report_date=when)
        elif name == IBM_REPORT_WKLY_SHEET_NAME:
            _build_report_wkly_sheet(
                ws, ibm_rows, report_date=when, snapshot_store=store
            )
        elif name == HP_REPORT_WKLY_SHEET_NAME:
            _build_report_wkly_sheet(
                ws, hp_rows, report_date=when, snapshot_store=store
            )
        elif name == IBM_FORECAST_SHEET_NAME:
            _build_forecast_sheet(ws, ibm_rows, report_date=when)
        elif name == HP_FORECAST_SHEET_NAME:
            _build_forecast_sheet(ws, hp_rows, report_date=when)
        elif name == IBM_FORECAST_WKLY_SHEET_NAME:
            _build_forecast_wkly_sheet(ws, ibm_rows, report_date=when)
        elif name == HP_FORECAST_WKLY_SHEET_NAME:
            _build_forecast_wkly_sheet(ws, hp_rows, report_date=when)
        elif "Forecast" in name:
            _build_forecast_sheet(ws, [], report_date=when)
        else:
            _build_stub_sheet(ws, report_date=when)
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _site_value(site: ExportSite | dict[str, Any], key: str) -> Any:
    if isinstance(site, dict):
        return site.get(key)
    return getattr(site, key, None)


def _util_fraction(used_bytes: float, total_bytes: float) -> float | None:
    if total_bytes <= 0:
        return None
    return used_bytes / total_bytes


def _row_from_snapshots(prior: dict | None, current: dict) -> dict:
    curr_usable = float(current.get("usable_bytes") or 0)
    curr_used = float(current.get("used_bytes") or 0)
    growth = None
    prior_usable_gib = None
    prior_used_gib = None
    prior_util = None
    if prior is not None:
        prior_usable = float(prior.get("usable_bytes") or 0)
        prior_used = float(prior.get("used_bytes") or 0)
        prior_usable_gib = bytes_to_gib(prior_usable)
        prior_used_gib = bytes_to_gib(prior_used)
        prior_util = _util_fraction(prior_used, prior_usable)
        growth = weekly_growth_fraction(prior_used, curr_used)

    return {
        "facility": current.get("facility") or "",
        "array_name": current.get("array_name") or "",
        "model": current.get("model") or "",
        "prior_usable_gib": prior_usable_gib,
        "prior_used_gib": prior_used_gib,
        "prior_util": prior_util,
        "curr_usable_gib": bytes_to_gib(curr_usable),
        "curr_used_gib": bytes_to_gib(curr_used),
        "curr_util": _util_fraction(curr_used, curr_usable),
        "weekly_growth": growth,
    }


def _coerce_utc(when: datetime | None) -> datetime:
    if when is None:
        return datetime.now(timezone.utc)
    if when.tzinfo is None:
        return when.replace(tzinfo=timezone.utc)
    return when.astimezone(timezone.utc)


def _format_report_date(when: datetime) -> str:
    return when.strftime("%B %d, %Y")


def _prior_week_date(when: datetime) -> str:
    return _format_report_date(when - timedelta(days=7))


def _build_home_sheet(ws: Worksheet, *, report_date: datetime) -> None:
    ws.title = HOME_SHEET_NAME
    _add_logos(ws)
    ws["B8"] = REPORT_TITLE
    ws["B8"].font = _TITLE_FONT
    ws["B9"] = f"Report date: {_format_report_date(report_date)}"
    ws["B11"] = "Sheets"
    ws["B11"].font = Font(bold=True)
    row = 12
    for title in ORDERED_SHEET_NAMES:
        ws.cell(row=row, column=2, value=title)
        row += 1
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 48


_FORECAST_WKLY_HORIZONS = (1, 4, 8, 12)
_FORECAST_WKLY_HEADER_LABELS = (
    "Facility",
    "Storage Array",
    "Model Number",
    "Date",
    "+1 Week",
    "+4 Week",
    "+8 Week",
    "+12 Week",
)
# Facility..Model + Date + four horizons → util cols F–J (6–10) when data starts at C.
_FORECAST_WKLY_UTIL_COLUMNS = (6, 7, 8, 9, 10)


def _build_stub_sheet(ws: Worksheet, *, report_date: datetime) -> None:
    _write_sheet_header(ws, report_date=report_date)


def _iso_week_display_date(week: str) -> str:
    try:
        year_str, week_str = week.split("-W", 1)
        monday = datetime.fromisocalendar(int(year_str), int(week_str), 1)
        return monday.strftime("%B %d, %Y")
    except (ValueError, TypeError):
        return week


def _weeks_for_report_wkly(rows: list[dict], snapshot_store: dict) -> list[str]:
    card_ids = [row.get("card_id") for row in rows if row.get("card_id") is not None]
    weeks = ordered_weeks_for_cards(snapshot_store, card_ids)
    if weeks:
        return weeks
    # No store weeks yet — still show current ISO week header chrome.
    return [iso_week_key()]


def _build_report_wkly_sheet(
    ws: Worksheet,
    rows: list[dict],
    *,
    report_date: datetime,
    snapshot_store: dict,
) -> None:
    weeks = _weeks_for_report_wkly(rows, snapshot_store)
    _add_logos(ws, sheet_title=ws.title)
    ws.cell(row=_META_ROW, column=2, value="Home")
    ws.cell(row=_META_ROW, column=_FIRST_DATA_COL, value="Date")
    ws.cell(row=_META_ROW, column=_FIRST_DATA_COL + 3, value="Values")

    identity_headers = ("Facility", "Storage Array", "Model Number")
    for offset, label in enumerate(identity_headers):
        cell = ws.cell(row=_HEADER_ROW, column=_FIRST_DATA_COL + offset, value=label)
        _style_header_cell(cell)

    util_columns: list[int] = []
    for week_i, week in enumerate(weeks):
        base = _FIRST_DATA_COL + 3 + week_i * 3
        date_cell = ws.cell(row=_DATE_ROW, column=base, value=_iso_week_display_date(week))
        date_cell.alignment = Alignment(horizontal="center")
        for j, label in enumerate(
            ("Useable Capacity (GiB)", "Used Capacity (GiB)", "Utilization % ")
        ):
            cell = ws.cell(row=_HEADER_ROW, column=base + j, value=label)
            _style_header_cell(cell)
        util_columns.append(base + 2)

    ws.row_dimensions[_HEADER_ROW].height = 30
    data_start = _HEADER_ROW + 1
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("facility") or "").lower(),
            str(row.get("array_name") or "").lower(),
        ),
    )
    last_facility: str | None = None
    for offset, row in enumerate(sorted_rows):
        excel_row = data_start + offset
        facility = row.get("facility")
        facility_value = None if facility == last_facility else facility
        if facility != last_facility:
            last_facility = facility
        ws.cell(row=excel_row, column=_FIRST_DATA_COL, value=facility_value)
        ws.cell(row=excel_row, column=_FIRST_DATA_COL + 1, value=row.get("array_name"))
        ws.cell(row=excel_row, column=_FIRST_DATA_COL + 2, value=row.get("model"))
        card_id = row.get("card_id")
        card_weeks = (snapshot_store or {}).get(str(card_id), {}) if card_id is not None else {}
        for week_i, week in enumerate(weeks):
            base = _FIRST_DATA_COL + 3 + week_i * 3
            snap = card_weeks.get(week) if isinstance(card_weeks, dict) else None
            if not isinstance(snap, dict):
                continue
            usable = float(snap.get("usable_bytes") or 0)
            used = float(snap.get("used_bytes") or 0)
            usable_cell = ws.cell(row=excel_row, column=base, value=bytes_to_gib(usable))
            usable_cell.number_format = "0.00"
            used_cell = ws.cell(row=excel_row, column=base + 1, value=bytes_to_gib(used))
            used_cell.number_format = "0.00"
            util = _util_fraction(used, usable)
            util_cell = ws.cell(row=excel_row, column=base + 2, value=util)
            util_cell.number_format = "0.0%"

    if sorted_rows and util_columns:
        end_row = data_start + len(sorted_rows) - 1
        _apply_utilization_icon_leds(
            ws, data_start, end_row, util_columns=tuple(util_columns)
        )


def _project_util(
    curr_util: float | None, weekly_growth: float | None, weeks_ahead: int
) -> float | None:
    if curr_util is None:
        return None
    if weekly_growth is None:
        return float(curr_util)
    projected = float(curr_util) * ((1.0 + float(weekly_growth)) ** weeks_ahead)
    return max(0.0, projected)


def _build_forecast_wkly_sheet(
    ws: Worksheet,
    rows: list[dict],
    *,
    report_date: datetime,
) -> None:
    _add_logos(ws, sheet_title=ws.title)
    ws.cell(row=_META_ROW, column=2, value="Home")
    ws.cell(row=_META_ROW, column=_FIRST_DATA_COL, value="Sum of Utilization %")
    for col, label in enumerate(_FORECAST_WKLY_HEADER_LABELS, start=_FIRST_DATA_COL):
        cell = ws.cell(row=_FORECAST_HEADER_ROW, column=col, value=label)
        _style_header_cell(cell)
    ws.cell(
        row=_DATE_ROW,
        column=_FIRST_DATA_COL + 3,
        value=_format_report_date(report_date),
    )
    ws.row_dimensions[_FORECAST_HEADER_ROW].height = 30
    data_start = _FORECAST_HEADER_ROW + 1
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("facility") or "").lower(),
            str(row.get("array_name") or "").lower(),
        ),
    )
    last_facility: str | None = None
    for offset, row in enumerate(sorted_rows):
        excel_row = data_start + offset
        facility = row.get("facility")
        facility_value = None if facility == last_facility else facility
        if facility != last_facility:
            last_facility = facility
        ws.cell(row=excel_row, column=_FIRST_DATA_COL, value=facility_value)
        ws.cell(row=excel_row, column=_FIRST_DATA_COL + 1, value=row.get("array_name"))
        ws.cell(row=excel_row, column=_FIRST_DATA_COL + 2, value=row.get("model"))
        curr_util = row.get("curr_util")
        growth = row.get("weekly_growth")
        values = [curr_util] + [
            _project_util(curr_util, growth, weeks) for weeks in _FORECAST_WKLY_HORIZONS
        ]
        for col, value in zip(_FORECAST_WKLY_UTIL_COLUMNS, values):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.number_format = "0.0%"
    if sorted_rows:
        end_row = data_start + len(sorted_rows) - 1
        _apply_utilization_icon_leds(
            ws, data_start, end_row, util_columns=_FORECAST_WKLY_UTIL_COLUMNS
        )


def _build_forecast_sheet(
    ws: Worksheet,
    rows: list[dict],
    *,
    report_date: datetime,
) -> None:
    header_row = _write_forecast_sheet_header(ws, report_date=report_date)
    data_start = header_row + 1
    _write_forecast_grouped_rows(ws, rows, data_start=data_start)
    if rows:
        end_row = data_start + len(rows) - 1
        _apply_utilization_icon_leds(
            ws, data_start, end_row, util_columns=_FORECAST_UTIL_COLUMNS
        )


def _build_data_sheet(
    ws: Worksheet,
    rows: list[dict],
    *,
    report_date: datetime,
) -> None:
    header_row = _write_sheet_header(ws, report_date=report_date)
    data_start = header_row + 1
    _write_grouped_facility_rows(ws, rows, data_start=data_start)
    if rows:
        end_row = data_start + len(rows) - 1
        _apply_utilization_icon_leds(ws, data_start, end_row)
        ws.auto_filter.ref = (
            f"{get_column_letter(_FIRST_DATA_COL)}{_HEADER_ROW}:"
            f"{get_column_letter(_FIRST_DATA_COL + len(_DATA_COLUMNS) - 1)}"
            f"{end_row}"
        )


def _scale_logo(img: XLImage, *, max_height: int) -> XLImage:
    if img.height and img.height > max_height:
        scale = max_height / float(img.height)
        img.height = int(img.height * scale)
        img.width = int(img.width * scale)
    return img


def _add_logos(ws: Worksheet, *, sheet_title: str = "") -> None:
    """Dell left, optional centered sheet title, Walgreens right."""
    title = (sheet_title or ws.title or "").strip()
    dell = _ASSETS_DIR / _DELL_LOGO_FILE
    wag = _ASSETS_DIR / _WALGREENS_LOGO_FILE
    if dell.is_file():
        try:
            img = _scale_logo(XLImage(str(dell)), max_height=52)
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            pass
    if title:
        cell = ws.cell(row=_BANNER_TITLE_ROW, column=5, value=title)
        cell.font = _SHEET_TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        try:
            ws.merge_cells(
                start_row=_BANNER_TITLE_ROW,
                start_column=5,
                end_row=_BANNER_TITLE_ROW + 1,
                end_column=8,
            )
        except ValueError:
            pass
    if wag.is_file():
        try:
            img = _scale_logo(XLImage(str(wag)), max_height=48)
            img.anchor = "J1"
            ws.add_image(img)
        except Exception:
            pass


def _style_header_cell(cell) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")


def _write_forecast_grouped_rows(
    ws: Worksheet,
    rows: list[dict],
    *,
    data_start: int,
) -> None:
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("facility") or "").lower(),
            str(row.get("array_name") or "").lower(),
        ),
    )
    last_facility: str | None = None
    for offset, row in enumerate(sorted_rows):
        excel_row = data_start + offset
        facility = row.get("facility")
        if facility == last_facility:
            facility_value = None
        else:
            facility_value = facility
            last_facility = facility
        ws.cell(row=excel_row, column=_FIRST_DATA_COL, value=facility_value)
        ws.cell(row=excel_row, column=_FIRST_DATA_COL + 1, value=row.get("array_name"))
        ws.cell(row=excel_row, column=_FIRST_DATA_COL + 2, value=row.get("model"))
        curr_util = row.get("curr_util")
        for col in _FORECAST_UTIL_COLUMNS:
            cell = ws.cell(row=excel_row, column=col, value=curr_util)
            cell.number_format = "0.0%"


def _write_grouped_facility_rows(
    ws: Worksheet,
    rows: list[dict],
    *,
    data_start: int,
) -> None:
    sorted_rows = sorted(
        rows,
        key=lambda row: (
            str(row.get("facility") or "").lower(),
            str(row.get("array_name") or "").lower(),
        ),
    )
    last_facility: str | None = None
    for offset, row in enumerate(sorted_rows):
        excel_row = data_start + offset
        facility = row.get("facility")
        if facility == last_facility:
            facility_value = None
        else:
            facility_value = facility
            last_facility = facility
        for idx, (key, number_format) in enumerate(_DATA_COLUMNS):
            col = _FIRST_DATA_COL + idx
            value = facility_value if key == "facility" else row.get(key)
            cell = ws.cell(row=excel_row, column=col, value=value)
            if number_format is not None:
                cell.number_format = number_format


def _apply_direct_utilization_fills(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    util_columns: tuple[int, ...] = _UTIL_COLUMNS,
) -> None:
    """Deprecated path kept for callers; prefer icon-set LEDs."""
    _apply_utilization_icon_leds(ws, start_row, end_row, util_columns=util_columns)


def _apply_utilization_icon_leds(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    util_columns: tuple[int, ...] = _UTIL_COLUMNS,
) -> None:
    """Circular traffic-light icons: green below 80%, yellow at/above 80%.

    reverse=True with thresholds [0, 0.8, 1.01] maps util in [0, 1] to green/yellow
    only (red threshold sits above 100%).
    """
    if end_row < start_row:
        return
    for col in util_columns:
        column = get_column_letter(col)
        cell_range = f"{column}{start_row}:{column}{end_row}"
        ws.conditional_formatting.add(
            cell_range,
            IconSetRule(
                "3TrafficLights1",
                "num",
                [0, UTIL_YELLOW_THRESHOLD, 1.01],
                showValue=True,
                percent=False,
                reverse=True,
            ),
        )


def _write_forecast_sheet_header(ws: Worksheet, *, report_date: datetime) -> int:
    _add_logos(ws, sheet_title=ws.title)
    ws.cell(row=_META_ROW, column=2, value="Home")
    ws.cell(row=_META_ROW, column=3, value="Sum of Utilization %")
    ws.cell(row=_META_ROW, column=6, value="Date")
    for col, label in enumerate(_FORECAST_HEADER_LABELS, start=_FIRST_DATA_COL):
        cell = ws.cell(row=_FORECAST_HEADER_ROW, column=col, value=label)
        _style_header_cell(cell)
    ws.cell(row=_DATE_ROW, column=_FIRST_DATA_COL + 3, value=_format_report_date(report_date))
    widths = (4, 10, 22, 24, 20, 14, 12, 12, 12, 12)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[_FORECAST_HEADER_ROW].height = 30
    return _FORECAST_HEADER_ROW


def _write_sheet_header(ws: Worksheet, *, report_date: datetime) -> int:
    _add_logos(ws, sheet_title=ws.title)
    ws.cell(row=_META_ROW, column=2, value="Home")
    ws.cell(row=_META_ROW, column=6, value="Date")
    ws.cell(row=_META_ROW, column=7, value="Values")
    prior = ws.cell(row=_DATE_ROW, column=6, value=_prior_week_date(report_date))
    current = ws.cell(row=_DATE_ROW, column=9, value=_format_report_date(report_date))
    prior.alignment = Alignment(horizontal="center")
    current.alignment = Alignment(horizontal="center")
    for col, label in enumerate(_HEADER_LABELS, start=_FIRST_DATA_COL):
        cell = ws.cell(row=_HEADER_ROW, column=col, value=label)
        _style_header_cell(cell)
    widths = (4, 10, 22, 24, 22, 16, 16, 14, 16, 16, 14, 16)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[_HEADER_ROW].height = 32
    return _HEADER_ROW
