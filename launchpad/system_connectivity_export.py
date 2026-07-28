"""Export System Connectivity scan results to Excel or CSV ZIP."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TOPIC_SHEETS: dict[str, str] = {
    "call_home": "Call Home",
    "dns": "DNS",
    "snmp": "SNMP",
    "ntp": "NTP",
}

TOPIC_CSV_NAMES: dict[str, str] = {
    "call_home": "call_home.csv",
    "dns": "dns.csv",
    "snmp": "snmp.csv",
    "ntp": "ntp.csv",
}

HEADERS: tuple[str, ...] = (
    "Site",
    "Card",
    "Host",
    "Vendor",
    "Profile",
    "Configured",
    "Status",
    "Details",
    "Error",
)

_FIELDS: tuple[str, ...] = (
    "site",
    "card_name",
    "host",
    "vendor",
    "profile",
    "configured",
    "status",
    "details",
    "error",
)

_TOPIC_KEYS = ("call_home", "dns", "snmp", "ntp")


def filter_payload_by_card_id(
    payload: dict[str, Any],
    *,
    card_id: int | None = None,
    card_name: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    """Return topic rows filtered to one card (None = all). Prefer card_id."""
    if card_id is None and not card_name:
        return {
            **{key: list(payload.get(key) or []) for key in _TOPIC_KEYS},
            "errors": list(payload.get("errors") or []),
        }

    def _matches(row: dict[str, Any]) -> bool:
        if card_id is not None and row.get("card_id") is not None:
            try:
                return int(row.get("card_id")) == int(card_id)
            except (TypeError, ValueError):
                return False
        if card_name:
            return str(row.get("card_name") or "").strip() == str(card_name).strip()
        return False

    return {
        **{key: [row for row in payload.get(key) or [] if _matches(row)] for key in _TOPIC_KEYS},
        "errors": [row for row in payload.get("errors") or [] if _matches(row)],
    }


def _rows(items: list[dict[str, Any]], fields: tuple[str, ...]) -> list[tuple[Any, ...]]:
    return [tuple(item.get(field, "") for field in fields) for item in items]


def _write_sheet(
    worksheet,
    headers: tuple[str, ...],
    rows: list[tuple[Any, ...]],
) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for column, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    worksheet.freeze_panes = "A2"
    for index, title in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            12, min(42, len(title) + 2)
        )
    if rows:
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{len(rows) + 1}"


def export_system_connectivity_xlsx(payload: dict[str, Any]) -> bytes:
    """Return a workbook with Call Home, DNS, SNMP, and NTP sheets."""
    workbook = Workbook()
    first = True
    for topic_key, sheet_name in TOPIC_SHEETS.items():
        rows = _rows(list(payload.get(topic_key) or []), _FIELDS)
        if first:
            sheet = workbook.active
            sheet.title = sheet_name
            first = False
        else:
            sheet = workbook.create_sheet(sheet_name)
        _write_sheet(sheet, HEADERS, rows)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _csv_bytes(headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def export_system_connectivity_csv_zip(payload: dict[str, Any]) -> bytes:
    """Return a ZIP containing call_home.csv, dns.csv, snmp.csv, and ntp.csv."""
    output = BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for topic_key, csv_name in TOPIC_CSV_NAMES.items():
            rows = _rows(list(payload.get(topic_key) or []), _FIELDS)
            archive.writestr(csv_name, _csv_bytes(HEADERS, rows))
    return output.getvalue()
