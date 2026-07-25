"""Export Hosts & Volumes Health scan results to Excel or CSV ZIP."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HOST_HEADERS = (
    "Card",
    "Site IP",
    "Vendor",
    "Host name",
    "Status",
)

VOLUME_HEADERS = (
    "Card",
    "Site IP",
    "Vendor",
    "Volume",
    "Pool/CPG",
    "Status",
)

_HOST_FIELDS = (
    "card_name",
    "host",
    "vendor",
    "host_name",
    "status",
)

_VOLUME_FIELDS = (
    "card_name",
    "host",
    "vendor",
    "volume_name",
    "pool_or_cpg",
    "status",
)


def filter_payload_by_card_id(
    payload: dict[str, Any],
    *,
    card_id: int | None = None,
    card_name: str | None = None,
) -> dict[str, list[dict[str, Any]]]:
    """Return hosts/volumes filtered to one card (None = all). Prefer card_id."""
    if card_id is None and not card_name:
        return {
            "hosts": list(payload.get("hosts") or []),
            "volumes": list(payload.get("volumes") or []),
        }

    def _matches(row: dict[str, Any]) -> bool:
        if card_id is not None and row.get("card_id") is not None:
            try:
                return int(row.get("card_id")) == int(card_id)
            except (TypeError, ValueError):
                return False
        if card_name:
            return str(row.get("card_name") or "").strip() == str(card_name).strip()
        return False

    return {
        "hosts": [row for row in payload.get("hosts") or [] if _matches(row)],
        "volumes": [row for row in payload.get("volumes") or [] if _matches(row)],
    }


def _rows(items: list[dict[str, Any]], fields: tuple[str, ...]) -> list[tuple[Any, ...]]:
    return [tuple(item.get(field, "") for field in fields) for item in items]


def _write_sheet(
    worksheet,
    headers: tuple[str, ...],
    rows: list[tuple[Any, ...]],
) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for column, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    worksheet.freeze_panes = "A2"
    for index, title in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            12, min(42, len(title) + 2)
        )
    if rows:
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{len(rows) + 1}"


def export_host_volume_health_xlsx(payload: dict[str, Any]) -> bytes:
    """Return a workbook with Hosts and Volumes sheets."""
    hosts = list(payload.get("hosts") or [])
    volumes = list(payload.get("volumes") or [])
    host_rows = _rows(hosts, _HOST_FIELDS)
    volume_rows = _rows(volumes, _VOLUME_FIELDS)

    workbook = Workbook()
    hosts_sheet = workbook.active
    hosts_sheet.title = "Hosts"
    _write_sheet(hosts_sheet, HOST_HEADERS, host_rows)
    _write_sheet(workbook.create_sheet("Volumes"), VOLUME_HEADERS, volume_rows)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _csv_bytes(headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def export_host_volume_health_csv_zip(payload: dict[str, Any]) -> bytes:
    """Return a ZIP containing hosts.csv and volumes.csv."""
    hosts = list(payload.get("hosts") or [])
    volumes = list(payload.get("volumes") or [])
    output = BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "hosts.csv",
            _csv_bytes(HOST_HEADERS, _rows(hosts, _HOST_FIELDS)),
        )
        archive.writestr(
            "volumes.csv",
            _csv_bytes(VOLUME_HEADERS, _rows(volumes, _VOLUME_FIELDS)),
        )
    return output.getvalue()
