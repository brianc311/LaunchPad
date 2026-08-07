"""Export Site Lookup inventory to Excel or CSV ZIP."""

from __future__ import annotations

import csv
import zipfile
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from launchpad.host_volume_health import status_is_offline_or_degraded

HOST_HEADERS = ("Host", "Status", "Type", "Ports", "Protocol")
VOLUME_HEADERS = ("Volume", "Status", "Capacity", "Pool/CPG", "UID")
POOL_HEADERS = ("Name", "Used %", "Used", "Free", "Total")
CG_HEADERS = ("Name", "Status", "Location", "Volumes", "Maps")
OFFLINE_HEADERS = ("Type", "Name", "Status", "Detail")


def is_hpe_lookup_card(card: dict | None) -> bool:
    profile = str((card or {}).get("device_profile") or "").lower()
    return "hpe" in profile or "3par" in profile or "primera" in profile


def pools_sheet_title(card: dict | None) -> str:
    return "CPGs" if is_hpe_lookup_card(card) else "Pools"


def consistency_groups_sheet_wanted(payload: dict) -> bool:
    if payload.get("consistency_groups_available"):
        return True
    rows = payload.get("consistency_groups") or []
    return isinstance(rows, list) and bool(rows)


def offline_inventory_rows(payload: dict) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for host in payload.get("hosts") or []:
        if not isinstance(host, dict):
            continue
        status = str(host.get("status") or host.get("state") or "")
        if not status_is_offline_or_degraded(status):
            continue
        name = str(host.get("host_name") or host.get("name") or "").strip()
        detail = " · ".join(
            part
            for part in (
                str(host.get("type") or "").strip(),
                (f"{host.get('port_count') or host.get('ports') or ''} ports").strip()
                if (host.get("port_count") or host.get("ports"))
                else "",
            )
            if part
        )
        out.append({"row_type": "host", "name": name, "status": status, "detail": detail})
    for vol in payload.get("volumes") or []:
        if not isinstance(vol, dict):
            continue
        status = str(vol.get("status") or vol.get("state") or "")
        if not status_is_offline_or_degraded(status):
            continue
        name = str(vol.get("name") or vol.get("vdisk_name") or "").strip()
        detail = " · ".join(
            part
            for part in (
                str(vol.get("pool") or vol.get("mdisk_grp_name") or "").strip(),
                str(vol.get("capacity") or "").strip(),
            )
            if part
        )
        out.append({"row_type": "volume", "name": name, "status": status, "detail": detail})
    return out


def _write_sheet(wb: Workbook, title: str, headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(list(headers))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(list(row))


def export_site_lookup_xlsx(payload: dict, *, include_offline: bool = False) -> bytes:
    card = payload.get("card") if isinstance(payload.get("card"), dict) else {}
    wb = Workbook()
    wb.remove(wb.active)
    hosts = [
        (
            h.get("host_name") or h.get("name") or "",
            h.get("status") or h.get("state") or "",
            h.get("type") or h.get("host_type") or "",
            h.get("port_count") or h.get("ports") or "",
            h.get("protocol") or "SCSI",
        )
        for h in (payload.get("hosts") or [])
        if isinstance(h, dict)
    ]
    volumes = [
        (
            v.get("name") or v.get("vdisk_name") or "",
            v.get("status") or v.get("state") or "",
            v.get("capacity") or "",
            v.get("pool") or v.get("mdisk_grp_name") or "",
            v.get("uid") or v.get("vdisk_UID") or "",
        )
        for v in (payload.get("volumes") or [])
        if isinstance(v, dict)
    ]
    pools = [
        (
            p.get("name") or "",
            p.get("used_pct") if p.get("used_pct") is not None else "",
            p.get("used_bytes") if p.get("used_bytes") is not None else "",
            p.get("free_bytes") if p.get("free_bytes") is not None else "",
            p.get("total_bytes") if p.get("total_bytes") is not None else "",
        )
        for p in (payload.get("pools") or [])
        if isinstance(p, dict)
    ]
    _write_sheet(wb, "Hosts", HOST_HEADERS, hosts)
    _write_sheet(wb, "Volumes", VOLUME_HEADERS, volumes)
    _write_sheet(wb, pools_sheet_title(card), POOL_HEADERS, pools)
    if consistency_groups_sheet_wanted(payload):
        cgs = [
            (
                g.get("name") or g.get("id") or "",
                g.get("status") or "",
                g.get("location") or "",
                len(g.get("volumes") or []) if isinstance(g.get("volumes"), list) else "",
                len(g.get("maps") or []) if isinstance(g.get("maps"), list) else "",
            )
            for g in (payload.get("consistency_groups") or [])
            if isinstance(g, dict)
        ]
        _write_sheet(wb, "Consistency Groups", CG_HEADERS, cgs)
    if include_offline:
        offline = [
            (r["row_type"], r["name"], r["status"], r["detail"])
            for r in offline_inventory_rows(payload)
        ]
        _write_sheet(wb, "Offline", OFFLINE_HEADERS, offline)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _csv_bytes(headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> bytes:
    buf = StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8")


def export_site_lookup_csv_zip(payload: dict) -> bytes:
    card = payload.get("card") if isinstance(payload.get("card"), dict) else {}
    hosts = [
        (
            h.get("host_name") or h.get("name") or "",
            h.get("status") or h.get("state") or "",
            h.get("type") or h.get("host_type") or "",
            h.get("port_count") or h.get("ports") or "",
            h.get("protocol") or "SCSI",
        )
        for h in (payload.get("hosts") or [])
        if isinstance(h, dict)
    ]
    volumes = [
        (
            v.get("name") or v.get("vdisk_name") or "",
            v.get("status") or v.get("state") or "",
            v.get("capacity") or "",
            v.get("pool") or v.get("mdisk_grp_name") or "",
            v.get("uid") or v.get("vdisk_UID") or "",
        )
        for v in (payload.get("volumes") or [])
        if isinstance(v, dict)
    ]
    pools = [
        (
            p.get("name") or "",
            p.get("used_pct") if p.get("used_pct") is not None else "",
            p.get("used_bytes") if p.get("used_bytes") is not None else "",
            p.get("free_bytes") if p.get("free_bytes") is not None else "",
            p.get("total_bytes") if p.get("total_bytes") is not None else "",
        )
        for p in (payload.get("pools") or [])
        if isinstance(p, dict)
    ]
    members: list[tuple[str, bytes]] = [
        ("Hosts.csv", _csv_bytes(HOST_HEADERS, hosts)),
        ("Volumes.csv", _csv_bytes(VOLUME_HEADERS, volumes)),
        (f"{pools_sheet_title(card)}.csv", _csv_bytes(POOL_HEADERS, pools)),
    ]
    if consistency_groups_sheet_wanted(payload):
        cgs = [
            (
                g.get("name") or g.get("id") or "",
                g.get("status") or "",
                g.get("location") or "",
                len(g.get("volumes") or []) if isinstance(g.get("volumes"), list) else "",
                len(g.get("maps") or []) if isinstance(g.get("maps"), list) else "",
            )
            for g in (payload.get("consistency_groups") or [])
            if isinstance(g, dict)
        ]
        members.append(("Consistency_Groups.csv", _csv_bytes(CG_HEADERS, cgs)))
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, data in members:
            zf.writestr(name, data)
    return buf.getvalue()
