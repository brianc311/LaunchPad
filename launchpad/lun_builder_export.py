"""Export LUN Builder plans to Excel or a CSV ZIP archive."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from launchpad.lun_builder_data import expand_lun_batch

HOST_HEADERS = (
    "LPAR Name",
    "Slot",
    "State",
    "Required",
    "Type",
    "Remote LPAR",
    "Remote Slot",
    "WWPN 1",
    "WWPN 2",
    "Physical FC Slot",
    "Managed System Name",
    "Managed System Serial",
    "Notes",
)

LUN_HEADERS = (
    "Volume Name",
    "Source Batch",
    "Size",
    "Shared",
    "Storage Profile",
    "Pool / CPG",
    "Host Names",
    "SCSI / LUN ID",
    "Card Hint",
    "Cluster",
)

_HOST_FIELDS = (
    "lpar_name",
    "slot",
    "state",
    "required",
    "type",
    "remote_lpar",
    "remote_slot",
    "wwpn1",
    "wwpn2",
    "physical_fc_slot",
    "managed_system_name",
    "managed_system_serial",
    "notes",
)

_LUN_FIELDS = (
    "name",
    "source_batch",
    "size",
    "shared",
    "storage_profile",
    "pool_or_cpg",
    "host_names",
    "scsi_or_lun_id",
    "card_hint",
    "cluster",
)


def _display_value(field: str, value: Any) -> Any:
    if field in {"required", "shared"}:
        return "Yes" if bool(value) else "No"
    if field == "host_names":
        return "; ".join(str(item) for item in (value or []))
    return value


def _rows(items: list[dict], fields: tuple[str, ...]) -> list[tuple[Any, ...]]:
    return [
        tuple(_display_value(field, item.get(field)) for field in fields)
        for item in items
    ]


def _expanded_luns(build: dict) -> list[dict]:
    expanded: list[dict] = []
    for lun in build.get("luns") or []:
        if isinstance(lun, dict):
            expanded.extend(expand_lun_batch(lun))
    return expanded


def _write_sheet(
    worksheet,
    headers: tuple[str, ...],
    rows: list[tuple[Any, ...]],
) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for column, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    worksheet.freeze_panes = "A2"
    for index, title in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            12, min(42, len(title) + 2)
        )
    if rows:
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{len(rows) + 1}"


def export_lun_build_xlsx(build: dict) -> bytes:
    """Return a styled workbook containing hosts and expanded LUN plan rows."""
    hosts = [host for host in build.get("hosts") or [] if isinstance(host, dict)]
    expanded = _expanded_luns(build)
    host_rows = _rows(hosts, _HOST_FIELDS)
    lun_rows = _rows(expanded, _LUN_FIELDS)
    by_system = sorted(
        expanded,
        key=lambda row: (
            str(row.get("storage_profile") or "").casefold(),
            str(row.get("card_hint") or "").casefold(),
            str(row.get("name") or "").casefold(),
        ),
    )

    workbook = Workbook()
    hosts_sheet = workbook.active
    hosts_sheet.title = "Hosts"
    _write_sheet(hosts_sheet, HOST_HEADERS, host_rows)
    _write_sheet(workbook.create_sheet("LUN Plan"), LUN_HEADERS, lun_rows)
    _write_sheet(
        workbook.create_sheet("By System"),
        LUN_HEADERS,
        _rows(by_system, _LUN_FIELDS),
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _csv_bytes(headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def export_lun_build_csv_zip(build: dict) -> bytes:
    """Return a ZIP containing flat host and expanded LUN CSV files."""
    hosts = [host for host in build.get("hosts") or [] if isinstance(host, dict)]
    expanded = _expanded_luns(build)
    output = BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("hosts.csv", _csv_bytes(HOST_HEADERS, _rows(hosts, _HOST_FIELDS)))
        archive.writestr("luns.csv", _csv_bytes(LUN_HEADERS, _rows(expanded, _LUN_FIELDS)))
    return output.getvalue()
