"""Export FlashCopy CG Status scan results to Excel."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

STATUS_HEADERS: tuple[str, ...] = (
    "Site",
    "Card",
    "Host",
    "CG name",
    "Status",
    "Maps",
    "Flash time",
    "Error",
)

STATUS_FIELDS: tuple[str, ...] = (
    "site",
    "card_name",
    "host",
    "name",
    "status",
    "map_count",
    "flash_time",
    "error",
)

_SHEET_NAME = "FC CG Status"


def filter_status_rows(rows: list[dict], *, bucket: str) -> list[dict]:
    """Return rows for the active Status tab bucket (empty/all = no filter)."""
    normalized = str(bucket or "").strip().lower()
    if normalized in ("", "all"):
        return list(rows)
    return [row for row in rows if str(row.get("bucket") or "") == normalized]


def _rows(items: list[dict[str, Any]], fields: tuple[str, ...]) -> list[tuple[Any, ...]]:
    return [tuple(item.get(field, "") for field in fields) for item in items]


def _write_sheet(
    worksheet,
    headers: tuple[str, ...],
    rows: list[tuple[Any, ...]],
) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for column, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    worksheet.freeze_panes = "A2"
    for index, title in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            12, min(42, len(title) + 2)
        )
    if rows:
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{len(rows) + 1}"


def export_fc_consistgrp_status_xlsx(rows: list[dict]) -> bytes:
    """Return a workbook with one FC CG Status sheet."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET_NAME
    _write_sheet(sheet, STATUS_HEADERS, _rows(rows, STATUS_FIELDS))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
