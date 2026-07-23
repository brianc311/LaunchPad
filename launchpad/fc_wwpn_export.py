"""Export FC WWPN / host / LUN mapping inventory to Excel."""

from __future__ import annotations

import csv
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from launchpad.database import Database
from launchpad.flashsystem_fc import analyze_fc_inventory
from launchpad.monitor import build_health_dashboard_entries
from launchpad.snapshot_schedule_export import filter_cards_by_groups
from launchpad.storage_presets import DEVICE_PROFILES

ProgressCallback = Callable[[str, int, int], None]

DEFAULT_FC_EXPORT_GROUPS = frozenset({"wag1", "wag2", "other"})


def parse_fc_export_groups(query: dict[str, list[str]]) -> set[str]:
    if "groups" not in query:
        return set(DEFAULT_FC_EXPORT_GROUPS)
    return {
        part.strip().lower()
        for raw in query.get("groups") or [""]
        for part in str(raw).split(",")
        if part.strip()
    }


def cards_for_fc_export(
    cards: list[dict[str, Any]],
    groups: set[str] | None,
) -> list[dict[str, Any]]:
    return filter_cards_by_groups(list(cards), groups)


def filter_cards_for_fc_export(
    cards: list[dict[str, Any]],
    *,
    card_id: str | None = None,
    card_name: str | None = None,
) -> list[dict[str, Any]]:
    cid = str(card_id or "").strip()
    if cid:
        return [card for card in cards if str(card.get("id", "")) == cid]
    name = str(card_name or "").strip().lower()
    if name:
        return [
            card
            for card in cards
            if str(card.get("name") or "").strip().lower() == name
        ]
    return list(cards)


MAPPINGS_HOST_HEADERS = (
    "ID",
    "Host",
    "Status",
    "Protocol",
    "WWPN count",
    "Host WWPNs",
)
MAPPINGS_LUN_HEADERS = (
    "Host",
    "Volume / VDisk",
    "SCSI / LUN ID",
    "VDisk ID",
    "Host WWPNs",
)
MAPPINGS_FABRIC_HEADERS = (
    "Node",
    "Local WWPN",
    "Remote WWPN",
    "Host",
    "State",
    "Local port",
)


def mappings_rows_from_card(
    card: dict[str, Any],
) -> tuple[list[tuple[Any, ...]], list[tuple[Any, ...]], list[tuple[Any, ...]]]:
    hosts: list[tuple[Any, ...]] = []
    maps: list[tuple[Any, ...]] = []
    fabric: list[tuple[Any, ...]] = []
    for fc_host in card.get("fc_hosts") or []:
        hosts.append(
            (
                fc_host.get("host_id"),
                fc_host.get("host_name"),
                fc_host.get("status"),
                fc_host.get("protocol"),
                fc_host.get("wwpn_count"),
                fc_host.get("wwpns"),
            )
        )
    for mapping in card.get("fc_mappings") or []:
        maps.append(
            (
                mapping.get("host_name"),
                mapping.get("vdisk_name"),
                mapping.get("scsi_id"),
                mapping.get("vdisk_id"),
                mapping.get("host_wwpns"),
            )
        )
    for login in card.get("fc_fabric") or []:
        fabric.append(
            (
                login.get("node_name"),
                login.get("local_wwpn"),
                login.get("remote_wwpn"),
                login.get("host_name"),
                login.get("state"),
                login.get("local_port"),
            )
        )
    return hosts, maps, fabric


def build_fc_mappings_workbook(
    cards: list[dict[str, Any]],
) -> tuple[Workbook, int, int, int]:
    host_rows: list[tuple[Any, ...]] = []
    map_rows: list[tuple[Any, ...]] = []
    fabric_rows: list[tuple[Any, ...]] = []
    for card in cards:
        hosts, maps, fabric = mappings_rows_from_card(card)
        host_rows.extend(hosts)
        map_rows.extend(maps)
        fabric_rows.extend(fabric)

    wb = Workbook()
    ws_hosts = wb.active
    ws_hosts.title = "Hosts"
    _write_rows(ws_hosts, MAPPINGS_HOST_HEADERS, host_rows)
    ws_maps = wb.create_sheet("LUN Mappings")
    _write_rows(ws_maps, MAPPINGS_LUN_HEADERS, map_rows)
    ws_fabric = wb.create_sheet("Fabric Logins")
    _write_rows(ws_fabric, MAPPINGS_FABRIC_HEADERS, fabric_rows)
    return wb, len(host_rows), len(map_rows), len(fabric_rows)


def _mappings_csv_bytes(headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(headers)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def export_fc_mappings_csv_zip(cards: list[dict[str, Any]]) -> bytes:
    host_rows: list[tuple[Any, ...]] = []
    map_rows: list[tuple[Any, ...]] = []
    fabric_rows: list[tuple[Any, ...]] = []
    for card in cards:
        hosts, maps, fabric = mappings_rows_from_card(card)
        host_rows.extend(hosts)
        map_rows.extend(maps)
        fabric_rows.extend(fabric)
    output = BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "hosts.csv", _mappings_csv_bytes(MAPPINGS_HOST_HEADERS, host_rows)
        )
        archive.writestr(
            "lun_mappings.csv", _mappings_csv_bytes(MAPPINGS_LUN_HEADERS, map_rows)
        )
        archive.writestr(
            "fabric_logins.csv",
            _mappings_csv_bytes(MAPPINGS_FABRIC_HEADERS, fabric_rows),
        )
    return output.getvalue()


@dataclass(frozen=True)
class FcExportResult:
    path: Path
    port_rows: int
    host_rows: int
    map_rows: int
    error_count: int
    generated_at: str


def _style_header(ws, headers: tuple[str, ...]) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
    ws.freeze_panes = "A2"


def _write_rows(ws, headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> None:
    _style_header(ws, headers)
    thin = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for r_index, row in enumerate(rows, start=2):
        for c_index, value in enumerate(row, start=1):
            cell = ws.cell(row=r_index, column=c_index, value=value or None)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [max(12, min(42, len(h) + 2)) for h in headers]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def _refresh_fc(entry) -> tuple[dict[str, Any], str | None]:
    from launchpad.health_server import get_health_server
    from launchpad.monitor import _register_entry

    server = get_health_server()
    server.ensure_running()
    _register_entry(server, entry)
    card = server.refresh_card(entry.card_id)
    return analyze_fc_inventory(card.command_results), card.error


def rows_from_card_api(card: dict[str, Any]) -> tuple[
    list[tuple[Any, ...]],
    list[tuple[Any, ...]],
    list[tuple[Any, ...]],
]:
    """Build port / host / map rows from a health API card payload."""
    site = card.get("category") or ""
    name = card.get("name") or ""
    host = card.get("host") or ""
    model = card.get("model") or card.get("device_profile") or ""
    port_rows: list[tuple[Any, ...]] = []
    host_rows: list[tuple[Any, ...]] = []
    map_rows: list[tuple[Any, ...]] = []

    for port in card.get("fc_ports") or []:
        port_rows.append(
            (
                site,
                name,
                host,
                model,
                port.get("node_name"),
                port.get("port_id") or port.get("fc_io_port_id"),
                port.get("wwpn"),
                port.get("status"),
                port.get("speed"),
                port.get("attachment") or port.get("type"),
                port.get("logged_in_count"),
                port.get("remote_wwpns"),
                port.get("fabric_hosts"),
            )
        )
    for fc_host in card.get("fc_hosts") or []:
        host_rows.append(
            (
                site,
                name,
                host,
                fc_host.get("host_id"),
                fc_host.get("host_name"),
                fc_host.get("status"),
                fc_host.get("protocol"),
                fc_host.get("wwpn_count"),
                fc_host.get("wwpns"),
            )
        )
    for mapping in card.get("fc_mappings") or []:
        map_rows.append(
            (
                site,
                name,
                host,
                mapping.get("host_name"),
                mapping.get("vdisk_name"),
                mapping.get("scsi_id"),
                mapping.get("vdisk_id"),
                mapping.get("host_wwpns"),
            )
        )
    return port_rows, host_rows, map_rows


def build_fc_wwpn_workbook(
    cards: list[dict[str, Any]],
) -> tuple[Workbook, int, int, int]:
    port_rows: list[tuple[Any, ...]] = []
    host_rows: list[tuple[Any, ...]] = []
    map_rows: list[tuple[Any, ...]] = []
    for card in cards:
        ports, hosts, maps = rows_from_card_api(card)
        port_rows.extend(ports)
        host_rows.extend(hosts)
        map_rows.extend(maps)

    wb = Workbook()
    ws_ports = wb.active
    ws_ports.title = "FC Ports WWPN"
    _write_rows(
        ws_ports,
        (
            "Location",
            "Site",
            "IP",
            "Model",
            "Canister / Node",
            "Port",
            "WWPN",
            "Status",
            "Speed",
            "Attachment",
            "Logins",
            "Remote WWPNs",
            "Fabric Hosts",
        ),
        port_rows,
    )

    ws_hosts = wb.create_sheet("Hosts")
    _write_rows(
        ws_hosts,
        (
            "Location",
            "Site",
            "IP",
            "Host ID",
            "Host Name",
            "Status",
            "Protocol",
            "WWPN Count",
            "Host WWPNs",
        ),
        host_rows,
    )

    ws_maps = wb.create_sheet("LUN Mappings")
    _write_rows(
        ws_maps,
        (
            "Location",
            "Site",
            "IP",
            "Host",
            "Volume / VDisk",
            "SCSI / LUN ID",
            "VDisk ID",
            "Host WWPNs",
        ),
        map_rows,
    )
    return wb, len(port_rows), len(host_rows), len(map_rows)


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_fc_wwpn_excel(
    db: Database,
    crypto_key: bytes,
    output_path: Path,
    *,
    progress: ProgressCallback | None = None,
) -> FcExportResult:
    entries = build_health_dashboard_entries(db, crypto_key)
    cards_by_id = {card.id: card for card in db.list_cards() if card.card_type == "ssh"}

    api_cards: list[dict[str, Any]] = []
    error_count = 0
    total = len(entries)

    for index, entry in enumerate(entries, start=1):
        if progress:
            progress(entry.name, index, total)
        card = cards_by_id.get(entry.card_id)
        model = DEVICE_PROFILES.get(entry.device_profile, entry.device_profile or "")
        site = (card.category if card else "") or ""
        try:
            fc, error = _refresh_fc(entry)
            if error and not fc.get("fc_available"):
                error_count += 1
            api_cards.append(
                {
                    "name": entry.name,
                    "host": entry.host,
                    "category": site,
                    "model": model,
                    "device_profile": entry.device_profile,
                    "fc_ports": fc.get("fc_ports") or [],
                    "fc_hosts": fc.get("fc_hosts") or [],
                    "fc_mappings": fc.get("fc_mappings") or [],
                }
            )
        except Exception:
            error_count += 1

    wb, port_count, host_count, map_count = build_fc_wwpn_workbook(api_cards)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return FcExportResult(
        path=output_path,
        port_rows=port_count,
        host_rows=host_count,
        map_rows=map_count,
        error_count=error_count,
        generated_at=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
    )
