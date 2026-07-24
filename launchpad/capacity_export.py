"""Export storage inventory Excel with live SSH capacity and pool stats filled in."""

from __future__ import annotations

import os
import re
import subprocess
import sys
import time
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import AbstractSet, Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from launchpad.database import Card, Database
from launchpad.flashsystem_health import analyze_health, pool_capacity_from_commands
from launchpad.flashsystem_parse import _format_bytes
from launchpad.monitor import HealthDashboardEntry, build_health_dashboard_entries

_IP_RE = re.compile(r"(\d{1,3}(?:\.\d{1,3}){3})")

HEADERS = (
    "Location",
    "Device SN",
    "IP",
    "Device name",
    "SN",
    "Model",
    "Capacity",
    "Pool Stats",
)

POOL_HEADERS = (
    "Location",
    "Device SN",
    "IP",
    "Pool Name",
    "Used %",
    "Used",
    "Total",
    "Free",
)

InventoryRow = tuple[str, str, str, str, str, str]

INVENTORY_ROWS: list[InventoryRow] = [
    ("Data center -WAG1", "DS8884", "172.19.196.29", "DS8884 Type 2833 model 986", "75-GXR40", "IBM"),
    ("Data center -WAG1", "WAG1_XIV_13557", "172.23.13.188", "XIV Gen 3- 314", "7813557", "IBM"),
    ("Data center -WAG1", "V7K237XW-WAG1", "172.19.195.127", "Storwize V7000 G2", "2076-524-78237xw", "IBM"),
    ("Data center -WAG2", "V7K37WP_wag2", "172.18.218.250", "Storwize V7000 G2", "78237WP", "IBM"),
    ("Data center -WAG1", "WAG1_FS9200_1", "172.19.197.122", "IBM FS9200", "2042020E782", "IBM"),
    ("Distribution center", "v5kPEN-g3v1", "10.245.6.251", "FlashSystem 5200", "V5K78F1928", "IBM"),
    ("Data center -WAG1", "wag1_fs9500_c1", "172.19.192.194", "IBM FlashSystem 9500", "78E4F68", "IBM"),
    ("Data center -WAG2", "wag2_fs9500_c1", "172.18.220.213", "IBM FlashSystem 9500", "78E4F71", "IBM"),
    ("Distribution center", "v5kjup-g3v1", "10.244.0.160", "FlashSystem 5200", "78F11W5", "IBM"),
    ("Distribution center", "v5kmtv-g3c1", "10.244.160.208", "FlashSystem 5200", "78F12DB", "IBM"),
    ("Distribution center", "v5kwax-g3v1", "10.244.34.84", "FlashSystem 5200", "78F12DG", "IBM"),
    ("Distribution center", "v5kmor-g3v1", "10.244.128.100", "FlashSystem 5200", "78F1088", "IBM"),
    ("Distribution center", "v7kper-g3v1", "10.244.237.131", "FlashSystem 7200", "78E36FN", "IBM"),
    ("Distribution center", "v7kand-g3v1", "10.244.25.158", "FlashSystem 7200", "78E31NF", "IBM"),
    ("Distribution center", "v5kwin-g3v1", "10.244.60.154", "FlashSystem 5200", "78F1925", "IBM"),
    ("Distribution center", "v7kcon-g3v1", "10.245.16.56", "FlashSystem 7200", "78E37V9", "IBM"),
    ("Distribution center", "v5kwoo-g3c1", "10.244.66.227", "FlashSystem 5200", "78F192B", "IBM"),
    ("Distribution center", "V7val-g3v1", "10.244.210.3", "FlashSystem 7300", "78E3KRZ", "IBM"),
    ("Distribution center", "V7pue-g3v1", "10.244.102.3", "FlashSystem 7300", "78E3KW9", "IBM"),
    ("Distribution center", "V5kNAZ-g3v1", "10.245.32.116", "FlashSystem 5200", "78F1GXV", "IBM"),
    ("Distribution center", "V5kHOU-g3v1", "10.245.8.106", "FlashSystem 5200", "78F1GZM", "IBM"),
    ("Remote", "v7ktmp-g2v1", "10.241.129.163", "Storwize V7000 G3", "7825YWT", "IBM"),
    ("Remote", "dvl_xiv_25037", "10.246.85.1", "XIV 114 / 2812", "7825037", "IBM"),
    ("Data center -WAG1", "WAG1_PVC_FS9500", "172.19.192.127", "IBM FlashSystem 9500", "78E4M8X", "IBM"),
    ("Data center -WAG2", "WAG2_PVC_FS9500", "172.31.194.127", "IBM FlashSystem 9500", "78E4M9K", "IBM"),
    ("Remote", "CPO-V7KRE554", "10.240.57.164", "Storwize V7000", "V7KRE554", "IBM"),
    ("Data center -WAG1", "svcpvcw1", "172.23.28.26", "IBM SAN Volume Controller (2145-SV1)", "78HMTG0", "IBM"),
    ("Data center -WAG2", "svcpvcw2", "172.31.194.69", "IBM SAN Volume Controller (2145-SV1)", "78HLPZ0", "IBM"),
]

ProgressCallback = Callable[[str, int, int], None]
InventoryFill = tuple[str, str]
ExtraRow = tuple[str, str, str, str, str, str, str, str]
PoolDetailRow = tuple[str, str, str, str, float, str, str, str]


@dataclass(frozen=True)
class ExportResult:
    path: Path
    filled_count: int
    pool_filled_count: int
    pool_rows_written: int
    error_count: int
    extra_rows: int
    generated_at: str


@dataclass(frozen=True)
class ExportSite:
    card_id: int
    name: str
    host: str
    serial_number: str
    category: str
    device_profile: str
    capacity_summary: dict[str, Any] | None
    pools: list[dict[str, Any]]
    error: str | None


def card_ids_included_for_export(
    card_ids: Iterable[int],
    *,
    include_monitor_off: bool,
    monitor_enabled: Mapping[int, bool],
) -> frozenset[int]:
    ids = [int(card_id) for card_id in card_ids]
    if include_monitor_off:
        return frozenset(ids)
    return frozenset(
        card_id for card_id in ids if bool(monitor_enabled.get(card_id, False))
    )


def filter_capacity_entries_by_card_id(
    included_card_ids: AbstractSet[int],
    *,
    card_id: int | None = None,
) -> frozenset[int]:
    if card_id is None:
        return frozenset(included_card_ids)
    selected = int(card_id)
    return frozenset({selected}) if selected in included_card_ids else frozenset()


def keep_inventory_row(
    *,
    matched_card_id: int | None,
    included_card_ids: AbstractSet[int],
    include_monitor_off: bool,
) -> bool:
    if include_monitor_off:
        return True
    return matched_card_id is not None and matched_card_id in included_card_ids


def format_capacity_text(
    capacity_summary: dict[str, Any] | None,
    *,
    error: str | None = None,
) -> str:
    if capacity_summary:
        pct = float(capacity_summary.get("used_pct") or 0)
        used = int(capacity_summary.get("used_bytes") or 0)
        total = int(capacity_summary.get("total_bytes") or 0)
        label = capacity_summary.get("name") or "System"
        if total > 0:
            return (
                f"{label}: {pct:.1f}% used "
                f"({_format_bytes(used)} / {_format_bytes(total)})"
            )
        if pct > 0:
            return f"{label}: {pct:.1f}% used"
    if error:
        return f"Error: {error[:160]}"
    return ""


def format_pool_stats_text(pools: list[dict[str, Any]] | None) -> str:
    """One line per pool for the main inventory sheet."""
    if not pools:
        return ""
    lines: list[str] = []
    for pool in pools:
        name = str(pool.get("name") or "Pool").strip()
        used_pct = float(pool.get("used_pct") or 0.0)
        used = int(pool.get("used_bytes") or 0)
        total = int(pool.get("total_bytes") or 0)
        free = int(pool.get("free_bytes") or 0)
        if total > 0:
            lines.append(
                f"{name}: {used_pct:.1f}% used "
                f"({_format_bytes(used)} / {_format_bytes(total)}, {_format_bytes(free)} free)"
            )
        else:
            lines.append(f"{name}: {used_pct:.1f}% used")
    return "\n".join(lines)


def _normalize(value: str) -> str:
    return (value or "").strip().lower()


def _normalize_ip(value: str) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    match = _IP_RE.search(raw)
    return match.group(1).lower() if match else raw.lower()


def _build_card_lookups(
    cards_by_id: dict[int, Card],
) -> tuple[dict[str, int], dict[str, int], dict[str, int]]:
    """Index LaunchPad SSH cards by IP (primary), serial, and device name."""
    by_ip: dict[str, int] = {}
    by_serial: dict[str, int] = {}
    by_name: dict[str, int] = {}
    for card_id, card in cards_by_id.items():
        ip_key = _normalize_ip(card.host)
        if ip_key:
            by_ip[ip_key] = card_id
        serial_key = _normalize(card.serial_number)
        if serial_key:
            by_serial[serial_key] = card_id
        name_key = _normalize(card.name)
        if name_key:
            by_name[name_key] = card_id
    return by_ip, by_serial, by_name


def _build_site_lookups(
    sites_by_id: dict[int, ExportSite],
) -> tuple[dict[str, int], dict[str, int], dict[str, int]]:
    """Index health-report sites by IP (primary), serial, and device name."""
    by_ip: dict[str, int] = {}
    by_serial: dict[str, int] = {}
    by_name: dict[str, int] = {}
    for card_id, site in sites_by_id.items():
        ip_key = _normalize_ip(site.host)
        if ip_key:
            by_ip[ip_key] = card_id
        serial_key = _normalize(site.serial_number)
        if serial_key:
            by_serial[serial_key] = card_id
        name_key = _normalize(site.name)
        if name_key:
            by_name[name_key] = card_id
    return by_ip, by_serial, by_name


def match_inventory_row(
    row: InventoryRow,
    by_ip: dict[str, int],
    by_serial: dict[str, int],
    by_name: dict[str, int],
    *,
    matched_card_ids: set[int],
) -> int | None:
    """Match an inventory row to a LaunchPad card. IP is tried first."""
    _location, device_sn, ip_addr, _device_name, serial, _model = row
    candidates: list[int] = []

    row_ip = _normalize_ip(ip_addr)
    if row_ip and row_ip in by_ip:
        candidates.append(by_ip[row_ip])

    row_serial = _normalize(serial)
    if row_serial and row_serial in by_serial:
        candidates.append(by_serial[row_serial])

    device_key = _normalize(device_sn)
    if device_key and device_key in by_name:
        candidates.append(by_name[device_key])
    elif device_key:
        for name_key, card_id in by_name.items():
            if device_key in name_key or name_key in device_key:
                candidates.append(card_id)
                break
        alt_key = device_key.replace("_", "-")
        for name_key, card_id in by_name.items():
            if alt_key in name_key.replace("_", "-"):
                candidates.append(card_id)
                break

    seen: set[int] = set()
    for card_id in candidates:
        if card_id in seen:
            continue
        seen.add(card_id)
        if card_id not in matched_card_ids:
            return card_id
    return None


def row_matches_card(row: InventoryRow, card: Card) -> bool:
    """Backward-compatible single-card match (IP first, then serial, then device SN)."""
    by_ip, by_serial, by_name = _build_card_lookups({card.id: card})
    return (
        match_inventory_row(row, by_ip, by_serial, by_name, matched_card_ids=set())
        == card.id
    )


def _pool_detail_rows_for_site(
    location: str,
    device_sn: str,
    ip_addr: str,
    pools: list[dict[str, Any]],
) -> list[PoolDetailRow]:
    rows: list[PoolDetailRow] = []
    for pool in pools:
        rows.append(
            (
                location,
                device_sn,
                ip_addr,
                str(pool.get("name") or ""),
                float(pool.get("used_pct") or 0.0),
                _format_bytes(int(pool.get("used_bytes") or 0)),
                _format_bytes(int(pool.get("total_bytes") or 0)),
                _format_bytes(int(pool.get("free_bytes") or 0)),
            )
        )
    return rows


def _styled_workbook(
    inventory_rows: list[tuple[str, ...]],
    inventory_fills: list[InventoryFill],
    extra_rows: list[ExtraRow],
    pool_detail_rows: list[PoolDetailRow],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Storage Capacity"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="D9E8F5")
    capacity_fill = PatternFill("solid", fgColor="FFF2CC")
    error_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="B4C6E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row_idx = 2
    for inv_row, (capacity, pool_stats) in zip(inventory_rows, inventory_fills, strict=True):
        location, device_sn, ip_addr, device_name, serial, model = inv_row
        values = (
            location,
            device_sn,
            ip_addr or None,
            device_name,
            serial,
            model,
            capacity or None,
            pool_stats or None,
        )
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0 and col not in (7, 8):
                cell.fill = alt_fill
            if col in (7, 8):
                cell.fill = error_fill if str(value or "").startswith("Error:") else capacity_fill
        row_idx += 1

    for extra in extra_rows:
        for col, value in enumerate(extra, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value or None)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col in (7, 8):
                cell.fill = error_fill if str(value or "").startswith("Error:") else capacity_fill
        row_idx += 1

    widths = (22, 20, 16, 34, 18, 10, 42, 48)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    last_row = row_idx - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{last_row}"

    ws_pools = wb.create_sheet("Pool Capacity")
    for col, title in enumerate(POOL_HEADERS, start=1):
        cell = ws_pools.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    pool_widths = (22, 20, 16, 24, 10, 14, 14, 14)
    for col, width in enumerate(pool_widths, start=1):
        ws_pools.column_dimensions[get_column_letter(col)].width = width

    for pool_row_index, row in enumerate(pool_detail_rows, start=2):
        for col_index, value in enumerate(row, start=1):
            cell = ws_pools.cell(row=pool_row_index, column=col_index, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_index == 5 and isinstance(value, (int, float)):
                cell.number_format = "0.0"

    if pool_detail_rows:
        ws_pools.freeze_panes = "A2"
        ws_pools.auto_filter.ref = (
            f"A1:{get_column_letter(len(POOL_HEADERS))}{len(pool_detail_rows) + 1}"
        )

    return wb


def _card_to_extra_row(
    card: Card,
    capacity_text: str,
    pool_stats_text: str,
) -> ExtraRow:
    return (
        card.category or card.name,
        card.name,
        card.host,
        card.device_profile or "",
        card.serial_number or "",
        "IBM" if "flashsystem" in (card.device_profile or "").lower() else (card.device_profile or ""),
        capacity_text,
        pool_stats_text,
    )


def _site_to_extra_row(
    site: ExportSite,
    capacity_text: str,
    pool_stats_text: str,
) -> ExtraRow:
    return (
        site.category or site.name,
        site.name,
        site.host,
        site.device_profile or "",
        site.serial_number or "",
        "IBM" if "flashsystem" in (site.device_profile or "").lower() else (site.device_profile or ""),
        capacity_text,
        pool_stats_text,
    )


def _refresh_entry_capacity(
    entry: HealthDashboardEntry,
) -> tuple[dict[str, Any] | None, list[dict[str, Any]], str | None]:
    from launchpad.health_server import get_health_server
    from launchpad.monitor import _register_entry

    server = get_health_server()
    server.ensure_running()
    _register_entry(server, entry)
    card = server.refresh_card(entry.card_id)
    analysis = analyze_health(entry.name, card.command_results, card.metrics)
    pools = pool_capacity_from_commands(card.command_results)
    return analysis.get("capacity_summary"), pools, card.error


def export_storage_capacity_excel(
    db: Database,
    crypto_key: bytes,
    output_path: Path,
    *,
    progress: ProgressCallback | None = None,
    include_monitor_off: bool = True,
    monitor_enabled: Mapping[int, bool] | None = None,
    card_id: int | None = None,
) -> ExportResult:
    entries = build_health_dashboard_entries(db, crypto_key)
    monitor_map = monitor_enabled or {}
    included = card_ids_included_for_export(
        [e.card_id for e in entries],
        include_monitor_off=include_monitor_off,
        monitor_enabled=monitor_map,
    )
    included = filter_capacity_entries_by_card_id(included, card_id=card_id)
    entries = [e for e in entries if e.card_id in included]
    cards_by_id = {card.id: card for card in db.list_cards() if card.card_type == "ssh"}
    by_ip, by_serial, by_name = _build_card_lookups(cards_by_id)

    capacity_by_card_id: dict[int, str] = {}
    pools_by_card_id: dict[int, list[dict[str, Any]]] = {}
    error_count = 0
    total = len(entries)

    for index, entry in enumerate(entries, start=1):
        if progress:
            progress(entry.name, index, total)
        try:
            summary, pools, error = _refresh_entry_capacity(entry)
            text = format_capacity_text(summary, error=error)
            capacity_by_card_id[entry.card_id] = text
            pools_by_card_id[entry.card_id] = pools
            if error and not summary:
                error_count += 1
        except Exception as exc:
            capacity_by_card_id[entry.card_id] = format_capacity_text(None, error=str(exc))
            pools_by_card_id[entry.card_id] = []
            error_count += 1

    matched_card_ids: set[int] = set()
    inventory_rows_exported: list[tuple[str, ...]] = []
    inventory_fills: list[InventoryFill] = []
    pool_detail_rows: list[PoolDetailRow] = []
    filled_count = 0
    pool_filled_count = 0

    for row in INVENTORY_ROWS:
        capacity_text = ""
        pool_stats_text = ""
        location, device_sn, ip_addr, _device_name, _serial, _model = row
        card_id = match_inventory_row(
            row,
            by_ip,
            by_serial,
            by_name,
            matched_card_ids=matched_card_ids,
        )
        if not keep_inventory_row(
            matched_card_id=card_id,
            included_card_ids=included,
            include_monitor_off=include_monitor_off,
        ):
            continue
        if card_id is not None:
            matched_card_ids.add(card_id)
            capacity_text = capacity_by_card_id.get(card_id, "")
            pools = pools_by_card_id.get(card_id, [])
            pool_stats_text = format_pool_stats_text(pools)
            pool_detail_rows.extend(
                _pool_detail_rows_for_site(location, device_sn, ip_addr, pools)
            )
            if capacity_text and not capacity_text.startswith("Error:"):
                filled_count += 1
            if pool_stats_text:
                pool_filled_count += 1
        inventory_rows_exported.append(row)
        inventory_fills.append((capacity_text, pool_stats_text))

    extra_rows: list[ExtraRow] = []
    for entry in entries:
        if entry.card_id in matched_card_ids:
            continue
        card = cards_by_id.get(entry.card_id)
        if not card:
            continue
        capacity_text = capacity_by_card_id.get(entry.card_id, "")
        pools = pools_by_card_id.get(entry.card_id, [])
        pool_stats_text = format_pool_stats_text(pools)
        extra_rows.append(_card_to_extra_row(card, capacity_text, pool_stats_text))
        pool_detail_rows.extend(
            _pool_detail_rows_for_site(
                card.category or card.name,
                card.name,
                card.host,
                pools,
            )
        )
        if capacity_text and not capacity_text.startswith("Error:"):
            filled_count += 1
        if pool_stats_text:
            pool_filled_count += 1

    extra_rows.sort(key=lambda row: (row[0].lower(), row[1].lower()))
    pool_detail_rows.sort(key=lambda row: (row[0].lower(), row[1].lower(), row[3].lower()))

    wb = _styled_workbook(inventory_rows_exported, inventory_fills, extra_rows, pool_detail_rows)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    return ExportResult(
        path=output_path,
        filled_count=filled_count,
        pool_filled_count=pool_filled_count,
        pool_rows_written=len(pool_detail_rows),
        error_count=error_count,
        extra_rows=len(extra_rows),
        generated_at=generated_at,
    )


def export_storage_capacity_excel_from_sites(
    sites: list[ExportSite],
    output_path: Path,
    *,
    include_monitor_off: bool,
    monitor_enabled: Mapping[int, bool],
    card_id: int | None = None,
) -> ExportResult:
    included = card_ids_included_for_export(
        [site.card_id for site in sites],
        include_monitor_off=include_monitor_off,
        monitor_enabled=monitor_enabled,
    )
    included = filter_capacity_entries_by_card_id(included, card_id=card_id)
    sites_by_id = {site.card_id: site for site in sites if site.card_id in included}
    by_ip, by_serial, by_name = _build_site_lookups(sites_by_id)

    capacity_by_card_id: dict[int, str] = {}
    pools_by_card_id: dict[int, list[dict[str, Any]]] = {}
    error_count = 0

    for card_id, site in sites_by_id.items():
        text = format_capacity_text(site.capacity_summary, error=site.error)
        capacity_by_card_id[card_id] = text
        pools_by_card_id[card_id] = site.pools
        if site.error and not site.capacity_summary:
            error_count += 1

    matched_card_ids: set[int] = set()
    inventory_rows_exported: list[tuple[str, ...]] = []
    inventory_fills: list[InventoryFill] = []
    pool_detail_rows: list[PoolDetailRow] = []
    filled_count = 0
    pool_filled_count = 0

    for row in INVENTORY_ROWS:
        capacity_text = ""
        pool_stats_text = ""
        location, device_sn, ip_addr, _device_name, _serial, _model = row
        card_id = match_inventory_row(
            row,
            by_ip,
            by_serial,
            by_name,
            matched_card_ids=matched_card_ids,
        )
        if not keep_inventory_row(
            matched_card_id=card_id,
            included_card_ids=included,
            include_monitor_off=include_monitor_off,
        ):
            continue
        if card_id is not None:
            matched_card_ids.add(card_id)
            capacity_text = capacity_by_card_id.get(card_id, "")
            pools = pools_by_card_id.get(card_id, [])
            pool_stats_text = format_pool_stats_text(pools)
            pool_detail_rows.extend(
                _pool_detail_rows_for_site(location, device_sn, ip_addr, pools)
            )
            if capacity_text and not capacity_text.startswith("Error:"):
                filled_count += 1
            if pool_stats_text:
                pool_filled_count += 1
        inventory_rows_exported.append(row)
        inventory_fills.append((capacity_text, pool_stats_text))

    extra_rows: list[ExtraRow] = []
    for card_id, site in sites_by_id.items():
        if card_id in matched_card_ids:
            continue
        capacity_text = capacity_by_card_id.get(card_id, "")
        pools = pools_by_card_id.get(card_id, [])
        pool_stats_text = format_pool_stats_text(pools)
        extra_rows.append(_site_to_extra_row(site, capacity_text, pool_stats_text))
        pool_detail_rows.extend(
            _pool_detail_rows_for_site(
                site.category or site.name,
                site.name,
                site.host,
                pools,
            )
        )
        if capacity_text and not capacity_text.startswith("Error:"):
            filled_count += 1
        if pool_stats_text:
            pool_filled_count += 1

    extra_rows.sort(key=lambda row: (row[0].lower(), row[1].lower()))
    pool_detail_rows.sort(key=lambda row: (row[0].lower(), row[1].lower(), row[3].lower()))

    wb = _styled_workbook(inventory_rows_exported, inventory_fills, extra_rows, pool_detail_rows)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    return ExportResult(
        path=output_path,
        filled_count=filled_count,
        pool_filled_count=pool_filled_count,
        pool_rows_written=len(pool_detail_rows),
        error_count=error_count,
        extra_rows=len(extra_rows),
        generated_at=generated_at,
    )


def open_exported_workbook(path: Path) -> None:
    """Open the saved workbook with the system default app (Excel)."""
    resolved = Path(path).resolve()
    if not resolved.exists():
        raise FileNotFoundError(resolved)

    time.sleep(0.2)

    if sys.platform == "win32":
        errors: list[str] = []
        try:
            os.startfile(str(resolved))  # type: ignore[attr-defined]
            return
        except OSError as exc:
            errors.append(f"startfile: {exc}")
        try:
            subprocess.Popen(
                ["cmd", "/c", "start", "", str(resolved)],
                close_fds=False,
            )
            return
        except OSError as exc:
            errors.append(f"cmd start: {exc}")
        raise OSError("; ".join(errors))

    if sys.platform == "darwin":
        subprocess.run(["open", str(resolved)], check=True)
        return

    subprocess.run(["xdg-open", str(resolved)], check=True)


def export_blank_inventory(output_path: Path) -> Path:
    """Write inventory template with empty Capacity and Pool Stats columns."""
    empty_fills = [("", "") for _ in INVENTORY_ROWS]
    wb = _styled_workbook(INVENTORY_ROWS, empty_fills, [], [])
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
