"""Export Snapshot Schedule cards to an Excel workbook (.xlsx)."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from launchpad.database import Database
from launchpad.flashsystem_parse import _format_bytes
from launchpad.snapshot_schedule_overrides import (
    format_one_offs_summary,
    normalize_override,
    parse_date_yyyy_mm_dd,
)
from launchpad.storage_presets import DEVICE_PROFILES

ProgressCallback = Callable[[str, int, int], None]

SCHEDULE_HEADERS = (
    "Status",
    "Site",
    "Location",
    "Model",
    "IP",
    "Pool",
    "Used %",
    "Free",
    "Frequency",
    "Interval Days",
    "Starts",
    "Mode",
    "Time",
    "Held",
    "One-offs",
    "Notes",
)


@dataclass(frozen=True)
class SnapshotExportResult:
    path: Path
    scheduled_count: int
    flagged_count: int
    error_count: int
    generated_at: str


def interval_days(used_pct: float, threshold: float) -> int:
    t = max(0.0, min(1.0, used_pct / threshold)) if threshold else 1.0
    return max(2, round(2 + t * 19))


def format_frequency(days: int) -> str:
    if days == 7:
        return "WEEKLY"
    if days == 14:
        return "BIWEEKLY"
    return f"EVERY {days} DAYS"


def pick_capacity(card_api: dict[str, Any]) -> dict[str, Any] | None:
    pools = list(card_api.get("pools") or [])
    if pools:
        pools.sort(key=lambda p: float(p.get("used_pct") or 0), reverse=True)
        pool = pools[0]
        return {
            "used_pct": float(pool.get("used_pct") or 0),
            "free_bytes": int(pool.get("free_bytes") or 0),
            "pool_name": str(pool.get("name") or "Pool"),
        }
    cap = card_api.get("capacity_summary")
    if cap and (cap.get("total_bytes") or cap.get("used_pct") is not None):
        return {
            "used_pct": float(cap.get("used_pct") or 0),
            "free_bytes": int(cap.get("free_bytes") or 0),
            "pool_name": str(cap.get("name") or "System"),
        }
    return None


def is_storage_like(card_api: dict[str, Any]) -> bool:
    profile = str(card_api.get("device_profile") or "").lower()
    if (
        "flashsystem" in profile
        or "storwize" in profile
        or "svc" in profile
        or profile.startswith("ibm_")
    ):
        return True
    return pick_capacity(card_api) is not None


def site_group(card: dict[str, Any]) -> str:
    hay = " ".join(
        str(card.get(key) or "").lower()
        for key in ("name", "category", "host", "model", "device_profile")
    )
    if "wag1" in hay:
        return "wag1"
    if "wag2" in hay:
        return "wag2"
    return "other"


def filter_cards_by_groups(
    cards: list[dict[str, Any]],
    groups: set[str] | None,
) -> list[dict[str, Any]]:
    if groups is None:
        return list(cards)
    allowed = {g.strip().lower() for g in groups if g and str(g).strip()}
    if not allowed or allowed >= {"wag1", "wag2", "other"}:
        if not allowed:
            return []
        if allowed >= {"wag1", "wag2", "other"}:
            return list(cards)
    return [card for card in cards if site_group(card) in allowed]


def build_schedule_rows(
    cards: list[dict[str, Any]],
    notes: dict[str, str],
    *,
    threshold: float = 80.0,
    groups: set[str] | None = None,
    overrides: dict[str, dict[str, Any]] | None = None,
) -> list[tuple[Any, ...]]:
    today = datetime.now().astimezone().replace(hour=0, minute=0, second=0, microsecond=0)
    prepared: list[dict[str, Any]] = []
    for card in filter_cards_by_groups(cards, groups):
        if not is_storage_like(card):
            continue
        cap = pick_capacity(card)
        if not cap:
            continue
        used_pct = float(cap["used_pct"])
        held = used_pct >= threshold
        days = None if held else interval_days(used_pct, threshold)
        prepared.append(
            {
                "card": card,
                "used_pct": used_pct,
                "free_bytes": cap["free_bytes"],
                "pool_name": cap["pool_name"],
                "held": held,
                "days": days,
            }
        )
    prepared.sort(key=lambda item: item["used_pct"])

    rows: list[tuple[Any, ...]] = []
    for index, item in enumerate(prepared, start=1):
        card = item["card"]
        held = item["held"]
        days = item["days"]
        ov_raw = (overrides or {}).get(str(card.get("id")))
        ov = normalize_override(ov_raw) if ov_raw else None
        mode = "auto"
        time_out = ""
        one_offs_text = ""
        custom_start_dt: datetime | None = None
        if ov and ov["held"]:
            held = True
            days = None
            mode = ov["mode"]
        elif ov and ov["mode"] == "custom":
            held = False
            days = int(ov["interval_days"])
            mode = "custom"
            parsed_start = parse_date_yyyy_mm_dd(ov.get("start_date") or "")
            if parsed_start:
                custom_start_dt = datetime(
                    parsed_start.year, parsed_start.month, parsed_start.day
                )
            time_out = str(ov.get("time") or "")
            one_offs_text = format_one_offs_summary(list(ov.get("one_offs") or []))
        elif ov:
            mode = ov["mode"]
        if held or days is None:
            frequency = "HOLD — EXPAND FIRST"
            starts = ""
            days_out: int | str = ""
            status = "Flagged / Hold"
        else:
            frequency = format_frequency(int(days))
            if custom_start_dt is not None:
                start = custom_start_dt
            else:
                start = today + timedelta(days=1 + (index % max(1, int(days))))
            starts = start.strftime("%b %d, %Y")
            days_out = int(days)
            status = "Scheduled"
        held_col = "Yes" if held else "No"
        model = card.get("model") or DEVICE_PROFILES.get(
            card.get("device_profile") or "", card.get("device_profile") or ""
        )
        rows.append(
            (
                status,
                card.get("name") or "",
                card.get("category") or "",
                model,
                card.get("host") or "",
                item["pool_name"],
                round(item["used_pct"], 1),
                _format_bytes(float(item["free_bytes"])),
                frequency,
                days_out,
                starts,
                mode,
                time_out,
                held_col,
                one_offs_text,
                notes.get(str(card.get("id")), "") or "",
            )
        )
    return rows


def _style_sheet(ws, headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    flagged_fill = PatternFill("solid", fgColor="FCE4D6")
    ok_fill = PatternFill("solid", fgColor="E2EFDA")

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin

    for r_index, row in enumerate(rows, start=2):
        status = str(row[0] or "")
        for c_index, value in enumerate(row, start=1):
            cell = ws.cell(row=r_index, column=c_index, value=value if value != "" else None)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c_index == 1:
                cell.fill = flagged_fill if status.startswith("Flagged") else ok_fill

    widths = (16, 20, 18, 28, 16, 18, 10, 12, 22, 12, 14, 10, 10, 8, 24, 40)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def build_snapshot_schedule_workbook(
    cards: list[dict[str, Any]],
    notes: dict[str, str],
    *,
    threshold: float = 80.0,
    groups: set[str] | None = None,
    overrides: dict[str, dict[str, Any]] | None = None,
) -> Workbook:
    rows = build_schedule_rows(
        cards, notes, threshold=threshold, groups=groups, overrides=overrides
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshot Schedule"
    _style_sheet(ws, SCHEDULE_HEADERS, rows)

    ws_meta = wb.create_sheet("Export Info")
    ws_meta["A1"] = "Enough-space threshold (% used)"
    ws_meta["B1"] = threshold
    ws_meta["A2"] = "Generated (UTC)"
    ws_meta["B2"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ws_meta["A3"] = "Scheduled rows"
    ws_meta["B3"] = sum(1 for row in rows if str(row[0]).startswith("Scheduled"))
    ws_meta["A4"] = "Flagged rows"
    ws_meta["B4"] = sum(1 for row in rows if str(row[0]).startswith("Flagged"))
    ws_meta["A5"] = "Site groups"
    if not groups:
        ws_meta["B5"] = "(none)"
    elif {"wag1", "wag2", "other"} <= {g.lower() for g in groups}:
        ws_meta["B5"] = "All"
    else:
        ws_meta["B5"] = ", ".join(sorted(g.upper() if g != "other" else "Other" for g in groups))
    ws_meta.column_dimensions["A"].width = 34
    ws_meta.column_dimensions["B"].width = 28
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _refresh_card_api(entry) -> dict[str, Any]:
    from launchpad.health_server import get_health_server
    from launchpad.monitor import _register_entry

    server = get_health_server()
    server.ensure_running()
    _register_entry(server, entry)
    card = server.refresh_card(entry.card_id)
    return card.to_api()


def export_snapshot_schedule_excel(
    db: Database,
    crypto_key: bytes,
    output_path: Path,
    *,
    threshold: float = 80.0,
    groups: set[str] | None = None,
    overrides: dict[str, dict[str, Any]] | None = None,
    progress: ProgressCallback | None = None,
) -> SnapshotExportResult:
    from launchpad.health_server import get_health_server
    from launchpad.monitor import build_health_dashboard_entries

    entries = build_health_dashboard_entries(db, crypto_key)
    cards: list[dict[str, Any]] = []
    error_count = 0
    total = len(entries)

    for index, entry in enumerate(entries, start=1):
        if progress:
            progress(entry.name, index, total)
        try:
            cards.append(_refresh_card_api(entry))
        except Exception:
            error_count += 1
            cards.append(
                {
                    "id": entry.card_id,
                    "name": entry.name,
                    "host": entry.host,
                    "category": entry.category,
                    "device_profile": entry.device_profile,
                    "model": DEVICE_PROFILES.get(entry.device_profile, entry.device_profile),
                    "pools": [],
                    "capacity_summary": None,
                }
            )

    server = get_health_server()
    notes = server.get_snapshot_notes()
    if overrides is None:
        overrides = server.get_snapshot_overrides()
    wb = build_snapshot_schedule_workbook(
        cards, notes, threshold=threshold, groups=groups, overrides=overrides
    )
    rows = build_schedule_rows(
        cards, notes, threshold=threshold, groups=groups, overrides=overrides
    )
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return SnapshotExportResult(
        path=output_path,
        scheduled_count=sum(1 for row in rows if str(row[0]).startswith("Scheduled")),
        flagged_count=sum(1 for row in rows if str(row[0]).startswith("Flagged")),
        error_count=error_count,
        generated_at=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
    )
