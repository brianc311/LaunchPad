"""Export Health Dashboard summary rows to Excel."""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from launchpad.flashsystem_parse import summarize_command_output

HEALTH_SUMMARY_HEADERS = (
    "Card",
    "Host / Site IP",
    "Profile / Model",
    "Monitor",
    "Status",
    "Issue count",
)

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_DEFAULT_SHEET_TITLE = "Sheet"
_TRUNCATED_SUFFIX = "… (truncated)"


@dataclass(frozen=True)
class HealthExcelSections:
    summary: bool = True
    issues: bool = True
    command_summaries: bool = True
    raw: bool = False


def _coerce_section_flag(value: str | bool) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in ("0", "false"):
        return False
    if text in ("1", "true"):
        return True
    return bool(text)


def parse_health_excel_sections(
    *,
    summary: str | bool = True,
    issues: str | bool = True,
    command_summaries: str | bool = True,
    raw: str | bool = False,
) -> HealthExcelSections:
    """Coerce 0/1/true/false query values; defaults match spec."""
    return HealthExcelSections(
        summary=_coerce_section_flag(summary),
        issues=_coerce_section_flag(issues),
        command_summaries=_coerce_section_flag(command_summaries),
        raw=_coerce_section_flag(raw),
    )


def excel_safe_sheet_title(name: str, *, used: set[str], max_len: int = 31) -> str:
    """Sanitize + truncate + disambiguate into `used`."""
    cleaned = _INVALID_SHEET_CHARS.sub("_", str(name or "").strip())
    cleaned = cleaned[:max_len] or _DEFAULT_SHEET_TITLE

    candidate = cleaned
    suffix_num = 2
    while candidate in used:
        suffix = f" ({suffix_num})"
        base = cleaned[: max_len - len(suffix)]
        candidate = f"{base}{suffix}"
        suffix_num += 1
    used.add(candidate)
    return candidate


def truncate_excel_cell(text: str, *, limit: int = 32767) -> str:
    """Append '… (truncated)' when over limit."""
    if len(text) <= limit:
        return text
    if len(_TRUNCATED_SUFFIX) >= limit:
        return _TRUNCATED_SUFFIX[:limit]
    return text[: limit - len(_TRUNCATED_SUFFIX)] + _TRUNCATED_SUFFIX


def _monitor_on(card_id: int, monitor_enabled: Mapping[int | str, bool]) -> bool:
    if card_id in monitor_enabled:
        return bool(monitor_enabled[card_id])
    return bool(monitor_enabled.get(str(card_id), False))


def derive_health_status(
    card: dict[str, Any],
    *,
    monitor_enabled: Mapping[int | str, bool],
) -> str:
    card_id = int(card.get("id", 0))
    if not _monitor_on(card_id, monitor_enabled):
        return "monitoring off"
    issues = card.get("health_issues") or []
    if issues:
        return "has issues"
    return "healthy"


def profile_or_model(card: dict[str, Any]) -> str:
    model = str(card.get("model") or "").strip()
    if model:
        return model
    return str(card.get("device_profile") or "").strip()


def health_summary_row(
    card: dict[str, Any],
    *,
    monitor_enabled: Mapping[int | str, bool],
) -> tuple[str, str, str, str, str, int]:
    card_id = int(card.get("id", 0))
    monitor_on = _monitor_on(card_id, monitor_enabled)
    issues = card.get("health_issues") or []
    return (
        str(card.get("name") or ""),
        str(card.get("host") or ""),
        profile_or_model(card),
        "on" if monitor_on else "off",
        derive_health_status(card, monitor_enabled=monitor_enabled),
        len(issues),
    )


def filter_health_summary_cards(
    cards: list[dict[str, Any]],
    *,
    card_id: int | None = None,
) -> list[dict[str, Any]]:
    if card_id is None:
        return list(cards)
    return [card for card in cards if int(card.get("id", -1)) == card_id]


def _summary_rows(
    cards: list[dict[str, Any]],
    *,
    monitor_enabled: Mapping[int | str, bool],
) -> list[tuple[str, str, str, str, str, int]]:
    sorted_cards = sorted(
        cards,
        key=lambda card: str(card.get("name") or "").lower(),
    )
    return [
        health_summary_row(card, monitor_enabled=monitor_enabled)
        for card in sorted_cards
    ]


def command_summary_text(item: dict[str, Any]) -> str:
    """Prefer item['summary']; else summarize_command_output(label, command, output)."""
    summary = item.get("summary")
    if summary:
        return str(summary)
    label = str(item.get("label") or "")
    command = str(item.get("command") or "")
    output = str(item.get("output") or "")
    return summarize_command_output(label, command, output)


def _summary_card_list(
    cards: list[dict[str, Any]],
    detail_card_ids: list[int] | None,
) -> list[dict[str, Any]]:
    if detail_card_ids is None:
        return list(cards)
    id_set = {int(card_id) for card_id in detail_card_ids}
    return [card for card in cards if int(card.get("id", -1)) in id_set]


def _detail_sections_enabled(sections: HealthExcelSections) -> bool:
    return sections.issues or sections.command_summaries or sections.raw


def _styled_summary_sheet(
    ws: Worksheet,
    rows: list[tuple[str, str, str, str, str, int]],
) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="D9E8F5")
    thin = Side(style="thin", color="B4C6E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, title in enumerate(HEALTH_SUMMARY_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    widths = (24, 18, 28, 10, 16, 12)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(len(HEALTH_SUMMARY_HEADERS))}{len(rows) + 1}"
        )


def _styled_summary_workbook(rows: list[tuple[str, str, str, str, str, int]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _styled_summary_sheet(ws, rows)
    return wb


def _write_site_sheet(
    ws: Worksheet,
    card: dict[str, Any],
    *,
    monitor_enabled: Mapping[int | str, bool],
    sections: HealthExcelSections,
) -> None:
    row = 1
    ws.cell(row=row, column=1, value=str(card.get("name") or ""))
    row += 1

    card_id = int(card.get("id", 0))
    monitor_on = _monitor_on(card_id, monitor_enabled)
    ws.cell(row=row, column=1, value=f"Host: {card.get('host') or ''}")
    ws.cell(row=row, column=2, value=f"Profile: {profile_or_model(card)}")
    ws.cell(row=row, column=3, value=f"Monitor: {'on' if monitor_on else 'off'}")
    row += 2

    if sections.issues:
        ws.cell(row=row, column=1, value="Issues")
        row += 1
        for col, title in enumerate(("Severity", "Category", "Message"), start=1):
            ws.cell(row=row, column=col, value=title)
        row += 1
        for issue in card.get("health_issues") or []:
            ws.cell(row=row, column=1, value=str(issue.get("severity") or ""))
            ws.cell(row=row, column=2, value=str(issue.get("category") or ""))
            ws.cell(row=row, column=3, value=str(issue.get("message") or ""))
            row += 1

    if sections.command_summaries or sections.raw:
        ws.cell(row=row, column=1, value="Commands")
        row += 1
        for item in card.get("command_results") or []:
            ws.cell(row=row, column=1, value=f"Label: {item.get('label') or ''}")
            row += 1
            ws.cell(row=row, column=1, value=f"Command: {item.get('command') or ''}")
            row += 1
            error = item.get("error")
            if error:
                ws.cell(row=row, column=1, value=f"Error: {error}")
                row += 1
            if sections.command_summaries:
                ws.cell(
                    row=row,
                    column=1,
                    value=f"Summary: {command_summary_text(item)}",
                )
                row += 1
            if sections.raw:
                raw_value = truncate_excel_cell(f"Raw: {item.get('output') or ''}")
                ws.cell(row=row, column=1, value=raw_value)
                row += 1
            row += 1


def build_health_workbook(
    cards: list[dict[str, Any]],
    *,
    monitor_enabled: Mapping[int | str, bool],
    sections: HealthExcelSections,
    detail_card_ids: list[int] | None = None,
) -> bytes:
    write_summary = sections.summary
    write_detail = detail_card_ids is not None and _detail_sections_enabled(sections)
    if not write_summary and not write_detail:
        raise ValueError("Nothing to export")

    detail_cards: list[dict[str, Any]] = []
    if write_detail:
        detail_ids = {int(card_id) for card_id in detail_card_ids or []}
        detail_cards = sorted(
            (card for card in cards if int(card.get("id", -1)) in detail_ids),
            key=lambda card: str(card.get("name") or "").lower(),
        )

    if not write_summary and not detail_cards:
        raise ValueError("Nothing to export")

    wb = Workbook()
    used_titles: set[str] = set()
    default_ws = wb.active

    if write_summary:
        summary_cards = _summary_card_list(cards, detail_card_ids)
        rows = _summary_rows(summary_cards, monitor_enabled=monitor_enabled)
        if write_detail:
            wb.remove(default_ws)
            ws = wb.create_sheet("Summary", 0)
        else:
            ws = default_ws
            ws.title = "Summary"
        used_titles.add("Summary")
        _styled_summary_sheet(ws, rows)
    elif write_detail:
        wb.remove(default_ws)

    if write_detail:
        for card in detail_cards:
            title = excel_safe_sheet_title(str(card.get("name") or ""), used=used_titles)
            ws = wb.create_sheet(title)
            _write_site_sheet(
                ws,
                card,
                monitor_enabled=monitor_enabled,
                sections=sections,
            )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_health_summary_workbook(
    cards: list[dict[str, Any]],
    *,
    monitor_enabled: Mapping[int | str, bool],
) -> bytes:
    return build_health_workbook(
        cards,
        monitor_enabled=monitor_enabled,
        sections=HealthExcelSections(
            summary=True,
            issues=False,
            command_summaries=False,
            raw=False,
        ),
        detail_card_ids=None,
    )
