"""Export Array FlashCopy CG summary rows to Excel."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SUMMARY_HEADERS: tuple[str, ...] = (
    "Site",
    "Name",
    "Status",
    "Flash time",
    "Progress",
    "Maps",
    "Host maps",
    "Size",
    "Policy",
    "Snaps/week",
)

SUMMARY_FIELDS: tuple[str, ...] = (
    "site",
    "name",
    "status",
    "flash_time",
    "progress_pct",
    "fc_map_count",
    "host_map_count",
    "total_size",
    "policy",
    "snaps_per_week",
)

_SHEET_NAME = "FC CG Summary"
_EMPTY_MULTISITE_SHEET = "Summary"
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
_MAX_SHEET_TITLE_LEN = 31


def sanitize_excel_sheet_name(name: str, *, used: set[str]) -> str:
    """Return a valid unique Excel worksheet title (max 31 chars)."""
    cleaned = _INVALID_SHEET_CHARS.sub("_", str(name or "").strip())
    cleaned = cleaned[:_MAX_SHEET_TITLE_LEN] or _EMPTY_MULTISITE_SHEET

    candidate = cleaned
    suffix = 2
    while candidate in used:
        suffix_text = f"_{suffix}"
        base = cleaned[: _MAX_SHEET_TITLE_LEN - len(suffix_text)]
        candidate = f"{base}{suffix_text}"
        suffix += 1
    used.add(candidate)
    return candidate


def _sanitize_sheet_title(site: str) -> str:
    """Sanitize a site name for use as an Excel sheet title."""
    return sanitize_excel_sheet_name(site, used=set())


def _row_site(row: dict[str, Any]) -> str:
    site = str(row.get("site") or row.get("card_name") or row.get("name") or "").strip()
    return site or "Unknown"


def _cell_value(item: dict[str, Any], field: str) -> Any:
    if field == "progress_pct":
        pct = item.get("progress_pct")
        if pct is None or pct == "":
            return ""
        return f"{pct}%"
    return item.get(field, "")


def _rows(items: list[dict[str, Any]], fields: tuple[str, ...]) -> list[tuple[Any, ...]]:
    return [tuple(_cell_value(item, field) for field in fields) for item in items]


def _write_sheet(
    worksheet,
    headers: tuple[str, ...],
    rows: list[tuple[Any, ...]],
) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for column, title in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    for row_index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    worksheet.freeze_panes = "A2"
    for index, title in enumerate(headers, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = max(
            12, min(42, len(title) + 2)
        )
    if rows:
        last_column = get_column_letter(len(headers))
        worksheet.auto_filter.ref = f"A1:{last_column}{len(rows) + 1}"


def export_fc_cg_summary_xlsx(rows: list[dict]) -> bytes:
    """Return a workbook with one FC CG Summary sheet."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET_NAME
    _write_sheet(sheet, SUMMARY_HEADERS, _rows(rows, SUMMARY_FIELDS))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_fc_cg_summary_multisite_xlsx(rows: list[dict]) -> bytes:
    """Return a workbook with one sheet per site (sites sorted A-Z)."""
    by_site: dict[str, list[dict]] = {}
    for row in rows:
        site = _row_site(row)
        enriched = {**row, "site": row.get("site") or site}
        by_site.setdefault(site, []).append(enriched)

    workbook = Workbook()
    used_titles: set[str] = set()

    if not by_site:
        sheet = workbook.active
        sheet.title = _EMPTY_MULTISITE_SHEET
        _write_sheet(sheet, SUMMARY_HEADERS, [])
    else:
        first = True
        for site in sorted(by_site.keys(), key=str.casefold):
            sheet_title = sanitize_excel_sheet_name(site, used=used_titles)
            site_rows = by_site[site]
            if first:
                sheet = workbook.active
                sheet.title = sheet_title
                first = False
            else:
                sheet = workbook.create_sheet(sheet_title)
            _write_sheet(sheet, SUMMARY_HEADERS, _rows(site_rows, SUMMARY_FIELDS))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
