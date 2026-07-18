"""Export contingency group host/volume/map inventory to Excel."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SUMMARY_HEADERS = (
    "ID",
    "Name",
    "Location",
    "Storage Hint",
    "Notes",
    "Updated At",
    "Host Count",
    "Volume Count",
    "Map Count",
)

HOST_HEADERS = (
    "Group ID",
    "Group Name",
    "Host Name",
    "Status",
    "Host Type",
    "Port Count",
    "Protocol",
    "WWPNs",
)

VOLUME_HEADERS = (
    "Group ID",
    "Group Name",
    "Volume Name",
    "Capacity",
    "Pool",
    "UID",
    "Protocol",
    "Role",
    "Source Volume",
)

MAP_HEADERS = (
    "Group ID",
    "Group Name",
    "Volume",
    "Host",
    "SCSI ID",
    "Role",
)


def _wwpns_cell(wwpns: Any) -> str:
    if not isinstance(wwpns, list):
        return ""
    parts = [str(w).strip() for w in wwpns if str(w or "").strip()]
    return ";".join(parts)


def _style_header(ws, headers: tuple[str, ...]) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin
    ws.freeze_panes = "A2"


def _write_rows(ws, headers: tuple[str, ...], rows: list[tuple[Any, ...]]) -> None:
    _style_header(ws, headers)
    thin = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for r_index, row in enumerate(rows, start=2):
        for c_index, value in enumerate(row, start=1):
            cell = ws.cell(row=r_index, column=c_index, value=value or None)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [max(12, min(42, len(h) + 2)) for h in headers]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"


def build_contingency_groups_workbook(groups: list[dict]) -> Workbook:
    summary_rows: list[tuple[Any, ...]] = []
    host_rows: list[tuple[Any, ...]] = []
    volume_rows: list[tuple[Any, ...]] = []
    map_rows: list[tuple[Any, ...]] = []

    for group in groups:
        group_id = str(group.get("id") or "")
        group_name = str(group.get("name") or "")
        hosts = list(group.get("hosts") or [])
        volumes = list(group.get("volumes") or [])
        maps = list(group.get("maps") or [])

        summary_rows.append(
            (
                group_id,
                group_name,
                group.get("location"),
                group.get("storage_hint"),
                group.get("notes"),
                group.get("updated_at"),
                len(hosts),
                len(volumes),
                len(maps),
            )
        )

        for host in hosts:
            if not isinstance(host, dict):
                continue
            host_rows.append(
                (
                    group_id,
                    group_name,
                    host.get("name"),
                    host.get("status"),
                    host.get("host_type"),
                    host.get("port_count"),
                    host.get("protocol"),
                    _wwpns_cell(host.get("wwpns")),
                )
            )

        for vol in volumes:
            if not isinstance(vol, dict):
                continue
            volume_rows.append(
                (
                    group_id,
                    group_name,
                    vol.get("name"),
                    vol.get("capacity"),
                    vol.get("pool"),
                    vol.get("uid"),
                    vol.get("protocol"),
                    vol.get("role") or "source",
                    vol.get("source_volume") or "",
                )
            )

        for mapping in maps:
            if not isinstance(mapping, dict):
                continue
            map_rows.append(
                (
                    group_id,
                    group_name,
                    mapping.get("volume"),
                    mapping.get("host"),
                    mapping.get("scsi_id"),
                    mapping.get("role") or "source",
                )
            )

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_rows(ws_summary, SUMMARY_HEADERS, summary_rows)

    ws_hosts = wb.create_sheet("Hosts")
    _write_rows(ws_hosts, HOST_HEADERS, host_rows)

    ws_volumes = wb.create_sheet("Volumes")
    _write_rows(ws_volumes, VOLUME_HEADERS, volume_rows)

    ws_maps = wb.create_sheet("Maps")
    _write_rows(ws_maps, MAP_HEADERS, map_rows)

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
