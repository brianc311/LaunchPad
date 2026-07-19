"""Parse LUN Builder uploads and map hosts from FC WWPN card data."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import PurePath
import re
from typing import Any
import zipfile

from openpyxl import load_workbook

from launchpad.lun_builder_data import normalize_host_row, normalize_lun_row


_HEADER_RE = re.compile(r"[^a-z0-9]+")
_HOST_ALIASES = {
    "lparname": "lpar_name",
    "hostname": "lpar_name",
    "host": "lpar_name",
    "name": "lpar_name",
    "slot": "slot",
    "state": "state",
    "required": "required",
    "type": "type",
    "remotelpar": "remote_lpar",
    "remoteslot": "remote_slot",
    "wwpn": "wwpn1",
    "wwpn1": "wwpn1",
    "wwpn2": "wwpn2",
    "physicalfcslot": "physical_fc_slot",
    "managedsystemname": "managed_system_name",
    "managedsystemserial": "managed_system_serial",
    "notes": "notes",
}
_LUN_ALIASES = {
    "purpose": "purpose",
    "volumename": "name",
    "name": "name",
    "sourcebatch": "source_batch",
    "count": "count",
    "size": "size",
    "shared": "shared",
    "storageprofile": "storage_profile",
    "poolorcpg": "pool_or_cpg",
    "poolcpg": "pool_or_cpg",
    "hostnames": "host_names",
    "scsiorlunid": "scsi_or_lun_id",
    "scsilunid": "scsi_or_lun_id",
    "cardhint": "card_hint",
    "cluster": "cluster",
    "group": "cluster",
}


def _header_key(value: Any) -> str:
    return _HEADER_RE.sub("", str(value or "").strip().lower())


def _mapped_row(headers: list[Any], values: list[Any], aliases: dict[str, str]) -> dict:
    row: dict[str, Any] = {}
    for header, value in zip(headers, values):
        field = aliases.get(_header_key(header))
        if field:
            row[field] = value
    return row


def _parse_rows(
    rows: list[list[Any]],
    *,
    kind: str,
    warnings: list[str],
) -> tuple[list[dict], list[dict]]:
    if not rows:
        return [], []
    headers = rows[0]
    aliases = _HOST_ALIASES if kind == "hosts" else _LUN_ALIASES
    recognized = {_header_key(header) for header in headers} & set(aliases)
    if not recognized:
        warnings.append(f"No recognized {kind} headers were found.")
        return [], []
    hosts: list[dict] = []
    luns: list[dict] = []
    for row_number, values in enumerate(rows[1:], start=2):
        if not any(value not in (None, "") for value in values):
            continue
        mapped = _mapped_row(headers, values, aliases)
        if kind == "hosts":
            cleaned = normalize_host_row(mapped)
            if cleaned is None:
                warnings.append(f"Hosts row {row_number} has no host name and was skipped.")
                continue
            hosts.append(cleaned)
        else:
            if mapped.get("source_batch"):
                mapped["purpose"] = mapped["source_batch"]
            cleaned = normalize_lun_row(mapped)
            if cleaned is None or not cleaned.get("purpose"):
                warnings.append(f"LUN row {row_number} has no purpose and was skipped.")
                continue
            luns.append(cleaned)
    return hosts, _collapse_expanded_luns(luns)


def _collapse_expanded_luns(luns: list[dict]) -> list[dict]:
    collapsed: list[dict] = []
    positions: dict[tuple[Any, ...], int] = {}
    for lun in luns:
        key = (
            lun.get("purpose"),
            lun.get("size"),
            lun.get("shared"),
            lun.get("storage_profile"),
            lun.get("pool_or_cpg"),
            tuple(lun.get("host_names") or []),
            lun.get("scsi_or_lun_id"),
            lun.get("card_hint"),
            lun.get("cluster"),
        )
        if key in positions:
            collapsed[positions[key]]["count"] += max(1, int(lun.get("count") or 1))
        else:
            positions[key] = len(collapsed)
            collapsed.append(dict(lun))
    return collapsed


def _csv_rows(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    return list(csv.reader(StringIO(text)))


def _csv_kind(filename: str, rows: list[list[Any]]) -> str:
    stem = PurePath(filename).stem.casefold()
    if "host" in stem:
        return "hosts"
    if rows:
        headers = {_header_key(header) for header in rows[0]}
        if headers & {"lparname", "hostname", "wwpn", "wwpn1", "wwpn2"}:
            return "hosts"
    return "luns"


def parse_lun_builder_upload(filename: str, content: bytes) -> dict:
    """Return normalized hosts, LUN specs, and non-fatal import warnings."""
    suffix = PurePath(filename).suffix.casefold()
    warnings: list[str] = []
    hosts: list[dict] = []
    luns: list[dict] = []
    if suffix == ".csv":
        rows = _csv_rows(content)
        parsed_hosts, parsed_luns = _parse_rows(
            rows,
            kind=_csv_kind(filename, rows),
            warnings=warnings,
        )
        hosts.extend(parsed_hosts)
        luns.extend(parsed_luns)
    elif suffix == ".xlsx":
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheets = {_header_key(name): name for name in workbook.sheetnames}
        for key, kind in (("hosts", "hosts"), ("lunplan", "luns")):
            sheet_name = sheets.get(key)
            if not sheet_name:
                warnings.append(f'Workbook has no "{kind.title()}" sheet.')
                continue
            rows = [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
            parsed_hosts, parsed_luns = _parse_rows(
                rows,
                kind=kind,
                warnings=warnings,
            )
            hosts.extend(parsed_hosts)
            luns.extend(parsed_luns)
        workbook.close()
    elif suffix == ".zip":
        with zipfile.ZipFile(BytesIO(content)) as archive:
            csv_names = [
                name for name in archive.namelist() if name.casefold().endswith(".csv")
            ]
            if not csv_names:
                warnings.append("ZIP archive contains no CSV files.")
            for name in csv_names:
                result = parse_lun_builder_upload(name, archive.read(name))
                hosts.extend(result["hosts"])
                luns.extend(result["luns"])
                warnings.extend(f"{name}: {warning}" for warning in result["warnings"])
    else:
        raise ValueError("Upload must be an .xlsx, .csv, or .zip file.")
    return {"hosts": hosts, "luns": luns, "warnings": warnings}


def merge_hosts(existing: list[dict], incoming: list[dict]) -> list[dict]:
    """Append hosts not already identified by lpar_name and wwpn1."""
    merged = [dict(host) for host in existing if isinstance(host, dict)]
    seen = {
        (
            str(host.get("lpar_name") or "").strip().casefold(),
            str(host.get("wwpn1") or "").strip().casefold(),
        )
        for host in merged
    }
    for host in incoming:
        if not isinstance(host, dict):
            continue
        key = (
            str(host.get("lpar_name") or "").strip().casefold(),
            str(host.get("wwpn1") or "").strip().casefold(),
        )
        if key not in seen:
            merged.append(dict(host))
            seen.add(key)
    return merged


def _wwpn_values(raw: Any) -> list[str]:
    values = raw if isinstance(raw, list) else re.split(r"[;,\s]+", str(raw or ""))
    return [str(value).strip() for value in values if str(value).strip()]


def map_fc_hosts(
    cards: list[dict],
    *,
    card_name: str | None = None,
    include_warnings: bool = False,
):
    """Map FC report hosts into LUN Builder host fields."""
    warnings: list[str] = []
    requested = str(card_name or "").strip()
    selected = [
        card
        for card in cards
        if not requested
        or str(card.get("name") or "").strip().casefold() == requested.casefold()
    ]
    if requested and not selected:
        warnings.append(f'FC WWPN card "{requested}" was not found.')
    hosts: list[dict] = []
    for card in selected:
        for raw_host in card.get("fc_hosts") or []:
            if not isinstance(raw_host, dict):
                continue
            name = str(raw_host.get("host_name") or raw_host.get("name") or "").strip()
            if not name:
                continue
            wwpns = _wwpn_values(
                raw_host.get("wwpns")
                or raw_host.get("wwpn")
                or raw_host.get("host_wwpns")
            )
            hosts.append(
                {
                    "lpar_name": name,
                    "wwpn1": wwpns[0] if wwpns else "",
                    "wwpn2": wwpns[1] if len(wwpns) > 1 else "",
                }
            )
    hosts = merge_hosts([], hosts)
    if include_warnings:
        return hosts, warnings
    return hosts
