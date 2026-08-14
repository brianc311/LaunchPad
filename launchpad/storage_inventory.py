"""Storage Inventory helpers — parsers, formatters, issue aggregation."""

from __future__ import annotations

import re
from io import BytesIO
from threading import Lock
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from launchpad.flashsystem_parse import _parse_colon_table
from launchpad.health_alert_state import issue_fingerprint_for_issue, issue_is_visible
from launchpad.storage_presets import HPE_SHELL_PROFILES, SVC_PROFILES, is_svc_fc_profile
from launchpad.system_connectivity import (
    hpe_call_home_na_row,
    is_system_connectivity_eligible,
    parse_ds_networkport_dns,
    parse_ds_showsp_call_home,
    parse_hpe_shownet_dns_ntp,
    parse_svc_call_home,
    parse_svc_dns,
    parse_svc_ntp_from_lssystem,
)

INVENTORY_COLUMNS: tuple[str, ...] = (
    "site",
    "host",
    "ip",
    "model",
    "serial",
    "location",
    "phone_home",
    "data_protection",
    "volume_protection",
    "smtp",
    "issues",
    "card_id",
    "profile",
    "vendor",
)

_INVENTORY_TOPICS: tuple[str, ...] = (
    "identity",
    "call_home",
    "dns",
    "ntp",
    "smtp",
    "data_protection",
)

_DS8884_PROFILE = "ibm_ds8884"


def is_hpe_inventory_profile(profile: str) -> bool:
    key = (profile or "").strip().lower()
    if not key:
        return False
    return key in HPE_SHELL_PROFILES or "3par" in key or "primera" in key


def is_storage_inventory_profile(profile: str) -> bool:
    key = (profile or "").strip().lower()
    if not key:
        return False
    if key in SVC_PROFILES or is_svc_fc_profile(key):
        return True
    if is_hpe_inventory_profile(key):
        return True
    if key == _DS8884_PROFILE or key.startswith("ibm_ds"):
        return True
    return False


def is_storage_inventory_eligible(card: dict) -> bool:
    """SSH FlashSystem / 3PAR / DS8884 cards — monitoring is not required."""
    if str(card.get("card_type") or "").lower() != "ssh":
        return False
    return is_storage_inventory_profile(str(card.get("device_profile") or ""))

_HPE_RCOPY_NOT_CONFIGURED_RE = re.compile(
    r"not\s+configured|no\s+remote\s+copy|disabled|none\s+found",
    re.IGNORECASE,
)
_HPE_RCOPY_TARGET_RE = re.compile(
    r"^\s*(?:Remote\s+Copy\s+)?(?:Group|Target|System)\s*[:=]\s*\S",
    re.IGNORECASE,
)


def _header_index(headers: list[str], *names: str) -> int | None:
    lowered = [h.strip().lower() for h in headers]
    for name in names:
        key = name.lower()
        if key in lowered:
            return lowered.index(key)
    return None


def _cell(row: list[str], idx: int | None) -> str:
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    return str(row[idx] or "").strip()


def inventory_commands_for_profile(profile: str) -> dict[str, list[str]]:
    key = (profile or "").strip().lower()
    empty = {topic: [] for topic in _INVENTORY_TOPICS}
    if key in SVC_PROFILES or is_svc_fc_profile(key):
        return {
            "identity": ["lssystem -delim :"],
            "call_home": ["lscloudcallhome -delim :"],
            "dns": ["lsdnsserver -delim :"],
            "ntp": ["lssystem -delim :"],
            "smtp": ["lsemailserver -delim :"],
            "data_protection": ["lsrcrelationship -delim :"],
        }
    if is_hpe_inventory_profile(key):
        return {
            "identity": ["showsys"],
            "call_home": [],
            "dns": ["shownet"],
            "ntp": ["shownet"],
            "smtp": [],
            "data_protection": ["showrcopy"],
        }
    if key == _DS8884_PROFILE:
        return {
            "identity": [],
            "call_home": ["dscli showsp"],
            "dns": ["dscli lsnetworkport"],
            "ntp": [],
            "smtp": [],
            "data_protection": [],
        }
    return empty


def wrap_inventory_commands_for_card(
    commands: dict[str, list[str]],
    *,
    dscli_path: str = "",
    dscli_hmc: str = "",
    username: str = "",
    password: str = "",
) -> dict[str, list[str]]:
    from launchpad.dscli_wrap import wrap_dscli_command_list

    return {
        topic: wrap_dscli_command_list(
            list(cmds),
            dscli_path=dscli_path,
            hmc_host=dscli_hmc,
            username=username,
            password=password,
        )
        for topic, cmds in commands.items()
    }


def parse_svc_lssystem_identity(output: str) -> tuple[str, str]:
    text = str(output or "")
    model = ""
    serial = ""
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped or ":" not in stripped:
            continue
        key, _, value = stripped.partition(":")
        token = key.strip().lower()
        val = value.strip()
        if token == "product_name" and val:
            model = val
        elif token == "id" and val:
            serial = val
    return model, serial


_VOLUME_PROTECTION_YES = frozenset({"enabled", "on", "yes", "true"})
_VOLUME_PROTECTION_NO = frozenset({"disabled", "off", "no", "false"})


def parse_svc_lssystem_volume_protection(output: str) -> tuple[str, str, str]:
    text = str(output or "")
    found = ""
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped or ":" not in stripped:
            continue
        key, _, value = stripped.partition(":")
        if key.strip().lower() != "volume_protection":
            continue
        found = value.strip()
        break
    if not found:
        return "unknown", "", "volume protection not reported"
    token = found.lower()
    if token in _VOLUME_PROTECTION_YES:
        return "yes", "configured", found
    if token in _VOLUME_PROTECTION_NO:
        return "no", "empty", found
    return "unknown", "", "volume protection not reported"


def parse_svc_lsemailserver(output: str) -> tuple[str, str, str]:
    text = str(output or "").strip()
    if not text:
        return "unknown", "", "empty email server output"
    headers, rows = _parse_colon_table(text)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not headers and not rows:
        if len(lines) == 1 and ":" in lines[0] and lines[0].lower().startswith("id"):
            return "no", "empty", "No IP — Not configured"
        return "unknown", "", "unrecognized email server output"
    if not rows:
        return "no", "empty", "No IP — Not configured"
    ip_idx = _header_index(headers, "IP_address", "ip_address", "IP", "ip")
    ips: list[str] = []
    for row in rows:
        ip = _cell(row, ip_idx)
        if not ip and len(row) >= 3:
            ip = str(row[2] or "").strip()
        if ip and re.match(r"^\d{1,3}(?:\.\d{1,3}){3}$", ip):
            ips.append(ip)
    if not ips:
        return "no", "empty", "No IP — Not configured"
    return "yes", "configured", ", ".join(ips)


def parse_svc_lsrcrelationship(output: str) -> tuple[str, str, str]:
    text = str(output or "").strip()
    if not text:
        return "unknown", "", "empty remote-copy relationship output"
    headers, rows = _parse_colon_table(text)
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not headers and not rows:
        if len(lines) == 1 and ":" in lines[0] and lines[0].lower().startswith("id"):
            return "no", "empty", "no remote-copy relationships"
        return "unknown", "", "unrecognized remote-copy relationship output"
    if not rows:
        return "no", "empty", "no remote-copy relationships"
    name_idx = _header_index(headers, "name")
    names: list[str] = []
    for row in rows:
        name = _cell(row, name_idx)
        if not name and len(row) > 1:
            name = str(row[1] or "").strip()
        if name:
            names.append(name)
    details = ", ".join(names) if names else "configured"
    return "yes", "configured", details


def parse_hpe_showrcopy_protection(output: str) -> tuple[str, str, str]:
    text = str(output or "").strip()
    if not text:
        return "unknown", "", "empty showrcopy output"
    if _HPE_RCOPY_NOT_CONFIGURED_RE.search(text):
        return "no", "empty", "not configured"
    targets: list[str] = []
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if _HPE_RCOPY_TARGET_RE.match(stripped):
            _, _, rest = stripped.partition(":")
            if not rest.strip():
                _, _, rest = stripped.partition("=")
            name = rest.strip()
            if name:
                targets.append(name)
            continue
        if stripped.lower().startswith(("group", "target", "system", "remote")):
            continue
        tokens = stripped.split()
        if len(tokens) >= 2 and not tokens[0].isdigit():
            continue
        if len(tokens) >= 2:
            candidate = tokens[1] if tokens[0].isdigit() else tokens[0]
            if candidate and candidate.lower() not in {"name", "status", "mode", "role"}:
                targets.append(candidate)
    if targets:
        return "yes", "configured", ", ".join(dict.fromkeys(targets))
    lowered = text.lower()
    if "remote copy" in lowered or "rcopy" in lowered:
        if re.search(r"\b0\b|\bnone\b|\bno\b", lowered):
            return "no", "empty", "not configured"
    non_empty = [ln for ln in text.splitlines() if ln.strip()]
    if len(non_empty) <= 2:
        return "no", "empty", "not configured"
    return "unknown", "", "unrecognized showrcopy output"


def format_phone_home_cell(*, configured: str, details: str, vendor: str) -> str:
    cfg = str(configured or "").strip().lower()
    if cfg == "n/a":
        return "n/a"
    if cfg == "unknown":
        return "unknown"
    if cfg == "no":
        return "No — Not configured"
    if cfg == "yes":
        short = str(details or "").strip()
        if short:
            return f"Yes — {short}"
        return f"Yes — {vendor}"
    return str(configured or "")


def format_yes_no_cell(
    *,
    configured: str,
    details: str = "",
    na_label: str = "n/a",
) -> str:
    cfg = str(configured or "").strip().lower()
    if cfg == "n/a":
        return na_label
    if cfg == "unknown":
        return "unknown"
    if cfg == "no":
        short = str(details or "").strip()
        return short if short else "No — Not configured"
    if cfg == "yes":
        short = str(details or "").strip()
        return short if short else "Yes"
    return str(configured or "")


def format_smtp_cell(*, configured: str, details: str) -> str:
    cfg = str(configured or "").strip().lower()
    if cfg in {"n/a", "unknown"}:
        return cfg
    if cfg == "no":
        short = str(details or "").strip()
        return short if short else "No IP — Not configured"
    if cfg == "yes":
        return str(details or "").strip()
    return str(configured or "")


def health_issue_messages(health_issues: list) -> list[str]:
    messages: list[str] = []
    for item in health_issues or []:
        if isinstance(item, dict):
            msg = str(item.get("message") or "").strip()
            if msg:
                messages.append(msg)
        else:
            text = str(item or "").strip()
            if text:
                messages.append(text)
    return messages


def build_issues_notes(
    *,
    phone_configured: str,
    data_protection_configured: str,
    smtp_configured: str,
    dns_configured: str,
    ntp_configured: str,
    health_issues: list,
    extra_errors: list[str],
    volume_protection_configured: str = "n/a",
) -> str:
    notes: list[str] = []
    if str(phone_configured or "").strip().lower() == "no":
        notes.append("Phone Home not configured")
    if str(data_protection_configured or "").strip().lower() == "no":
        notes.append("Data Protection not configured")
    if str(volume_protection_configured or "").strip().lower() == "no":
        notes.append("Volume Protection not configured")
    if str(smtp_configured or "").strip().lower() == "no":
        notes.append("SMTP not configured")
    if str(dns_configured or "").strip().lower() == "no":
        notes.append("DNS not configured")
    if str(ntp_configured or "").strip().lower() == "no":
        notes.append("NTP not configured")
    notes.extend(health_issue_messages(health_issues))
    for err in extra_errors or []:
        text = str(err or "").strip()
        if text:
            notes.append(text)
    return "; ".join(notes)


def split_inventory_issue_fields(
    *,
    phone_configured: str,
    data_protection_configured: str,
    smtp_configured: str,
    dns_configured: str,
    ntp_configured: str,
    health_issues: list,
    extra_errors: list[str],
    card_id: int | str | None = None,
    alert_state: dict | None = None,
    now: float = 0.0,
    volume_protection_configured: str = "n/a",
) -> dict[str, str]:
    items = list(health_issues or [])
    extras = list(extra_errors or [])
    issues = build_issues_notes(
        phone_configured=phone_configured,
        data_protection_configured=data_protection_configured,
        smtp_configured=smtp_configured,
        dns_configured=dns_configured,
        ntp_configured=ntp_configured,
        health_issues=items,
        extra_errors=extras,
        volume_protection_configured=volume_protection_configured,
    )
    config_notes = build_issues_notes(
        phone_configured=phone_configured,
        data_protection_configured=data_protection_configured,
        smtp_configured=smtp_configured,
        dns_configured=dns_configured,
        ntp_configured=ntp_configured,
        health_issues=[],
        extra_errors=extras,
        volume_protection_configured=volume_protection_configured,
    )
    recent_health: list = []
    older_health: list = []
    for issue in items:
        if not isinstance(issue, dict):
            continue
        visible = True
        if alert_state is not None and card_id is not None:
            fp = issue_fingerprint_for_issue(card_id, issue)
            visible = issue_is_visible(alert_state, fp, now=now)
        if visible:
            recent_health.append(issue)
        else:
            older_health.append(issue)
    recent_parts = [part for part in (config_notes, "; ".join(health_issue_messages(recent_health))) if part]
    return {
        "issues": issues,
        "issues_recent": "; ".join(recent_parts),
        "issues_older": "; ".join(health_issue_messages(older_health)),
    }


def row_has_issues(row: dict) -> bool:
    return bool(str(row.get("issues") or "").strip())


BLANK_SITE_LABEL = "(no site)"
_SITE_UNKNOWN_FIELDS: tuple[str, ...] = ("phone_home", "data_protection", "smtp", "volume_protection")


def group_inventory_rows_by_site(rows: list[dict] | None) -> list[tuple[str, list[dict]]]:
    buckets: dict[str, list[dict]] = {}
    for row in rows or []:
        label = str(row.get("site") or "").strip() or BLANK_SITE_LABEL
        buckets.setdefault(label, []).append(row)
    return [(name, buckets[name]) for name in sorted(buckets, key=lambda item: item.lower())]


def _row_has_unknown(row: dict) -> bool:
    for field in _SITE_UNKNOWN_FIELDS:
        raw = str(row.get(field) or "").strip().lower()
        if field == "volume_protection" and not raw:
            return True
        if raw == "unknown":
            return True
    return False


def site_status(rows: list[dict] | None) -> str:
    items = list(rows or [])
    if any(row_has_issues(row) for row in items):
        return "red"
    if any(_row_has_unknown(row) for row in items):
        return "orange"
    return "green"


def inventory_totals(rows: list[dict]) -> dict:
    total = len(rows or [])
    with_issues = sum(1 for row in rows or [] if row_has_issues(row))
    return {"total_devices": total, "devices_with_issues": with_issues}


def build_inventory_row(
    *,
    site: str,
    host: str,
    ip: str,
    model: str,
    serial: str,
    location: str,
    vendor: str,
    profile: str,
    card_id: int | None,
    phone: tuple[str, str, str],
    data_protection: tuple[str, str, str],
    smtp: tuple[str, str, str],
    dns: tuple[str, str, str],
    ntp: tuple[str, str, str],
    health_issues: list | None = None,
    extra_errors: list[str] | None = None,
    alert_state: dict | None = None,
    now: float = 0.0,
    volume_protection: tuple[str, str, str] = ("n/a", "", ""),
) -> dict:
    phone_cfg, _phone_status, phone_details = phone
    dp_cfg, _dp_status, dp_details = data_protection
    smtp_cfg, _smtp_status, smtp_details = smtp
    dns_cfg, _dns_status, _dns_details = dns
    ntp_cfg, _ntp_status, _ntp_details = ntp
    vp_cfg, _vp_status, _vp_details = volume_protection
    split = split_inventory_issue_fields(
        phone_configured=phone_cfg,
        data_protection_configured=dp_cfg,
        smtp_configured=smtp_cfg,
        dns_configured=dns_cfg,
        ntp_configured=ntp_cfg,
        health_issues=health_issues or [],
        extra_errors=extra_errors or [],
        card_id=card_id,
        alert_state=alert_state,
        now=now,
        volume_protection_configured=vp_cfg,
    )
    row: dict[str, Any] = {
        "site": site,
        "host": host,
        "ip": ip,
        "model": model,
        "serial": serial,
        "location": location,
        "phone_home": format_phone_home_cell(
            configured=phone_cfg,
            details=phone_details,
            vendor=vendor,
        ),
        "data_protection": format_yes_no_cell(
            configured=dp_cfg,
            details=dp_details,
        ),
        "volume_protection": format_yes_no_cell(
            configured=vp_cfg,
        ),
        "smtp": format_smtp_cell(configured=smtp_cfg, details=smtp_details),
        "issues": split["issues"],
        "issues_recent": split["issues_recent"],
        "issues_older": split["issues_older"],
        "profile": profile,
        "vendor": vendor,
    }
    if card_id is not None:
        row["card_id"] = int(card_id)
    return row


_INVENTORY_HEADERS: tuple[str, ...] = (
    "Site",
    "Host",
    "IP Address",
    "Model",
    "Serial Number (SN)",
    "Location",
    "Phone Home",
    "Data Protection",
    "Volume Protection",
    "SMTP IP(s)",
    "Issues / Notes",
)

_INVENTORY_FIELDS: tuple[str, ...] = (
    "site",
    "host",
    "ip",
    "model",
    "serial",
    "location",
    "phone_home",
    "data_protection",
    "volume_protection",
    "smtp",
    "issues",
)

_ISSUES_SUMMARY_HEADERS: tuple[str, ...] = (
    "Site",
    "Host",
    "IP Address",
    "Model",
    "Serial Number (SN)",
    "Issues / Notes",
)

_ISSUES_SUMMARY_FIELDS: tuple[str, ...] = (
    "site",
    "host",
    "ip",
    "model",
    "serial",
    "issues",
)

_ISSUE_ROW_FILL = PatternFill(
    start_color="FFCDD2",
    end_color="FFCDD2",
    fill_type="solid",
)


def export_storage_inventory_xlsx(
    rows: list[dict],
    *,
    generated_at: str | None = None,
) -> bytes:
    """Return a workbook with Inventory and Issues Summary sheets."""
    totals = inventory_totals(rows)
    generated = str(generated_at or "").strip()
    meta = (
        f"Generated: {generated} | Total Devices: {totals['total_devices']} | "
        f"Devices with Issues: {totals['devices_with_issues']}"
    )

    workbook = Workbook()
    inventory = workbook.active
    inventory.title = "Inventory"
    inventory["A1"] = meta

    for column, title in enumerate(_INVENTORY_HEADERS, start=1):
        inventory.cell(row=2, column=column, value=title)

    for row_index, row in enumerate(rows or [], start=3):
        highlight = row_has_issues(row)
        for column, field in enumerate(_INVENTORY_FIELDS, start=1):
            cell = inventory.cell(row=row_index, column=column, value=row.get(field, ""))
            if highlight:
                cell.fill = _ISSUE_ROW_FILL

    summary = workbook.create_sheet("Issues Summary")
    for column, title in enumerate(_ISSUES_SUMMARY_HEADERS, start=1):
        summary.cell(row=1, column=column, value=title)

    issue_rows = [row for row in rows or [] if row_has_issues(row)]
    for row_index, row in enumerate(issue_rows, start=2):
        for column, field in enumerate(_ISSUES_SUMMARY_FIELDS, start=1):
            summary.cell(row=row_index, column=column, value=row.get(field, ""))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def empty_storage_inventory_progress() -> dict[str, Any]:
    return {"running": False, "done": 0, "total": 0, "current": ""}


class StorageInventoryProgress:
    def __init__(self) -> None:
        self._lock = Lock()
        self._data = empty_storage_inventory_progress()

    def begin(self, total: int) -> None:
        with self._lock:
            self._data = {
                "running": True,
                "done": 0,
                "total": int(total),
                "current": "",
            }

    def start_card(self, name: str) -> None:
        with self._lock:
            self._data["current"] = str(name or "")

    def finish_card(self) -> None:
        with self._lock:
            self._data["done"] = int(self._data.get("done") or 0) + 1

    def end(self) -> None:
        with self._lock:
            self._data["running"] = False
            self._data["current"] = ""

    def snapshot(self) -> dict[str, Any]:
        with self._lock:
            return {
                "running": bool(self._data["running"]),
                "done": int(self._data.get("done") or 0),
                "total": int(self._data.get("total") or 0),
                "current": str(self._data.get("current") or ""),
            }


__all__ = [
    "BLANK_SITE_LABEL",
    "INVENTORY_COLUMNS",
    "StorageInventoryProgress",
    "build_inventory_row",
    "build_issues_notes",
    "empty_storage_inventory_progress",
    "export_storage_inventory_xlsx",
    "format_phone_home_cell",
    "format_smtp_cell",
    "format_yes_no_cell",
    "group_inventory_rows_by_site",
    "health_issue_messages",
    "hpe_call_home_na_row",
    "inventory_commands_for_profile",
    "wrap_inventory_commands_for_card",
    "inventory_totals",
    "is_hpe_inventory_profile",
    "is_storage_inventory_eligible",
    "is_storage_inventory_profile",
    "is_system_connectivity_eligible",
    "parse_ds_networkport_dns",
    "parse_ds_showsp_call_home",
    "parse_hpe_showrcopy_protection",
    "parse_hpe_shownet_dns_ntp",
    "parse_svc_call_home",
    "parse_svc_dns",
    "parse_svc_lsemailserver",
    "parse_svc_lsrcrelationship",
    "parse_svc_lssystem_identity",
    "parse_svc_lssystem_volume_protection",
    "parse_svc_ntp_from_lssystem",
    "row_has_issues",
    "site_status",
    "split_inventory_issue_fields",
]
